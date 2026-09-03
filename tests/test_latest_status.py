"""最新ステータス Excel の生成を検証する。

管理表（Excel）と履歴（CSV）は tmp_path に本物を作り、`write_latest_status()` の
引数へ直接渡す（`_paths` の共有定数を monkeypatch すると `service.py` 側のローカル
束縛と食い違う経路が残るため、明示引数で閉じたテストにする）。
"""

import time
from pathlib import Path

from openpyxl import load_workbook

import comken.services.salesforce_downloader._paths as _paths_module
from comken.constants import Color
from comken.core.table.model import Table
from comken.services.salesforce_downloader.history import HistoryRow, record
from comken.services.salesforce_downloader.latest_status import write_latest_status
from comken.services.salesforce_downloader.master import ReportEntry
from comken.toolbox.excel import Excel

URL_A = "https://example--sandbox.sandbox.my.salesforce.com/lightning/r/Report/00O5g00000ABCDE/view"
URL_B = "https://example--sandbox.sandbox.my.salesforce.com/lightning/r/Report/00O5g00000FGHIJ/view"

HEADERS = [
    "ID",
    "グループ名",
    "担当者",
    "概要",
    "Salesforce URL",
    "保存先",
    "有効",
    "備考",
]


def make_master(path: Path, rows: list[list]) -> Path:
    """管理表（Excel）を作る。"""
    table_rows = [dict(zip(HEADERS, row, strict=True)) for row in rows]
    with Excel(path) as book:
        book.create_data_sheet("管理表").create_table("管理表", Table(HEADERS, table_rows))
    return path


def _entry(folder: Path, *, key: str, summary: str, url: str) -> ReportEntry:
    return ReportEntry(
        key=key,
        group_name="営業事務グループ",
        assignee="山田",
        summary=summary,
        url=url,
        folder=folder,
        enabled=True,
        allow_empty=False,
        note="",
    )


def _latest_rows(output: Path) -> list[dict[str, str]]:
    """生成された Excel を ``dict`` のリストで返す（PY_ 接頭辞のデータシート）。

    openpyxl は空セルを ``None`` として返すので、空文字に揃える。空セルが
    ``""`` のまま返るとテストで ``None`` との比較になって分かりにくいため、
    ここでは ``None`` を一律 ``""`` に正規化する。
    """
    book = load_workbook(output)
    sheet = book["PY_最新ステータス"]
    iterator = sheet.iter_rows(values_only=True)
    headers = list(next(iterator))
    return [
        {
            header: ("" if value is None else str(value))
            for header, value in zip(headers, row, strict=True)
        }
        for row in iterator
    ]


class TestWriteLatestStatus:
    """管理表 × 履歴 → 最新ステータス Excel。"""

    def test_uses_latest_history_row_per_key(self, tmp_path):
        """同じ管理番号に複数の履歴があれば、実行日時が最大の行を採用する。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "レポート管理表.xlsx",
            [
                [
                    "1001",
                    "営業事務グループ",
                    "山田",
                    "顧客一覧",
                    URL_A,
                    str(folder),
                    "○",
                    "",
                ],
                [
                    "1002",
                    "経理グループ",
                    "佐藤",
                    "売上実績",
                    URL_B,
                    str(folder),
                    "○",
                    "",
                ],
            ],
        )
        history_path = tmp_path / "ダウンロード履歴.csv"

        entry1 = _entry(folder, key="1001", summary="顧客一覧", url=URL_A)
        entry2 = _entry(folder, key="1002", summary="売上実績", url=URL_B)
        # 1001: 古い失敗 → 新しい成功。``now()`` の精度が秒なので 1 秒待って
        # 確実に新しい行として認識させる（文字列比較なので同秒だと古い行が残る）
        record(
            history_path,
            entry=entry1,
            project="P",
            row=HistoryRow(
                succeeded=False,
                fetched_from_salesforce=False,
                saved_to_file=None,
                cause="プログラム",
                error_code="TypeError",
                error="boom",
            ),
        )
        time.sleep(1)
        record(
            history_path,
            entry=entry1,
            project="P",
            row=HistoryRow(
                succeeded=True,
                fetched_from_salesforce=True,
                saved_to_file=True,
                file_name="1001.csv",
            ),
        )
        # 1002: 古い成功 → 新しい失敗
        record(
            history_path,
            entry=entry2,
            project="P",
            row=HistoryRow(
                succeeded=True,
                fetched_from_salesforce=True,
                saved_to_file=True,
                file_name="1002.csv",
            ),
        )
        time.sleep(1)
        record(
            history_path,
            entry=entry2,
            project="P",
            row=HistoryRow(
                succeeded=False,
                fetched_from_salesforce=True,
                saved_to_file=None,
                cause="データなし",
                error_code="EmptyReportError",
                error="0行",
            ),
        )

        output = tmp_path / "最新ステータス.xlsx"
        write_latest_status(master, history_path, output)

        rows = _latest_rows(output)
        by_key = {row["管理番号"]: row for row in rows}
        assert by_key["1001"]["成否"] == "成功"
        assert by_key["1001"]["原因区分"] == ""
        assert by_key["1002"]["成否"] == "失敗"
        assert by_key["1002"]["原因区分"] == "データなし"
        assert "0行" in by_key["1002"]["エラー内容"]

    def test_missing_history_is_marked_not_run(self, tmp_path):
        """履歴が無い管理番号は「未実行」として 1 行出す。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "レポート管理表.xlsx",
            [
                [
                    "1001",
                    "営業事務グループ",
                    "山田",
                    "顧客一覧",
                    URL_A,
                    str(folder),
                    "○",
                    "",
                ],
                [
                    "1002",
                    "経理グループ",
                    "佐藤",
                    "売上実績",
                    URL_B,
                    str(folder),
                    "○",
                    "",
                ],
            ],
        )
        history_path = tmp_path / "ダウンロード履歴.csv"
        # 1001 だけ履歴がある
        record(
            history_path,
            entry=_entry(folder, key="1001", summary="顧客一覧", url=URL_A),
            project="P",
            row=HistoryRow(
                succeeded=True,
                fetched_from_salesforce=True,
                saved_to_file=True,
                file_name="1001.csv",
            ),
        )

        output = tmp_path / "最新ステータス.xlsx"
        write_latest_status(master, history_path, output)
        rows = _latest_rows(output)
        by_key = {row["管理番号"]: row for row in rows}
        assert by_key["1001"]["成否"] == "成功"
        assert by_key["1002"]["成否"] == "未実行"
        assert by_key["1002"]["最新実行日時"] in (None, "")

    def test_failure_rows_are_filled_with_pink(self, tmp_path):
        """失敗行のセルは `Color.PINK` で塗りつぶされる。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "レポート管理表.xlsx",
            [
                [
                    "1001",
                    "営業事務グループ",
                    "山田",
                    "顧客一覧",
                    URL_A,
                    str(folder),
                    "○",
                    "",
                ],
                [
                    "1002",
                    "経理グループ",
                    "佐藤",
                    "売上実績",
                    URL_B,
                    str(folder),
                    "○",
                    "",
                ],
            ],
        )
        history_path = tmp_path / "ダウンロード履歴.csv"
        record(
            history_path,
            entry=_entry(folder, key="1001", summary="顧客一覧", url=URL_A),
            project="P",
            row=HistoryRow(
                succeeded=True,
                fetched_from_salesforce=True,
                saved_to_file=True,
                file_name="1001.csv",
            ),
        )
        record(
            history_path,
            entry=_entry(folder, key="1002", summary="売上実績", url=URL_B),
            project="P",
            row=HistoryRow(
                succeeded=False,
                fetched_from_salesforce=True,
                saved_to_file=None,
                cause="データなし",
                error_code="EmptyReportError",
                error="0行",
            ),
        )

        output = tmp_path / "最新ステータス.xlsx"
        write_latest_status(master, history_path, output)

        book = load_workbook(output)
        sheet = book["PY_最新ステータス"]
        # 2 行目（1001: 成功）と 3 行目（1002: 失敗）の塗りつぶしを比較
        success_fill = sheet.cell(row=2, column=4).fill
        failure_fill = sheet.cell(row=3, column=4).fill
        assert _rgb(success_fill) != f"00{Color.PINK}"
        assert _rgb(failure_fill) == f"00{Color.PINK}"
        # 行の全セルが同じ色で塗られていること
        for column in range(1, 7):
            assert _rgb(sheet.cell(row=3, column=column).fill) == f"00{Color.PINK}"

    def test_default_paths_follow_paths_module(self, tmp_path, monkeypatch):
        """引数を省略すると ``_paths.LATEST_STATUS_PATH`` へ書き出す。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "レポート管理表.xlsx",
            [
                [
                    "1001",
                    "営業事務グループ",
                    "山田",
                    "顧客一覧",
                    URL_A,
                    str(folder),
                    "○",
                    "",
                ]
            ],
        )
        history_path = tmp_path / "ダウンロード履歴.csv"
        record(
            history_path,
            entry=_entry(folder, key="1001", summary="顧客一覧", url=URL_A),
            project="P",
            row=HistoryRow(
                succeeded=True,
                fetched_from_salesforce=True,
                saved_to_file=True,
                file_name="1001.csv",
            ),
        )
        default_output = tmp_path / "default.xlsx"
        monkeypatch.setattr(_paths_module, "MASTER_PATH", master)
        monkeypatch.setattr(_paths_module, "HISTORY_PATH", history_path)
        monkeypatch.setattr(_paths_module, "LATEST_STATUS_PATH", default_output)

        write_latest_status()

        assert default_output.is_file()
        rows = _latest_rows(default_output)
        assert rows[0]["管理番号"] == "1001"
        assert rows[0]["成否"] == "成功"


def _rgb(fill) -> str:
    """openpyxl の ``PatternFill`` から、比較しやすい RGB 文字列を返す。"""
    fg = fill.fgColor
    return str(fg.rgb) if fg is not None else ""
