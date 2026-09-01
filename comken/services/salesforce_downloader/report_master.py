r"""comken/services/salesforce_downloader/report_master.py — Excel の表を設定として読む。

**非エンジニアが Excel で編集する一覧**を、型付きの行として読み込む。
「どのレポートを取るか」「どのファイルをコピーするか」のように**行が増えていく設定**は、
config.ini より表のほうが扱いやすい（並べ替え・フィルタ・コピーができる）。

使い方は、1行につき1列を宣言するだけ。

    from dataclasses import dataclass
    from pathlib import Path

    from comken.services.salesforce_downloader.report_master import MasterRow, column


    @dataclass(frozen=True, kw_only=True)
    class Report(MasterRow):
        \"\"\"レポート管理表の1行。\"\"\"

        SHEET_NAME = "管理表"
        PATH = Path(r"\\server\share\レポート管理表.xlsx")

        key: int = column("ID", unique=True, help="社内で決める管理番号")
        summary: str = column("概要", help="人が読んで分かる説明")
        folder: Path = column("保存先", help="落としたファイルを置くフォルダ")
        enabled: bool = column("有効", default=True, help="使わなくなったら「無効」")

    Report.create_template(path)   # 記入例と「記入方法」シート付きの雛形を作る
    for report in Report.load():   # 読む（型変換・検証込み）
        print(report.summary, report.folder)

**Python の名前は英語、Excel の見出しは日本語**にできる。`column()` の第1引数が
Excel の見出しで、スペースを含む見出し（`Salesforce URL`）も扱える。

**`kw_only=True` を付ける。** 付けないと「既定値のある列の後ろに、既定値のない列を
書けない」という dataclass の制約に引っかかり、**列を足すときに並び順を気にする**ことに
なる。Excel の列は増えるものなので、どこにでも書けるようにしておく
（Excel 側の並び順は元から自由。見出しの名前で引くため）。

**空欄をどう扱うかは、既定値の有無で決まる。**

| 宣言 | セルが空欄 | 列（見出し）ごと無い |
|---|---|---|
| `column("備考", default="")` | 既定値を使う | 既定値を使う |
| `column("担当")` | **エラー** | **エラー** |

**既定値は「空欄でよい」という宣言**として使う。意味が反転する列（有効/無効のような）に
既定値を付けてはいけない——**書き忘れが「有効」になり、意図と逆の結果になる**。
そういう列は既定値を持たせず、必ず書かせる。

既定値のある列は、**見出しごと無くても読める**。列を1つ足した瞬間に既存の管理表が
すべて読めなくなると業務が止まるため（共有サーバーを更新すると全プロジェクトへ伝播する）。
値が要る列を足したときは、管理表に足すまで止まる。

型は注釈から決まる（`int` / `str` / `bool` / `Path`）。**列の定義はここ1か所**なので、
読み込む型と Excel の見出しがズレることがない。
"""

import dataclasses
import datetime as dt
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, ClassVar, Self

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from comken.constants import Color
from comken.core.data import is_true_word
from comken.core.table.model import Table as CoreTable
from comken.core.timer import measure
from comken.exceptions import (
    ExcelApplicationNotAvailableError,
    MasterColumnNotFoundError,
    MasterDuplicateValueError,
    MasterRowValueError,
    MasterSheetNotDefinedError,
)
from comken.toolbox.excel import Excel

# フィールドの metadata に入れるときのキー
_SPEC_KEY = "comken_master_column"

# 「有効」列などで真として扱う値（小文字で比較する）。
# 英語の "true" 表記は config.ini と同じ判定にするため、ここには含めず
# is_true_word() で共通判定する（_to_bool 側で合わせて見る）
_TRUE_WORDS = ("有効", "○", "o", "yes", "1", "on", "はい")

# 見出し行を除いた1行目が Excel の何行目か（見出しが1行目のため）
_FIRST_DATA_ROW = 2

# 雛形の全セルに付けるフォント名。Windows 標準ではないため、入っていない PC では
# Excel が代替フォントを使う（コメント作成時点 2026-08）。これは予想される動作で、
# 意図した見た目にならずとも雛形自体は問題なく使える
_TEMPLATE_FONT_NAME = "Noto Sans JP"

# ドロップダウンを適用するデータ行数。**行を足したときにも効くよう**、雛形時点での
# 想定行数より十分大きく取っておく（テンプレから1000行以上増える運用は基本無い）
_DATA_VALIDATION_ROWS = 1000

# 記入例の行に付ける背景色。本物のデータと見分けが付くよう、薄い灰色系で
# `Color.LIGHT_GRAY` を使う（強く出すと「エラー行」に見えるため）
_EXAMPLE_FILL_COLOR = Color.LIGHT_GRAY
# openpyxl の PatternFill を作って使い回す（毎回作り直すのを避ける）
_EXAMPLE_FILL = PatternFill(fill_type="solid", fgColor=_EXAMPLE_FILL_COLOR)


@dataclass(frozen=True)
class ColumnSpec:
    """1つの列の決まりごと。

    Attributes:
        header: Excel の見出し（そのまま1行目に書かれる）。
        unique: True なら、同じ値が2つ以上あるとエラーにする。
        choices: 書ける値を限る場合の一覧。
        help: 「記入方法」シートとエラーメッセージに使う説明。
    """

    header: str
    unique: bool = False
    choices: tuple[str, ...] | None = None
    help: str = ""


def column(
    header: str,
    *,
    unique: bool = False,
    choices: tuple[str, ...] | None = None,
    default: Any = dataclasses.MISSING,
    help: str = "",  # 「記入方法」シートに出る説明。help 以外の名前だと意味が伝わらない
) -> Any:
    """列を1つ宣言する（dataclass のフィールドに使う）。

    Args:
        header: **Excel の見出し**。Python の名前と違ってよく、スペースも使える。
        unique: 同じ値が2つ以上あればエラーにする（管理番号など）。
        choices: 書ける値を限る（`("定期", "個別")` など）。
        default: 空欄のときの値。省略すると**空欄はエラー**になる。
        help: 何を書く列かの説明。「記入方法」シートとエラーメッセージに出る。
    """
    spec = ColumnSpec(header=header, unique=unique, choices=choices, help=help)
    metadata = {_SPEC_KEY: spec}
    if default is dataclasses.MISSING:
        return field(metadata=metadata)
    return field(default=default, metadata=metadata)


class MasterRow:
    """Excel の表の1行。`@dataclass(frozen=True, kw_only=True)` と一緒に継承して使う。

    クラス変数:
        SHEET_NAME: 読み書きするシート名。
        PATH: 既定のファイル。指定すると `load()` を引数なしで呼べる。
        GUIDE_INTRO: 「記入方法」シートの冒頭に出す説明文。設定した管理表だけを
            向いた案内（Salesforce なら Salesforce の）をここに置くと、非エンジニアが
            表を開いた瞬間にこの表で何ができるかが分かる。空のままでも雛形は作れる。
    """

    SHEET_NAME: ClassVar[str] = "管理表"
    PATH: ClassVar[Path | None] = None
    GUIDE_INTRO: ClassVar[str] = ""

    # ── 読む ────────────────────────────────────────────────────────────────
    @classmethod
    @measure
    def load(cls, path: str | Path | None = None) -> list[Self]:
        """表を読んで、行のリストを返す。

        Args:
            path: Excel のパス。省略時はクラス変数 `PATH`。

        Returns:
            宣言した順のまま、1行ずつのインスタンス。

        Raises:
            MasterSheetNotDefinedError: path も PATH も無い場合。
            MasterColumnNotFoundError: 宣言した見出しが表に無い場合。
            MasterRowValueError: 値が型・選択肢に合わない場合。
            MasterDuplicateValueError: unique の列に同じ値がある場合。
        """
        source = Path(path) if path is not None else cls.PATH
        if source is None:
            raise MasterSheetNotDefinedError(cls.__name__)

        # 共有関数 ``read_raw_rows`` に生 dict 化を任せ、ここでは型変換・検証に
        # 集中する（同じロジックを ``load_schedule`` 側でも使う）
        raw_rows = read_raw_rows(source, cls.SHEET_NAME)

        rows: list[Self] = []
        seen: dict[str, set] = {}
        for offset, raw in enumerate(raw_rows):
            if _is_blank(raw):
                continue  # 表の下に残った空行は読み飛ばす
            row_number = offset + _FIRST_DATA_ROW
            _require_headers(cls, raw, source)
            rows.append(cls._build(raw, row_number, seen, source))
        return rows

    @classmethod
    def _build(cls, raw: dict, row_number: int, seen: dict[str, set], source: Path) -> Self:
        """1行ぶんの生の値を、型付きのインスタンスにする。"""
        values = {}
        for name, spec, value_type in cls._columns():
            value = _convert(raw.get(spec.header), value_type, spec, row_number, cls)
            if value is _EMPTY:
                default = _default_of(cls, name)
                if default is dataclasses.MISSING:
                    raise MasterRowValueError(
                        row_number,
                        spec.header,
                        "",
                        "空のままにできません。",
                    )
                value = default
            if spec.unique:
                if value in seen.setdefault(name, set()):
                    raise MasterDuplicateValueError(spec.header, value, source)
                seen[name].add(value)
            values[name] = value
        return cls(**values)

    # ── 雛形を作る ───────────────────────────────────────────────────────────
    @classmethod
    @measure
    def create_template(cls, path: str | Path, examples: list[dict] | None = None) -> Path:
        """記入例と「記入方法」シートが入った雛形を作る。

        **空の表を渡されるより、1行埋まっているほうが何をどう書くか伝わる**ので、
        記入例を入れておく（使う前に消す案内も「記入方法」に書く）。

        **`choices` を宣言した列には、自動でドロップダウン（Excel の入力規則）が付く。**
        ドロップダウン・案内文・エラーメッセージは `column()` の `choices` と `help` から
        組み立てるため、宣言を1か所に保ったまま入力補助が出る。

        **雛形全体のフォントは Noto Sans JP。** 既存のフォント属性（太字など）は、
        フォント名だけ書き換える方式で残す。

        Args:
            path: 作成先（.xlsx）。
            examples: 記入例。{Python の名前: 値} の形で渡す。
        """
        path = Path(path)
        columns = cls._columns()
        headers = [spec.header for _, spec, _ in columns]
        rows = [
            {
                spec.header: _to_cell(example.get(name, ""), spec, value_type)
                for name, spec, value_type in columns
            }
            for example in (examples or [])
        ]

        # 空の雛形でも Excel テーブルを成立させるため、API が要求する見出しだけを
        # 持つ Table を作る。実データが無い場合の仮行は create_table が保持しない。
        template_table = CoreTable(headers, rows)
        with Excel(path) as excel:
            excel.create_data_sheet(cls.SHEET_NAME).create_table(cls.__name__, template_table)

        book = load_workbook(path)
        sheet = book[f"PY_{cls.SHEET_NAME}"]

        # **全セルに雛形用のフォントを当てる。** 既存のフォント属性（太字など）は
        # そのまま使い回し、`name` だけ書き換える（後勝ちで上書きすると太字まで消える）
        cls._apply_template_font(sheet, len(rows))
        # **`choices` がある列にドロップダウンを付ける。** データ行の先頭から
        # 十分な行数ぶんの範囲に適用し、あとから行を足しても効くようにする
        cls._apply_choice_validations(sheet, columns, len(rows))

        _auto_width(sheet)
        sheet.freeze_panes = "A2"

        # 記入例の行に薄い背景色を付ける。**本物のデータと見分けが付く**ように
        # するための目印で、エラー行に見える色は避ける（強すぎる色は業務側が
        # 「何か起きたのか」と不安になるため）
        for offset in range(len(rows)):
            row_number = _FIRST_DATA_ROW + offset
            for col in range(1, len(headers) + 1):
                sheet.cell(row=row_number, column=col).fill = _EXAMPLE_FILL

        cls._write_guide(book)
        book.save(path)
        book.close()
        return path

    @classmethod
    def _apply_template_font(cls, sheet: Worksheet, example_count: int) -> None:
        """雛形（表シート）の全セルに雛形用のフォント名を設定する。

        既存の設定（太字など）は `Font` オブジェクトをそのまま使い回し、`name` だけを
        書き換える。**他の属性（太字・サイズなど）に触らないため、雛形のもともとの
        見出し書式（太字）を崩さない。**
        """
        last_row = max(_FIRST_DATA_ROW + example_count - 1, 1)
        for row in sheet.iter_rows(
            min_row=1, max_row=last_row, min_col=1, max_col=sheet.max_column
        ):
            for cell in row:
                _set_template_font(cell)

    @classmethod
    def _apply_choice_validations(
        cls, ws: Worksheet, columns: list[tuple[str, ColumnSpec, type]], example_count: int
    ) -> None:
        """`choices` を宣言した列に Excel の入力規則（ドロップダウン）を付ける。

        列ごとに `DataValidation` を作り、**データ行2行目から十分な下まで**の範囲に
        適用する（あとから行を足したときにもドロップダウンが効くよう、十分な行数を取る）。
        入力時メッセージとエラーメッセージは `column()` の `help` と `choices` から組み立て、
        宣言を1か所に保つ。
        """
        last_row = _FIRST_DATA_ROW + example_count - 1 + _DATA_VALIDATION_ROWS
        for offset, (name, spec, _) in enumerate(columns, start=1):
            if not spec.choices:
                continue  # `choices` を宣言していない列には付けない
            letter = _column_letter(offset)
            choices_text = "、".join(f"「{choice}」" for choice in spec.choices)
            prompt = f"{spec.help}\n書き方: {choices_text}".strip()
            error = f"『{'』か『'.join(spec.choices)}』のいずれかを入力してください。"
            # showDropDown=False は「ボタンを表示する」指定（Excel の API は逆）。
            # 非エンジニアが見て選択できる必要があるため True（=ボタンを表示しない）
            # にはしない
            validation = DataValidation(
                type="list",
                formula1=f'"{",".join(spec.choices)}"',
                allow_blank=_default_of(cls, name) is not dataclasses.MISSING,
                showDropDown=False,
                showErrorMessage=True,
                errorTitle="書き方が違います",
                error=error,
                showInputMessage=True,
                promptTitle=spec.header,
                prompt=prompt,
            )
            validation.add(f"{letter}{_FIRST_DATA_ROW}:{letter}{last_row}")
            ws.add_data_validation(validation)

    @classmethod
    def _write_guide(cls, book: Workbook) -> None:
        """「記入方法」シートを書く。非エンジニアが1枚で分かるようにする。

        `GUIDE_INTRO` が設定されていれば冒頭（見出しより上）に書き出す。
        `docs/` を読まない編集者への唯一の案内になるため、各管理表に特化した
        説明を置く。改行は同じセル内に表示される。

        書き終わったら雛形用フォントをセルに適用する（表シート側で `_apply_template_font`
        したのと同じフォント名に揃える）。`set_bold` で付けた太字も保持される
        よう、**太字設定 → フォント適用** の順で行う。
        """
        sheet = book.create_sheet("記入方法")
        # 冒頭の説明文。設定が無ければ何も書かない（空の欄を増やさない）
        if cls.GUIDE_INTRO:
            sheet.cell(row=1, column=1, value=cls.GUIDE_INTRO)
            # 改行を含む説明文も1セルなので、空ける行数は変えない
            header_row = 3
        else:
            header_row = 1
        for column, value in enumerate(("列", "何を書くか", "書けない場合"), start=1):
            sheet.cell(row=header_row, column=column, value=value)
        for offset, (name, spec, value_type) in enumerate(cls._columns(), start=1):
            row_number = header_row + offset
            required = _default_of(cls, name) is dataclasses.MISSING
            note = "空欄にできません" if required else "空欄にできます"
            if spec.choices:
                note = f"「{'」か「'.join(spec.choices)}」と書いてください"
            elif value_type is bool:
                note = "「○」か「×」と書いてください"
            for column, value in enumerate((spec.header, spec.help, note), start=1):
                sheet.cell(row=row_number, column=column, value=value)

        last = header_row + len(cls._columns()) + 2
        sheet.cell(row=last, column=1, value="注意")
        sheet.cell(row=last, column=2, value="1行目の見出しは変えないでください（列名で読みます）")
        sheet.cell(row=last + 1, column=2, value="記入例の行は、実際に使うときに消してください")
        for column in range(1, 4):
            sheet.cell(row=header_row, column=column).font = Font(bold=True)
        _auto_width(sheet, max_width=80)
        sheet.freeze_panes = f"A{header_row + 1}"

        # 太字属性を**保ったまま**フォント名を差し替える。先に `set_bold` してから
        # `_set_template_font` を呼ぶことで、既存の `bold=True` を引き継げる
        for row in sheet.iter_rows():
            for cell in row:
                _set_template_font(cell)

    # ── 列の情報 ─────────────────────────────────────────────────────────────
    @classmethod
    def header(cls, name: str) -> str:
        """Python の名前から、Excel の見出しを返す。

        メッセージに出す・Excel を直接触るときに使う。見出しを直接書くと、
        宣言を変えたときにズレるため。

            ReportEntry.header("summary")   # → "概要"

        Raises:
            KeyError: そのフィールドが宣言されていない場合。
        """
        for field_name, spec, _ in cls._columns():
            if field_name == name:
                return spec.header
        raise KeyError(f"{cls.__name__} に {name} という列はありません")

    @classmethod
    def headers(cls) -> list[str]:
        """Excel の見出しを宣言順で返す。"""
        return [spec.header for _, spec, _ in cls._columns()]

    @classmethod
    def _columns(cls) -> list[tuple[str, ColumnSpec, Any]]:
        """(Python の名前, 列の決まり, 型注釈) を宣言順で返す。

        `column()` を使っていないフィールドは、**名前をそのまま見出しにする**
        （見出しと同じ名前が使えるなら、宣言は型注釈だけで済む）。

        Returns:
            ``(name, spec, value_type)`` のリスト。``value_type`` は ``Field.type``
            の値で、dataclass の型注釈そのもの（class か文字列の前方参照）。

        Raises:
            TypeError: ``cls`` に ``@dataclass`` が付いていない場合（継承側で
                付け忘れた）。実行時に ``dataclasses.is_dataclass()`` で守って
                ``__dataclass_fields__`` を直接読む。
        """
        # MasterRow 本体は dataclass ではないが、継承先は必ず @dataclass を付ける
        # （docstring 冒頭のサンプル参照）。is_dataclass() で守ったうえで
        # __dataclass_fields__ を直接読めば ``fields()`` のプロトコル合致問題を
        # 避けつつ、誤用には明示的な例外で気づける
        if not dataclasses.is_dataclass(cls):
            raise TypeError(
                f"{cls.__name__} には @dataclass を付けてください（MasterRow を継承するとき）"
            )
        found: list[tuple[str, ColumnSpec, Any]] = []
        for name, item in cls.__dataclass_fields__.items():
            spec = item.metadata.get(_SPEC_KEY) or ColumnSpec(header=name)
            found.append((name, spec, item.type))
        return found

    def __init_subclass__(cls, **kwargs: Any) -> None:
        # dataclass を付け忘れると fields() が空になり、原因の分からない失敗になる。
        # 継承した時点では判定できないので、ここでは何もせず load() 側で確かめる
        super().__init_subclass__(**kwargs)


class _Empty:
    """空欄を表す番兵（None も空文字も「書かれた値」と区別したいため）。"""


_EMPTY = _Empty()


def _is_blank(raw: dict) -> bool:
    """すべての列が空の行か。"""
    return all(value in (None, "") for value in raw.values())


def read_raw_rows(source: Path, sheet_name: str) -> CoreTable:
    """指定シートを「Excel の生 dict のリスト」として読む（実体は `Table`、dict の
    イテレータとして使える）。

    ``MasterRow`` に紐付かない共通の下読みとして使う。``MasterRow.load()`` も
    ``schedule.load_schedule()`` も、Excel の開き方・未計算の数式判定をここで揃えて
    二重実装を避ける。

    **読み取り専用で開く。** 書き込みモードだと存在しないパスを渡されたときに
    空のブックを新規作成し、その後のテーブル解決が「対象テーブルを一意に決められません」
    で落ちる。管理表は共有サーバー (UNC) に置く運用が前提で、現実の失敗は
    「サーバーが落ちた」「パスが変わった」「権限が無い」のいずれか。
    **業務担当者が画面で見ても原因が分かるよう、ファイル不在は
    ``ExcelFileNotFoundError`` がそのまま上がる経路にする。**

    Raises:
        ExcelFileNotFoundError: ``source`` が存在しない場合。
        SheetNotFoundError: ``sheet_name`` がブックに無い場合。
        ExcelApplicationNotAvailableError: 未計算の数式セルが含まれていた場合。
    """
    with Excel(source, read_only=True) as excel:
        raw_rows = excel.data_sheet(sheet_name).table().read()
    if any(
        isinstance(value, str) and value.startswith("=")
        for raw_row in raw_rows
        for value in raw_row.values()
    ):
        raise ExcelApplicationNotAvailableError(
            source,
            RuntimeError("管理表に未計算の数式があります"),
        )
    return raw_rows


def _require_headers(cls: type[MasterRow], raw: dict, source: Path) -> None:
    """宣言した見出しが表にあるか確かめる。

    **既定値のある列は、見出しごと無くてもよい。** 列を1つ足した瞬間に、既存の管理表が
    すべて読めなくなると業務が止まる（共有サーバーを更新すると全プロジェクトへ伝播するため）。
    既定値を付けて足せば、**既存の管理表はそのまま動き、必要な人だけ Excel に列を足せる**。
    """
    for name, spec, _ in cls._columns():
        if spec.header in raw:
            continue
        if _default_of(cls, name) is not dataclasses.MISSING:
            continue  # 既定値があるので、列が無くても埋められる
        raise MasterColumnNotFoundError(spec.header, sorted(raw), source, cls.SHEET_NAME)


def _default_of(cls: type[MasterRow], name: str) -> Any:
    """そのフィールドの既定値（無ければ MISSING）。"""
    if not dataclasses.is_dataclass(cls):
        # 誤用（MasterRow のまま呼んだ）には読んでも空クラスを返すより、
        # 何がおかしいかを伝える方が有益
        raise TypeError(
            f"{cls.__name__} には @dataclass を付けてください（MasterRow を継承するとき）"
        )
    for name_, item in cls.__dataclass_fields__.items():
        if name_ == name:
            return item.default
    return dataclasses.MISSING


def _convert(value: Any, value_type: Any, spec: ColumnSpec, row: int, cls: type) -> Any:
    """セルの値を、宣言した型へ変換する。空欄なら _EMPTY を返す。"""
    if value is None or (isinstance(value, str) and not value.strip()):
        return _EMPTY

    text = str(value).strip()
    if spec.choices and text not in spec.choices:
        raise MasterRowValueError(
            row, spec.header, value, f"「{'」か「'.join(spec.choices)}」と書いてください。"
        )

    if value_type is bool or value_type == "bool":
        return _to_bool(value)
    if value_type is int or value_type == "int":
        return _to_int(value, text, spec, row)
    if value_type is Path or value_type == "Path":
        return Path(text)
    if value_type is dt.time or value_type == "dt.time":
        return _to_time(value)
    if value_type is str or value_type == "str":
        # **Excel は数値セルを float で返すことがある。** そのまま `str()` すると
        # `1001` が `"1001.0"` になるため、整数値は整数文字列として返す
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return text
    return text


def _to_bool(value: Any) -> bool:
    """セルの値が「有効」を表す語かどうかを判定する。

    ``_convert`` の bool 列変換に使うほか、``schedule.py`` の「有効」「月末指定」
    のような単独の真偽判定にもそのまま流用できるよう、``text`` を内部で
    導出して単一引数にしてある。
    """
    if isinstance(value, bool):
        return value
    text = str(value).strip()
    return is_true_word(text) or text.lower() in _TRUE_WORDS


def _to_time(value: Any) -> dt.time | None:
    """セルの値を時刻へ変換する。空欄は None。

    Excel の時刻セルは ``datetime`` で返ることがあるため、``datetime`` /
    ``time`` / ISO 形式の文字列のいずれも受け付ける。schedule.py の
    「取得時刻」系の列で使う。
    """
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, dt.time):
        return value.replace(second=0, microsecond=0)
    return dt.time.fromisoformat(str(value).strip())


def _to_int(value: Any, text: str, spec: ColumnSpec, row: int) -> int:
    if isinstance(value, bool):
        raise MasterRowValueError(row, spec.header, value, "数字を入れてください。")
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)  # Excel は数値を小数で返すことがある
    if not text.isdigit():
        raise MasterRowValueError(row, spec.header, value, "数字だけで書いてください（例: 1001）。")
    return int(text)


def _to_cell(value: Any, spec: ColumnSpec, value_type: Any) -> Any:
    """記入例をセルに書ける形にする。"""
    if isinstance(value, bool):
        if (value_type is bool or value_type == "bool") and spec.choices:
            return spec.choices[0] if value else spec.choices[1]
        # choices を宣言していない真偽列の既定表記。**案内文（「○」か「×」）と
        # そろえる。** ここだけ「有効/無効」を書くと、記入方法シートの案内と
        # 雛形の中身が食い違い、どちらに従えばよいか分からなくなる。
        # 読む側は `_TRUE_WORDS` が「有効」も受け付けるので、既存の管理表は壊れない
        return "○" if value else "×"
    if isinstance(value, Path):
        return str(value)
    return value


def _column_letter(index: int) -> str:
    """1 -> A, 27 -> AA。"""
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


def _set_template_font(cell: Any) -> None:
    """セルに雛形用のフォント名（Noto Sans JP）を当てる。

    既存のフォント属性（太字・サイズ・色）はそのまま残し、`name` だけを書き換える。
    **`Font(...)` で全項目を指定すると太字などが消える**ため、openpyxl の現プロパティを
    引き継ぐ形で作る。
    """
    existing = cell.font
    cell.font = Font(
        name=_TEMPLATE_FONT_NAME,
        size=existing.size,
        bold=existing.bold,
        italic=existing.italic,
        color=existing.color,
    )


def _auto_width(sheet: Worksheet, *, max_width: int | None = None) -> None:
    """セル内容に合わせて列幅を設定する。"""
    for column_index, cells in enumerate(sheet.iter_cols(), start=1):
        width = max((len(str(cell.value or "")) for cell in cells), default=0) + 2
        if max_width is not None:
            width = min(width, max_width)
        sheet.column_dimensions[get_column_letter(column_index)].width = width
