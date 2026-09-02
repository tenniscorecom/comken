"""comken/services/salesforce_downloader/schedule.py — 取得時刻を判定する。"""

import datetime as dt
import logging
import re
from collections.abc import Mapping
from dataclasses import dataclass
from pathlib import Path
from typing import Self, cast

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from comken.core.holidays import (
    BusinessDayNotFoundError,
    HolidayCalendar,
    default_calendar,
    nth_business_day_of_month,
)
from comken.core.table.model import Table as CoreTable
from comken.exceptions import (
    DownloaderError,
    ExcelFileNotFoundError,
    ScheduleDuplicateKeyError,
    ScheduleIntervalMissingError,
    ScheduleRequiredValueMissingError,
    ScheduleRowValueError,
    ScheduleWeekdayInvalidError,
    SheetNotFoundError,
    UnsupportedScheduleFrequencyError,
)
from comken.services.salesforce_downloader.report_master import (
    _EXAMPLE_FILL,
    _FIRST_DATA_ROW,
    ColumnSpec,
    _apply_template_font,
    _auto_width,
    _is_blank,
    _set_template_font,
    _to_bool,
    _to_time,
    read_raw_rows,
)
from comken.toolbox.excel import Excel

FREQUENCY_HOURLY = "1時間ごと"
FREQUENCY_DAILY = "毎日"
FREQUENCY_WEEKLY = "毎週"
FREQUENCY_MONTHLY = "毎月"
HOLIDAY_SKIP = "取得しない"
WEEKDAY_NAMES = ("月", "火", "水", "木", "金", "土", "日")

logger = logging.getLogger(__name__)

# 「第N営業日」表記の正規表現。N は 1 以上の整数。「曜日付き」（例: 「第2営業日（月曜）」）
# のような表記は受け付けず、もっとも素直な形の入力を要求する
_NTH_BUSINESS_DAY_PATTERN = re.compile(r"^第(\d+)営業日$")

# レポート管理表と同じブック内のスケジュール管理シート名。**管理表本体と
# 一緒に置かれる**ため、このシートが無い管理表でもエラーにせず空とみなす
# （後方互換。詳細は load_schedule() を参照）
SCHEDULE_SHEET_NAME = "スケジュール"

# 「スケジュール」シートの列（見出し）のフルセット。``ScheduleRule.from_row`` が
# 受け付ける列を全て並べたもの。``load_schedule`` のテスト
# (``tests/test_schedule_load.py`` の ``SCHEDULE_HEADERS``) はサブセットだが、
# 雛形としては不足列（``取得間隔（分）`` / ``日付``）も最初から出して、利用者が
# 「列を足す」工程を踏まずに済む形にする。
SCHEDULE_HEADERS_FULL: tuple[str, ...] = (
    "スケジュールキー",
    "レポートキー",
    "取得頻度",
    "曜日",
    "取得時刻",
    "取得間隔（分）",
    "日付",
    "祝日対応",
    "有効",
)

# 雛形に入れる列の宣言（``report_master.ColumnSpec``）。``choices`` を付けた列に
# は Excel のドロップダウン（入力規則）が付く（``create_schedule_template`` 内で
# 直接 ``DataValidation`` を組み立てる。``column()`` ヘルパーは ``MasterRow``
# 前提のため使わない）。
SCHEDULE_COLUMN_SPECS: tuple[ColumnSpec, ...] = (
    ColumnSpec(
        header="スケジュールキー",
        help="スケジュールを識別する名前。1行ごとに違う値を書いてください。"
        "重複しているとエラーになります",
    ),
    ColumnSpec(
        header="レポートキー",
        help="「管理表」シートの「ID」列の値。定期実行したいレポートの管理番号を書いてください",
    ),
    ColumnSpec(
        header="取得頻度",
        choices=(FREQUENCY_HOURLY, FREQUENCY_DAILY, FREQUENCY_WEEKLY, FREQUENCY_MONTHLY),
        help="「毎日」「毎週」「毎月」「1時間ごと」から選んでください",
    ),
    ColumnSpec(
        header="曜日",
        choices=WEEKDAY_NAMES,
        help="「毎週」のときに曜日を選んでください"
        "（「毎日」「1時間ごと」のときは空欄で構いません）",
    ),
    ColumnSpec(
        header="取得時刻",
        help="「毎日」「毎週」「毎月」「1時間ごと」共通の実行開始時刻。"
        "HH:MM 形式（例: 09:00）で書いてください"
        "（「1時間ごと」のときはここが開始時刻になります）。"
        "「毎日」「毎週」「毎月」で、時刻を問わず1日のうちいつでも取得してよい"
        "場合（例: 前日以前の確定済みデータのように、いつ取っても中身が変わらない"
        "レポート）は空欄のままにできます（「1時間ごと」では必須です）",
    ),
    ColumnSpec(
        header="取得間隔（分）",
        help="「1時間ごと」のとき、何分おきに取得するか。60（1時間）などを数字だけで書いてください",
    ),
    ColumnSpec(
        header="日付",
        help="「毎月」のとき、月の何日に取得するか。"
        "1〜31の数字、月末なら「月末」、第N営業日なら「第2営業日」のように書いてください",
    ),
    ColumnSpec(
        header="祝日対応",
        help="「取得しない」と書くと祝日はスキップします。それ以外（空欄含む）は祝日でも取得します",
    ),
    ColumnSpec(
        header="有効",
        choices=("○", "×"),
        help="「○」か「×」と書いてください（「×」にすると、その行は取得対象から外れます）",
    ),
)

# 雛形に書き込む記入例。2行で「毎週」と「1時間ごと」を見せる。**``備考`` 列は
# スケジュール表に存在しない**ため、削除を促す案内は「記入方法」シート側に書く。
_SCHEDULE_EXAMPLES: tuple[dict[str, object], ...] = (
    {
        "スケジュールキー": "S001",
        "レポートキー": "1001",
        "取得頻度": FREQUENCY_WEEKLY,
        "曜日": "月",
        "取得時刻": "09:00",
        "取得間隔（分）": "",
        "日付": "",
        "祝日対応": HOLIDAY_SKIP,
        "有効": "○",
    },
    {
        "スケジュールキー": "S002",
        "レポートキー": "1002",
        "取得頻度": FREQUENCY_HOURLY,
        "曜日": "",
        "取得時刻": "09:00",
        "取得間隔（分）": 60,
        "日付": "",
        "祝日対応": "",
        "有効": "○",
    },
)

# ガイドシート（``記入方法``）にスケジュール用注意文として書く案内。雛形だけで
# 分かり、``create_schedule_template`` の引数 ``examples`` を ``None`` にした
# ときのデフォルト examples と整合する。
_SCHEDULE_GUIDE_NOTE = "スケジュールの記入例の行は、実際に使うときに消してください"

# 雛形に付ける「ドロップダウン範囲の行数」。``report_master`` と同じ値を使う。
_DATA_VALIDATION_ROWS = 1000

# 条件付き書式で「今回は使わない列」をグレーアウトする際の塗り色。
# ``_EXAMPLE_FILL``（記入例の薄灰色）とは別に、もう少し濃いグレーにして
# 「入力例」と「使わない列」を視覚的に区別できるようにする
_UNUSED_COLUMN_FILL = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

# 雛形用のフォント名（``report_master`` と同じ「Noto Sans JP」）。PC に
# 入っていないときは Excel が代替フォントを選ぶが、雛形としては問題なく使える。
_TEMPLATE_FONT_NAME = "Noto Sans JP"

# ガイドシート名。``MasterRow.create_template`` と同じ ``記入方法`` を共有する。
_GUIDE_SHEET_NAME = "記入方法"

# 既存のガイドシートへ追記するときの空き行数（視覚的な区切り）。
_GUIDE_APPEND_BLANK_ROWS = 2

# ガイドシートで見出しとして使う3列（``_write_guide`` と同じ）。
_GUIDE_HEADERS: tuple[str, str, str] = ("列", "何を書くか", "書けない場合")


@dataclass(frozen=True)
class ScheduleRule:
    """取得スケジュール管理表の1行。"""

    schedule_key: str
    report_key: str
    frequency: str
    run_time: dt.time | None = None
    interval_minutes: int | None = None
    weekday: int | None = None
    day_of_month: int | None = None
    month_end: bool = False
    nth_business_day: int | None = None
    holiday_policy: str = HOLIDAY_SKIP
    enabled: bool = True

    @classmethod
    def from_row(cls, row: Mapping[str, object]) -> Self:
        """日本語カラム名の辞書からスケジュールを作る。"""
        day_of_month, month_end, nth_business_day = _parse_day_of_month(row.get("日付"))
        return cls(
            schedule_key=_required_text(row, "スケジュールキー"),
            report_key=_required_text(row, "レポートキー"),
            frequency=_required_text(row, "取得頻度"),
            run_time=_to_time(row.get("取得時刻")),
            interval_minutes=_parse_int(row.get("取得間隔（分）")),
            weekday=_parse_weekday(row.get("曜日")),
            day_of_month=day_of_month,
            month_end=month_end,
            nth_business_day=nth_business_day,
            holiday_policy=_text_or_default(row, "祝日対応", HOLIDAY_SKIP),
            enabled=_to_bool(row.get("有効")),
        )

    def is_due(
        self,
        now: dt.datetime,
        *,
        holidays: set[dt.date] | frozenset[dt.date] = frozenset(),
        calendar: HolidayCalendar | None = None,
    ) -> bool:
        """指定時刻にこのスケジュールを実行すべきか判定する。

        ``calendar`` は「日付」列に「第N営業日」を指定した行の判定にのみ使う
        （``comken.core.holidays.nth_business_day_of_month`` に渡す）。省略時は
        ``default_calendar()`` にフォールバックする。``holidays`` 引数（祝日の
        ``set[date]``）は独立に残しており、「第N営業日」以外での祝日判定に使う。

        ``FREQUENCY_DAILY`` / ``FREQUENCY_WEEKLY`` / ``FREQUENCY_MONTHLY`` で
        ``run_time is None`` のときは「時刻条件なし」を意味し、日付条件が合えば常に
        ``True`` を返す（例: 前日以前の確定済みデータのように、いつ取っても同じ内容の
        レポート用）。``FREQUENCY_HOURLY`` は対象外で、``run_time`` が無いと
        ``ScheduleIntervalMissingError`` を投げる。
        """
        if not self.enabled or not self._date_matches(now.date(), holidays, calendar):
            return False
        if self.frequency == FREQUENCY_HOURLY:
            return self._is_hourly_due(now)
        if self.frequency in {FREQUENCY_DAILY, FREQUENCY_WEEKLY, FREQUENCY_MONTHLY}:
            return self.run_time is None or now.time() >= self.run_time
        raise UnsupportedScheduleFrequencyError(self.frequency)

    def _date_matches(
        self,
        date: dt.date,
        holidays: set[dt.date] | frozenset[dt.date],
        calendar: HolidayCalendar | None = None,
    ) -> bool:
        if self.holiday_policy == HOLIDAY_SKIP and date in holidays:
            return False
        if self.weekday is not None and date.weekday() != self.weekday:
            return False
        if self.day_of_month is not None and date.day != self.day_of_month:
            return False
        if self.nth_business_day is not None:
            cal = calendar if calendar is not None else default_calendar()
            try:
                target = nth_business_day_of_month(
                    date.replace(day=1), self.nth_business_day, calendar=cal
                )
            except BusinessDayNotFoundError:
                # 「第N営業日」がその月の営業日数を超える設定ミスのケース。
                # ここで呼び出し元（``download_scheduled``）全体を止めると、
                # 同じ管理表内の他レポートの取得まで巻き添えになるため、
                # この日は対象外として扱いログだけ残す
                logger.warning(
                    "スケジュール %s の「第%d営業日」指定が %s年%s月の営業日数を"
                    "超えています。この日は対象外として扱います。",
                    self.schedule_key,
                    self.nth_business_day,
                    date.year,
                    date.month,
                )
                return False
            if date != target:
                return False
        return not self.month_end or (date + dt.timedelta(days=1)).month != date.month

    def _is_hourly_due(self, now: dt.datetime) -> bool:
        if self.run_time is None or not self.interval_minutes:
            raise ScheduleIntervalMissingError()
        if now.time() < self.run_time:
            return False
        start_minutes = self.run_time.hour * 60 + self.run_time.minute
        now_minutes = now.hour * 60 + now.minute
        return (now_minutes - start_minutes) % self.interval_minutes == 0


def _required_text(row: Mapping[str, object], column: str) -> str:
    value = str(row.get(column, "")).strip()
    if not value:
        raise ScheduleRequiredValueMissingError(column)
    return value


def _text_or_default(row: Mapping[str, object], column: str, default: str) -> str:
    value = str(row.get(column, "")).strip()
    return value or default


def _parse_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    return int(str(value).strip())


def _parse_day_of_month(value: object) -> tuple[int | None, bool, int | None]:
    """「日付」列を ``(day_of_month, month_end, nth_business_day)`` に分解する。

    空欄は「指定なし」、数字 1〜31 は ``day_of_month``、文字列「月末」は
    ``month_end=True``、``"第N営業日"``（N は 1 以上の整数）は
    ``nth_business_day=N`` として扱う。``_parse_int`` と同じ緩さで解釈し、
    想定外の値（例: ``"来月"``）は ``int()`` 由来の ``ValueError`` をそのまま
    投げる（専用のエラー型は用意しない）。

    Returns:
        3 要素のタプル。常にどれか 1 つだけが立ち、残りは「指定なし」になる。
    """
    if value in (None, ""):
        return None, False, None
    text = str(value).strip()
    if text == "月末":
        return None, True, None
    match = _NTH_BUSINESS_DAY_PATTERN.match(text)
    if match:
        return None, False, int(match.group(1))
    return _parse_int(text), False, None


def _parse_weekday(value: object) -> int | None:
    if value in (None, ""):
        return None
    text = str(value).strip().removesuffix("曜日")
    if text not in WEEKDAY_NAMES:
        raise ScheduleWeekdayInvalidError(value)
    return WEEKDAY_NAMES.index(text)


def load_schedule(path: str | Path | None = None) -> list[ScheduleRule]:
    """スケジュール管理シートを読んで、``ScheduleRule`` のリストを返す。

    **シートが存在しない場合はエラーにせず空リストを返す。** この機能を
    使っていない既存の管理表（「スケジュール」シートをまだ追加していないもの）が、
    このシートの有無で読み込みごと壊れないようにするため（後方互換）。

    空行は読み飛ばす（``ReportEntry`` と同じ扱い）。**スケジュールキーが
    重複している行はエラー**にする（``ReportEntry.key`` が ``unique=True``
    であるのと同じ考え方）。**行の中で値が壊れていた場合**は、``ScheduleRule.from_row``
    が投げた例外を ``ScheduleRowValueError`` で受け直して**行番号付き**で
    再送出する（業務担当者がどの行を直せばいいか分かるようにするため）。

    存在しない ``レポートキー`` を指している行はここではエラーにしない。
    レポート管理表との突き合わせは呼び出し側 ``download_scheduled()`` の責務。

    Args:
        path: 管理表（Excel）のパス。``None`` のときは ``MASTER_PATH``。

    Returns:
        宣言順に並んだ ``ScheduleRule`` のリスト。

    Raises:
        ScheduleDuplicateKeyError: スケジュールキーが重複している行がある。
        ScheduleRowValueError: 値の整合性エラー（行番号付き）。
        ExcelFileNotFoundError: ``path`` が存在しない場合。
    """
    if path is None:
        from comken.services.salesforce_downloader._paths import MASTER_PATH

        path = MASTER_PATH
    source = Path(path)

    # **シートが無い場合は空リストを返す。** この機能をまだ使っていない管理表を
    # 読み込み時に壊さないため。``ExcelFileNotFoundError`` などの「ファイル自体に
    # 関するエラー」はそのまま上位へ伝える
    try:
        raw_rows = read_raw_rows(source, SCHEDULE_SHEET_NAME)
    except SheetNotFoundError:
        return []

    rules: list[ScheduleRule] = []
    seen_keys: set[str] = set()
    for offset, raw in enumerate(raw_rows):
        if _is_blank(raw):
            continue  # 表の下に残った空行は読み飛ばす
        row_number = offset + _FIRST_DATA_ROW
        try:
            rule = ScheduleRule.from_row(raw)
        except DownloaderError as e:
            # ``ScheduleDuplicateKeyError`` も ``DownloaderError`` のサブクラスだが
            # この時点ではまだ送出される経路が無い（``from_row`` は送らない）ので
            # そのまま行番号を足して再送出する
            raise ScheduleRowValueError(row_number, str(e)) from e
        if rule.schedule_key in seen_keys:
            raise ScheduleDuplicateKeyError(rule.schedule_key, row_number, source)
        seen_keys.add(rule.schedule_key)
        rules.append(rule)
    return rules


def create_schedule_template(
    path: str | Path, examples: list[dict[str, object]] | None = None
) -> Path:
    """既存のレポート管理表（Excel）に「スケジュール」シートを追加する。

    **新しいブックは作らない。** ``path`` は ``PY_管理表`` シートをすでに持つ
    既存の管理表ファイルを指す前提。無ければ ``ExcelFileNotFoundError``。
    ``"スケジュール"`` シートが既にあれば ``SheetAlreadyExistsError``
    （既存データを消す事故を防ぐため、上書きしない）。

    雛形には次のものを入れる:

    - Excel の構造化テーブル（``openpyxl`` の ``Table``）。**``PY_T_スケジュール``**
      という名前で作成する（``Sheet.create_table`` の規約）
    - 記入例 2 行（「毎週」と「1時間ごと」を見せる）。**``report_master.py`` の
      ``_EXAMPLE_FILL`` と同じ薄灰色**を塗り、``SCHEDULE_HEADERS_FULL`` の全列に
      適用する
    - ``choices`` 付きの列には Excel のドロップダウン（入力規則）。``祝日対応``
      列は自由記述のため付けない
    - 全セルに ``Noto Sans JP`` のフォント（``report_master._apply_template_font``
      と ``_set_template_font`` を流用）
    - 既存の ``"記入方法"`` シートへの追記。**同じ見出し形式
      （列／何を書くか／書けない場合）**で ``SCHEDULE_COLUMN_SPECS`` の列ごとの
      説明を追加し、記入例は消す案内の注意文を1行足す。``"記入方法"`` シートが
      無いブックを渡された場合はエラーにせず新規作成してよい（``ReportEntry``
      側の説明が無いので、スケジュール部分だけの簡易版になる）

    Args:
        path: 既存のレポート管理表（``.xlsx``）。
        examples: 雛形に入れる記入例。キーは **Excel の見出し**（日本語）。
            ``None`` のときは ``_SCHEDULE_EXAMPLES``（2 行）を使う。

    Returns:
        ``path`` をそのまま返す（``with`` で保存まで行う）。

    Raises:
        ExcelFileNotFoundError: ``path`` が存在しない場合。
        SheetAlreadyExistsError: ``"スケジュール"`` シートが既に存在する場合。
    """
    source = Path(path)
    if not source.exists():
        # 書き込みモードで開くと「ブックが無い＝新規ブック作成」になり、
        # 既存の管理表が消えてしまう。先に明示的に弾く
        raise ExcelFileNotFoundError(source)

    rows = list(examples) if examples is not None else list(_SCHEDULE_EXAMPLES)

    # 構造化テーブル（``Sheet.create_table`` の要件）に渡す Table を作る
    table_rows = [
        {header: row.get(header, "") for header in SCHEDULE_HEADERS_FULL}
        for row in rows
    ]
    schedule_table = CoreTable(list(SCHEDULE_HEADERS_FULL), table_rows)

    # スケジュール用データシートを作成。既存なら ``SheetAlreadyExistsError``
    with Excel(source) as excel:
        excel.create_data_sheet(SCHEDULE_SHEET_NAME).create_table(
            "スケジュール", schedule_table
        )

    # フォント・ドロップダウン・列幅・記入例の背景色は ``openpyxl`` で直接当てる
    # （``Sheet.create_table`` は値・テーブル定義だけを書き、書式は触らないため）
    book = load_workbook(source)
    schedule_sheet = book[f"PY_{SCHEDULE_SHEET_NAME}"]

    _apply_template_font(schedule_sheet, len(rows))
    _apply_schedule_choice_validations(schedule_sheet, len(rows))
    _apply_schedule_conditional_formatting(schedule_sheet, len(rows))
    _auto_width(schedule_sheet)
    schedule_sheet.freeze_panes = "A2"

    # 記入例の行に薄灰色を塗る。**本物のデータと見分けが付く**ようにするための
    # 目印。 ``report_master._EXAMPLE_FILL`` と同じ ``PatternFill`` を使う
    for offset in range(len(rows)):
        row_number = _FIRST_DATA_ROW + offset
        for col in range(1, len(SCHEDULE_HEADERS_FULL) + 1):
            schedule_sheet.cell(row=row_number, column=col).fill = _EXAMPLE_FILL

    _append_schedule_guide(book)

    book.save(source)
    book.close()
    return source


def _apply_schedule_choice_validations(sheet: Worksheet, example_count: int) -> None:
    """スケジュール用シートの ``choices`` 列に Excel のドロップダウン（入力規則）を付ける。

    ``report_master._apply_choice_validations`` と同じ実装方針
    （``DataValidation(type="list", formula1=..., showDropDown=False,
    showErrorMessage=True, ...）``）を採るが、``choices`` の宣言は
    ``ColumnSpec`` 経由ではなく ``SCHEDULE_COLUMN_SPECS`` から直接組み立てる
    （``column()`` ヘルパーは ``MasterRow`` 前提のため使わない）。

    ``祝日対応`` 列は ``choices`` を宣言しないのでドロップダウンは付かない
    （「取得しない」以外は「祝日でも取得する」自由記述のため、選択肢を
    固定しない方が業務上の表現を縛らない）。

    **「取得頻度」は空欄を許さない。** ``ScheduleRule.from_row`` が
    ``_required_text`` で必須にしている列なので、Excel の画面上でも
    空欄のまま次のセルへ進めるとエラーメッセージが出るようにする
    （空欄のまま保存されると、`load_schedule()` を呼ぶまで気づけない）。
    """
    required_headers = {"取得頻度"}
    last_row = _FIRST_DATA_ROW + example_count - 1 + _DATA_VALIDATION_ROWS
    for offset, spec in enumerate(SCHEDULE_COLUMN_SPECS, start=1):
        if not spec.choices:
            continue  # ``choices`` を宣言していない列には付けない
        letter = get_column_letter(offset)
        choices_text = "、".join(f"「{choice}」" for choice in spec.choices)
        prompt = f"{spec.help}\n書き方: {choices_text}".strip()
        error = f"『{'』か『'.join(spec.choices)}』のいずれかを入力してください。"
        # ``showDropDown=False`` は Excel API の慣例で「ボタンを表示する」指定。
        # 非エンジニアが画面で選べる必要があるため True（=ボタンを表示しない）
        # にはしない
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(spec.choices)}"',
            allow_blank=spec.header not in required_headers,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="書き方が違います",
            error=error,
            showInputMessage=True,
            promptTitle=spec.header,
            prompt=prompt,
        )
        validation.add(f"{letter}{_FIRST_DATA_ROW}:{letter}{last_row}")
        sheet.add_data_validation(validation)


def _apply_schedule_conditional_formatting(sheet: Worksheet, example_count: int) -> None:
    """「曜日」「日付」列を、該当しない ``取得頻度`` のときグレーアウトする。

    「曜日」列は ``取得頻度`` が「毎週」でないとき、「日付」列は「毎月」でない
    ときグレーアウトする。**入力自体は禁止しない**（Excel の条件付き書式は
    見た目だけで、セル保護やデータ検証とは独立）。VBA を使わない方針を維持する
    ため、動的な入力制限は行わない。

    列位置は ``SCHEDULE_COLUMN_SPECS`` から「曜日」「日付」を探して動的に
    求める。``SCHEDULE_HEADERS_FULL`` の並びと ``SCHEDULE_COLUMN_SPECS`` の
    並びは1対1で対応しているという既存前提に乗っているので、ハードコードの
    列文字を使わない実装にしている（列の並びが変わっても ``SCHEDULE_HEADERS_FULL``
    と ``SCHEDULE_COLUMN_SPECS`` を更新するだけで追随できる）。

    ``FormulaRule`` の数式は範囲の左上セル基準の相対参照として書く。``$C{行}``
    のように列を絶対参照・行を相対参照にすることで、範囲内の各行がそれぞれ
    同じ行の C 列（取得頻度）を見るようになる（Excel の条件付き書式の標準的な
    書き方）。
    """
    last_row = _FIRST_DATA_ROW + example_count - 1 + _DATA_VALIDATION_ROWS
    weekday_letter = _resolve_schedule_column("曜日")
    date_letter = _resolve_schedule_column("日付")
    frequency_letter = _resolve_schedule_column("取得頻度")

    weekday_range = f"{weekday_letter}{_FIRST_DATA_ROW}:{weekday_letter}{last_row}"
    date_range = f"{date_letter}{_FIRST_DATA_ROW}:{date_letter}{last_row}"
    sheet.conditional_formatting.add(
        weekday_range,
        FormulaRule(
            formula=[
                f'${frequency_letter}{_FIRST_DATA_ROW}<>"{FREQUENCY_WEEKLY}"',
            ],
            fill=_UNUSED_COLUMN_FILL,
            stopIfTrue=False,
        ),
    )
    sheet.conditional_formatting.add(
        date_range,
        FormulaRule(
            formula=[
                f'${frequency_letter}{_FIRST_DATA_ROW}<>"{FREQUENCY_MONTHLY}"',
            ],
            fill=_UNUSED_COLUMN_FILL,
            stopIfTrue=False,
        ),
    )


def _resolve_schedule_column(header: str) -> str:
    """``SCHEDULE_COLUMN_SPECS`` から ``header`` を持つ列の列文字を返す。

    ``SCHEDULE_HEADERS_FULL`` の並びと ``SCHEDULE_COLUMN_SPECS`` の並びは1対1で
    対応するという既存前提に基づき、``SCHEDULE_HEADERS_FULL`` 側でも同じ
    インデックスを返す。``SCHEDULE_HEADERS_FULL.index(header) + 1`` を直接
    書いても同じだが、両者の順序が食い違っているときに気づけるよう一元化する。
    """
    for offset, spec in enumerate(SCHEDULE_COLUMN_SPECS, start=1):
        if spec.header == header:
            spec_index = offset
            break
    else:
        raise ValueError(f"{header} は SCHEDULE_COLUMN_SPECS に存在しません")
    headers_index = SCHEDULE_HEADERS_FULL.index(header) + 1
    if spec_index != headers_index:
        raise ValueError(
            f"{header} の位置が SCHEDULE_COLUMN_SPECS と SCHEDULE_HEADERS_FULL で"
            f"一致しません（{spec_index} vs {headers_index}）"
        )
    return get_column_letter(spec_index)


def _append_schedule_guide(book: object) -> None:
    """``"記入方法"`` シートに、スケジュール列の説明を追記する。

    既存シートがある場合は末尾（既存「注意」の下）から書き足し、無ければ
    新規作成してスケジュール部分だけを入れる。**既存の内容を壊さない**ため、
    シートの最終使用行を読み取ってから ``_GUIDE_APPEND_BLANK_ROWS`` 行空きを
    挟んでからセクションを始める。
    """
    if _GUIDE_SHEET_NAME in book.sheetnames:  # type: ignore[attr-defined]
        guide, start_row = _resume_existing_guide(book[_GUIDE_SHEET_NAME])  # type: ignore[attr-defined]
    else:
        guide = book.create_sheet(_GUIDE_SHEET_NAME)  # type: ignore[attr-defined]
        start_row = _init_new_guide(guide)

    _write_schedule_section(guide, start_row)
    _normalize_guide_fonts(guide, start_row)


def _resume_existing_guide(guide: Worksheet) -> tuple[Worksheet, int]:
    """既存の ``"記入方法"`` シートの末尾を特定して、追記開始行を返す。

    ``max_row`` は書式だけの空行でも広がるので「値の入った最大行」を自分で数える。
    戻り値の ``start_row`` はその最終行から ``_GUIDE_APPEND_BLANK_ROWS`` 行ぶん
    下で、視覚的な余白を空けてからセクションを書き始める位置。
    """
    last_used = 0
    for row in guide.iter_rows(values_only=False):
        for cell in row:
            if cell.value not in (None, ""):
                # ``Cell.row`` は openpyxl の型ヒント上 ``int | None`` だが、
                # ``iter_rows`` の戻り値経由では必ず ``int``。``cast`` で
                # 静的解析を黙らせつつ実行時の振る舞いは変えない
                row_index = cast(int, cell.row)
                if row_index > last_used:
                    last_used = row_index
    return guide, last_used + _GUIDE_APPEND_BLANK_ROWS


def _init_new_guide(guide: Worksheet) -> int:
    """新規 ``"記入方法"`` シートに見出しを書き、追記開始行を返す。

    戻り値の ``start_row`` は見出し（列 / 何を書くか / 書けない場合）の 1 行下に
    スケジュールセクションの見出しが来るように ``+ 2`` してある。
    """
    start_row = 1
    for column, value in enumerate(_GUIDE_HEADERS, start=1):
        cell = guide.cell(row=start_row, column=column, value=value)
        cell.font = Font(bold=True, name=_TEMPLATE_FONT_NAME)
    return start_row + 2


def _guide_note_text(spec: ColumnSpec) -> str:
    """``SCHEDULE_COLUMN_SPECS`` の各列に対する「書けない場合」欄の文言を返す。

    ``choices`` がある列は選択肢を列挙し、時刻・間隔・日付の列は形式を示し、
    それ以外は「空欄にできます」を返す。
    """
    if spec.choices:
        return f"「{'」か「'.join(spec.choices)}」と書いてください"
    if spec.header == "取得時刻":
        return (
            "HH:MM 形式で書いてください（例: 09:00）。"
            "「毎日」「毎週」「毎月」で時刻を問わず1日のうちいつでも取得してよい"
            "場合は空欄にできます（「1時間ごと」では必須）"
        )
    if spec.header == "取得間隔（分）":
        return "数字だけで書いてください（例: 60）"
    if spec.header == "日付":
        return "1〜31の数字、月末なら「月末」、第N営業日なら「第2営業日」のように書いてください"
    return "空欄にできます"


def _write_schedule_section(guide: Worksheet, start_row: int) -> None:
    """スケジュールセクション（見出し・列の説明・注意）を見出し行の直下から書き足す。"""
    heading_cell = guide.cell(row=start_row, column=1, value="スケジュール")
    heading_cell.font = Font(bold=True, name=_TEMPLATE_FONT_NAME)
    body_start = start_row + 1

    for offset, spec in enumerate(SCHEDULE_COLUMN_SPECS):
        row_number = body_start + offset
        note = _guide_note_text(spec)
        for column, value in enumerate((spec.header, spec.help, note), start=1):
            cell = guide.cell(row=row_number, column=column, value=value)
            cell.font = Font(name=_TEMPLATE_FONT_NAME)

    note_row = body_start + len(SCHEDULE_COLUMN_SPECS) + 1
    for column, value in enumerate(("注意", _SCHEDULE_GUIDE_NOTE, ""), start=1):
        cell = guide.cell(row=note_row, column=column, value=value)
        cell.font = Font(
            bold=(column == 1), name=_TEMPLATE_FONT_NAME
        )


def _normalize_guide_fonts(guide: Worksheet, start_row: int) -> None:
    """``"記入方法"`` シート全体のフォント名と列幅を整える。

    太字属性を保ったまま ``name`` を差し替える（``_set_template_font`` を一度通すと
    ``Font(bold=True)`` が消えるため、太字は別途保持する）。列幅は ``report_master``
    と同じ ``max_width=80`` を上限に自動調整する。

    既存シートへの追記時は ``freeze_panes`` を変えない（管理表セクションの
    スクロール位置がずれないように）。新規作成時だけ 1行目を固定する。
    """
    for row in guide.iter_rows():
        for cell in row:
            if cell.font.bold:
                cell.font = Font(
                    name=_TEMPLATE_FONT_NAME,
                    size=cell.font.size,
                    bold=True,
                    italic=cell.font.italic,
                    color=cell.font.color,
                )
            else:
                _set_template_font(cell)
    _auto_width(guide, max_width=80)
    if start_row == 1:
        guide.freeze_panes = f"A{start_row + 1}"


__all__ = [
    "ScheduleRule",
    "SCHEDULE_HEADERS_FULL",
    "SCHEDULE_SHEET_NAME",
    "create_schedule_template",
    "load_schedule",
]
