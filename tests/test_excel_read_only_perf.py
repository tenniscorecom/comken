"""``Excel(read_only=True)`` の動作契約を追加する。

- 読み取り専用では UNC パスでも作業コピー（=ローカルコピー）を作らない。
- シート経由でしか使わないメソッドでは通常 Workbook を遅延オープンする。
- ``read_computed_rows`` のような値だけ読む経路では通常 Workbook を開かず、
  openpyxl の read_only ストリームだけで完結する。
- テーブル API（``Sheet.table()``）は read_only でも従来どおり動く。

既存の ``test_excel_unified.py`` の隣に置くと統一感が出るが、隣接ファイルが
肥大するので独立ファイルにした。
"""

from __future__ import annotations

from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from comken.core.table import Table
from comken.toolbox.excel import Excel


def _book(path, value: str = "x") -> None:
    wb = Workbook()
    wb.active["A1"] = value
    wb.save(path)


def test_read_only_skips_local_copy_on_unc_path(tmp_path) -> None:
    """read_only=True のときは UNC パスでも作業コピー（ローカルコピー）を作らない。"""
    path = tmp_path / "unc.xlsx"
    _book(path)
    with (
        patch.object(Excel, "_is_unc_path", return_value=True),
        Excel(path, read_only=True) as excel,
    ):
        assert excel._local_copy_path is None
        assert excel._working_path == path


def test_read_only_skips_local_copy_on_unc_when_explicit(tmp_path) -> None:
    """``local_copy=True`` を明示したときは read_only でもコピーされることを許容する。"""
    path = tmp_path / "unc-explicit.xlsx"
    _book(path)
    with (
        patch.object(Excel, "_is_unc_path", return_value=True),
        Excel(path, read_only=True, local_copy=True) as excel,
    ):
        assert excel._local_copy_path is not None
        assert excel._local_copy_path.exists()


def test_read_only_normal_workbook_lazy(tmp_path) -> None:
    """``excel.sheet(...)`` を呼ぶまで ``load_workbook(read_only=False)`` は発火しない。"""
    path = tmp_path / "lazy.xlsx"
    _book(path)

    with (
        patch("comken.toolbox.excel.workbook.load_workbook") as load_spy,
        Excel(path, read_only=True),
    ):
        # __enter__ は終わってもまだ Excel 経由の値参照をしていないので Workbook を遅延。
        assert not load_spy.called

        # stream だけ開くパス（read_computed_rows）で通常 Workbook は開かない
        _read_with_excel(path)
        # stream の load_workbook は read_only=True で呼ばれることだけ確認
        assert load_spy.called
        for call in load_spy.call_args_list:
            kwargs = call.kwargs
            assert kwargs.get("read_only", False) is True


def _read_with_excel(path) -> int:
    """ベンチ用ヘルパ：read_computed_rows で行を読む。"""
    total = 0
    with Excel(path, read_only=True) as excel:
        rows = excel.read_computed_rows("Sheet", min_row=1)
        total += sum(len(r) for r in rows)
    return total


def test_read_only_stream_only_uses_openpyxl_read_only(tmp_path) -> None:
    """read_only の read_computed_rows は openpyxl load_workbook(read_only=True) だけを使う。"""
    path = tmp_path / "stream.xlsx"
    wb = Workbook()
    wb.active.append(["id", "name"])
    wb.active.append(["1", "a"])
    wb.active.append(["2", "b"])
    wb.save(path)

    normal_calls: list[tuple] = []
    stream_calls: list[tuple] = []

    def record(*args, **kwargs):
        if kwargs.get("read_only", False):
            stream_calls.append((args, kwargs))
        else:
            normal_calls.append((args, kwargs))
        return load_workbook(*args, **kwargs)

    with (
        patch("comken.toolbox.excel.workbook.load_workbook", side_effect=record),
        Excel(path, read_only=True) as excel,
    ):
        excel.read_computed_rows("Sheet", min_row=1)

    # 通常 Workbook は開かれない（__enter__ で read_only=True のときは遅延）
    assert normal_calls == []
    # ストリーム Workbook は read_only=True で 1 回以上開く
    assert stream_calls, "ストリーム Workbook が開かれていない"


def test_read_only_table_api_works(tmp_path) -> None:
    """``Sheet.table()`` が read_only でも従来どおり動く。"""
    path = tmp_path / "table.xlsx"
    with Excel(path) as excel:
        excel.create_data_sheet("顧客").create_table(
            "顧客",
            Table(["ID", "名前"], [{"ID": "001", "名前": "山田"}]),
        )
    with Excel(path, read_only=True) as excel:
        assert excel.data_sheet("顧客").table().read() == [
            {"ID": "001", "名前": "山田"}
        ]


def test_read_only_excel_table_opens_workbook_on_demand(tmp_path) -> None:
    """``Sheet.table()`` は開いた時点で通常 Workbook を遅延オープンする。"""
    path = tmp_path / "table-lazy.xlsx"
    with Excel(path) as excel:
        excel.create_data_sheet("顧客").create_table(
            "顧客",
            Table(["ID"], [{"ID": "001"}]),
        )

    normal_calls: list[tuple] = []

    def record(*args, **kwargs):
        if not kwargs.get("read_only", False):
            normal_calls.append((args, kwargs))
        return load_workbook(*args, **kwargs)

    with (
        patch("comken.toolbox.excel.workbook.load_workbook", side_effect=record),
        Excel(path, read_only=True) as excel,
    ):
        assert normal_calls == []  # __enter__ では遅延
        excel.data_sheet("顧客").table().read()
        # Sheet.table() を呼んだ時点で通常 Workbook が遅延オープン
        assert normal_calls, "通常 Workbook が遅延オープンされていない"


def test_read_only_cached_rows_no_formula_skips_formula_stream(tmp_path) -> None:
    """数式が無いブックでは ``data_only=False`` のストリームを開かない。"""
    path = tmp_path / "no-formula.xlsx"
    with Excel(path) as excel:
        sheet = excel.create_sheet("結果")
        sheet.write_value("A1", "a")
        sheet.write_value("B1", "b")
    formula_workbook_open_count = 0

    original_open = Excel._open_stream_workbook

    def counting_open(self, *, data_only):
        nonlocal formula_workbook_open_count
        if not data_only:
            formula_workbook_open_count += 1
        return original_open(self, data_only=data_only)

    with (
        patch.object(Excel, "_open_stream_workbook", counting_open),
        Excel(path, read_only=True) as excel,
    ):
        rows = excel.read_computed_rows("結果", min_row=1)
    assert rows == [("a", "b")]
    # data_only=False ストリームは 1 度も開かれない
    assert formula_workbook_open_count == 0
