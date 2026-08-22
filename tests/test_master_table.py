"""Excel の表を設定として読む仕組みを、実際に Excel を作って検証する。

Salesforce に依存しない汎用部分だけを見る（Salesforce 固有の列は
test_salesforce_downloader.py 側）。
"""

from dataclasses import dataclass
from pathlib import Path

import pytest
from openpyxl import load_workbook

from comken.core.table import Table
from comken.exceptions import (
    ExcelApplicationNotAvailableError,
    MasterColumnNotFoundError,
    MasterDuplicateValueError,
    MasterRowValueError,
    MasterSheetNotDefinedError,
)
from comken.services.salesforce_downloader.report_master import MasterRow, column
from comken.toolbox.excel import Excel


@dataclass(frozen=True, kw_only=True)
class Item(MasterRow):
    """検証用の1行。"""

    SHEET_NAME = "一覧"

    key: str = column("ID", unique=True, help="管理番号")
    name: str = column("名前", help="人が読んで分かる名前")
    source: Path = column("コピー元", help="共有サーバー上のファイル")  # noqa: RUF009
    mode: str = column("方式", choices=("毎日", "手動"), help="毎日は自動で取ります")
    enabled: bool = column("有効", choices=("○", "×"), help="「○」か「×」と書いてください")
    note: str = column("備考", default="", help="編集者の覚え書き")


HEADERS = ["ID", "名前", "コピー元", "方式", "有効", "備考"]
ROW_A = ["1001", "受注一覧", r"\\server\受注\data.csv", "毎日", "○", ""]
ROW_B = ["1002", "在庫", r"\\server\在庫\data.csv", "手動", "×", ""]


def make_sheet(path: Path, rows: list[list], headers: list[str] | None = None) -> Path:
    actual_headers = headers or HEADERS
    table_rows = [dict(zip(actual_headers, row, strict=False)) for row in rows]
    with Excel(path) as excel:
        excel.create_data_sheet("一覧").create_table("一覧", Table(actual_headers, table_rows))
    return path


class TestLoad:
    """宣言した型のとおりに読める。"""

    def test_reads_rows_with_declared_types(self, tmp_path):
        items = Item.load(make_sheet(tmp_path / "一覧.xlsx", [ROW_A, ROW_B]))
        assert [item.key for item in items] == ["1001", "1002"]
        assert isinstance(items[0].source, Path)  # 型注釈のとおり Path になる
        assert items[0].enabled is True
        assert items[1].enabled is False  # 「×」は False

    def test_blank_rows_are_skipped(self, tmp_path):
        """表の下に残った空行は読み飛ばす。"""
        items = Item.load(make_sheet(tmp_path / "一覧.xlsx", [ROW_A, [None] * 6]))
        assert len(items) == 1

    def test_blank_in_required_column_raises(self, tmp_path):
        """既定値の無い列が空欄なら止める（→ 理由は TestBlankPolicy）。"""
        row = ["1001", "受注一覧", r"\\server\a.csv", "毎日", None, ""]
        with pytest.raises(MasterRowValueError):
            Item.load(make_sheet(tmp_path / "一覧.xlsx", [row]))

    def test_blank_without_default_raises(self, tmp_path):
        """既定値の無い列が空なら、その行と列を示して止める。"""
        row = ["1001", None, r"\\server\a.csv", "毎日", "○", ""]
        with pytest.raises(MasterRowValueError) as e:
            Item.load(make_sheet(tmp_path / "一覧.xlsx", [row]))
        assert "2 行目" in str(e.value)  # 見出しが1行目なので、最初のデータは2行目
        assert "名前" in str(e.value)

    def test_path_without_default_raises(self, tmp_path):
        with pytest.raises(MasterSheetNotDefinedError):
            Item.load()  # PATH も引数も無い


class TestValidation:
    """非エンジニアが編集する表なので、どこが変かを示して止める。"""

    def test_value_outside_choices_raises(self, tmp_path):
        row = ["1001", "受注一覧", r"\\server\a.csv", "毎週", "○", ""]
        with pytest.raises(MasterRowValueError) as e:
            Item.load(make_sheet(tmp_path / "一覧.xlsx", [row]))
        assert "「毎日」か「手動」" in str(e.value)  # 書ける値を示す

    def test_non_numeric_key_raises(self, tmp_path):
        row = ["A001", "受注一覧", r"\\server\a.csv", "毎日", "○", ""]
        # str 型の key 列は「数字」を要求しない（むしろ数字以外も使える）。
        # 代わりに、重複していないことの検証として、別のシナリオで確認する
        items = Item.load(make_sheet(tmp_path / "一覧.xlsx", [row]))
        assert items[0].key == "A001"

    def test_numeric_string_key_is_preserved(self, tmp_path):
        """Excel が `1001` を数値セルで返すとき、`"1001"` として読める。"""
        # openpyxl 経由で、数値セル（float）として `1001` を書く
        path = make_sheet(
            tmp_path / "一覧.xlsx",
            [[1001, "受注一覧", r"\\server\a.csv", "毎日", "○", ""]],
        )

        items = Item.load(path)
        assert items[0].key == "1001"  # "1001.0" ではない

    def test_duplicate_unique_value_raises(self, tmp_path):
        rows = [ROW_A, ["1001", "別の名前", r"\\server\b.csv", "手動", "○", ""]]
        with pytest.raises(MasterDuplicateValueError) as e:
            Item.load(make_sheet(tmp_path / "一覧.xlsx", rows))
        assert "ID" in str(e.value)

    def test_missing_header_raises_with_existing_headers(self, tmp_path):
        """見出しを変えられたら、今ある見出しを示して止める。"""
        headers = ["ID", "名称", "コピー元", "方式", "有効", "備考"]  # 「名前」を「名称」に変えた
        with pytest.raises(MasterColumnNotFoundError) as e:
            Item.load(make_sheet(tmp_path / "一覧.xlsx", [ROW_A], headers))
        assert "名前" in str(e.value)
        assert "名称" in str(e.value)  # 今ある見出しも出す

    def test_value_outside_bool_choices_raises(self, tmp_path):
        """`enabled` を「○」「×」以外にするとエラー。表記が1つに絞られる。"""
        row = ["1001", "受注一覧", r"\\server\a.csv", "毎日", "有効", ""]
        with pytest.raises(MasterRowValueError) as e:
            Item.load(make_sheet(tmp_path / "一覧.xlsx", [row]))
        assert "「○」か「×」" in str(e.value)


class TestTemplate:
    """雛形は、そのまま読み込める状態で作られる。"""

    def test_generated_template_can_be_loaded(self, tmp_path):
        """**雛形と読み込みで列がズレない**（同じ宣言から作るため）。"""
        example = {
            "key": "1001",
            "name": "受注一覧",
            "source": Path(r"\\server\受注\data.csv"),
            "mode": "毎日",
            "enabled": True,
        }
        path = Item.create_template(tmp_path / "一覧.xlsx", [example])
        items = Item.load(path)
        assert len(items) == 1
        assert items[0].key == "1001"
        assert items[0].enabled is True  # True は「○」として書かれ、読み戻せる

    def test_bool_choices_round_trip(self, tmp_path):
        """bool 列の独自表記は、雛形へ書いて読み戻しても値が変わらない。"""

        @dataclass(frozen=True, kw_only=True)
        class WithBoolChoices(MasterRow):
            SHEET_NAME = "一覧"

            key: str = column("ID")
            is_allowed: bool = column("許可", choices=("○", "×"))

        examples = [
            {"key": "1", "is_allowed": True},
            {"key": "2", "is_allowed": False},
        ]
        path = WithBoolChoices.create_template(tmp_path / "一覧.xlsx", examples)

        sheet = load_workbook(path)["PY_一覧"]
        assert [sheet["B2"].value, sheet["B3"].value] == ["○", "×"]
        assert [row.is_allowed for row in WithBoolChoices.load(path)] == [True, False]

    def test_headers_are_written_in_declaration_order(self, tmp_path):
        path = Item.create_template(tmp_path / "一覧.xlsx")
        sheet = load_workbook(path)["PY_一覧"]
        assert [cell.value for cell in sheet[1]] == HEADERS

    def test_guide_sheet_lists_every_column(self, tmp_path):
        """非エンジニアが1枚で分かるよう、列ごとの説明を書く。"""
        path = Item.create_template(tmp_path / "一覧.xlsx")
        guide = load_workbook(path)["記入方法"]
        text = "\n".join(str(cell.value) for row in guide.iter_rows() for cell in row)
        for header in HEADERS:
            assert header in text
        assert "管理番号" in text  # help がそのまま出る
        assert "「毎日」か「手動」" in text  # 選択肢は書き方として出す

    def test_multiline_guide_intro_uses_one_cell_row(self, tmp_path):
        @dataclass(frozen=True, kw_only=True)
        class WithMultilineGuide(MasterRow):
            SHEET_NAME = "一覧"
            GUIDE_INTRO = "1行目\n2行目"

            key: str = column("ID")

        path = WithMultilineGuide.create_template(tmp_path / "一覧.xlsx")
        guide = load_workbook(path)["記入方法"]

        assert guide["A1"].value == "1行目\n2行目"
        assert [cell.value for cell in guide[3]] == ["列", "何を書くか", "書けない場合"]
        assert guide.freeze_panes == "A4"

    def test_table_is_created(self, tmp_path):
        """Excel のテーブルにしておくと、行を足すのが楽になる。"""
        path = Item.create_template(tmp_path / "一覧.xlsx", [{"key": "1", "name": "a"}])
        assert "PY_T_Item" in load_workbook(path)["PY_一覧"].tables

    def test_choice_columns_get_dropdown_in_template(self, tmp_path):
        """`choices` のある列には Excel のドロップダウン（入力規則）が付く。"""
        path = Item.create_template(tmp_path / "一覧.xlsx", [{"key": "1", "name": "a"}])
        ws = load_workbook(path)["PY_一覧"]
        validations = list(ws.data_validations.dataValidation)
        # 「方式」「有効」の 2 列にドロップダウンが付く。「名前」「コピー元」「ID」「備考」
        # には付かない
        ranges = sorted(str(v.sqref) for v in validations)
        assert ranges == ["D2:D1002", "E2:E1002"]
        formulas = {str(v.sqref): v.formula1 for v in validations}
        assert formulas["D2:D1002"] == '"毎日,手動"'
        assert formulas["E2:E1002"] == '"○,×"'

    def test_non_choice_columns_have_no_dropdown(self, tmp_path):
        """`choices` を宣言していない列にはドロップダウンが付かない。"""
        path = Item.create_template(tmp_path / "一覧.xlsx", [{"key": "1", "name": "a"}])
        ws = load_workbook(path)["PY_一覧"]
        # 「名前」列（B2:B1002）に validation が無いこと
        for v in ws.data_validations.dataValidation:
            assert str(v.sqref) != "B2:B1002"

    def test_template_font_is_noto_sans_jp(self, tmp_path):
        """雛形（表シート・記入方法シートとも）のフォントが Noto Sans JP。"""
        from openpyxl import load_workbook

        path = Item.create_template(tmp_path / "一覧.xlsx", [{"key": "1", "name": "a"}])
        wb = load_workbook(path)
        cell = wb["PY_一覧"]["A1"]
        assert cell.font.name == "Noto Sans JP"
        # ガイドシートの全セルのフォント名も Noto Sans JP
        guide = wb["記入方法"]
        for row in guide.iter_rows(min_row=1, max_row=10):
            for cell in row:
                assert cell.font.name == "Noto Sans JP"

    def test_example_rows_are_marked_with_note_and_fill(self, tmp_path):
        """記入例には「備考」に案内文が書いてあり、薄い背景色が付く。"""
        from openpyxl import load_workbook

        path = Item.create_template(
            tmp_path / "一覧.xlsx", [{"key": "1", "name": "a", "note": "記入例です"}]
        )
        ws = load_workbook(path)["PY_一覧"]
        # 記入例の備考列に案内文が書かれている
        assert ws["F2"].value == "記入例です"
        # 記入例全体に薄い背景色
        assert (
            ws["A2"].fill.fgColor.value.endswith("D9D9D9")
            or ws["A2"].fill.fgColor.rgb == "00D9D9D9"
        )


class TestColumnAdded:
    """あとから列を足しても、既存の表が読めなくならないこと。

    共有サーバーを更新すると全プロジェクトへ伝播するので、**列を1つ足した瞬間に
    既存の管理表がすべて読めなくなる**と業務が止まる。
    """

    def test_new_column_with_default_is_filled(self, tmp_path):
        """既定値のある列は、見出しごと無くても埋められる。"""

        @dataclass(frozen=True, kw_only=True)
        class WithNewColumn(MasterRow):
            SHEET_NAME = "一覧"

            key: str = column("ID", unique=True)
            name: str = column("名前")
            source: Path = column("コピー元")  # noqa: RUF009
            mode: str = column("方式", choices=("毎日", "手動"))
            enabled: bool = column("有効", choices=("○", "×"))
            memo: str = column("備考", default="")  # ← あとから足した列

        items = WithNewColumn.load(make_sheet(tmp_path / "一覧.xlsx", [ROW_A]))
        assert items[0].memo == ""  # 既定値で埋まる
        assert items[0].name == "受注一覧"  # 既存の列はそのまま読める

    def test_new_column_without_default_still_raises(self, tmp_path):
        """値が要る列を足したなら、管理表に足すまで止める。"""

        @dataclass(frozen=True, kw_only=True)
        class WithRequiredColumn(MasterRow):
            SHEET_NAME = "一覧"

            key: str = column("ID", unique=True)
            name: str = column("名前")
            source: Path = column("コピー元")  # noqa: RUF009
            mode: str = column("方式", choices=("毎日", "手動"))
            enabled: bool = column("有効", choices=("○", "×"))
            owner: str = column("担当", help="この一覧の持ち主")  # 既定値なし

        with pytest.raises(MasterColumnNotFoundError) as e:
            WithRequiredColumn.load(make_sheet(tmp_path / "一覧.xlsx", [ROW_A]))
        assert "担当" in str(e.value)

    def test_extra_column_in_excel_is_ignored(self, tmp_path):
        """宣言していない列が Excel にあっても無視する（列を消したとき）。"""
        headers = [*HEADERS, "使わない列"]
        rows = [[*ROW_A, "なにか"]]
        items = Item.load(make_sheet(tmp_path / "一覧.xlsx", rows, headers))
        assert items[0].key == "1001"


class TestHeaderAccess:
    """Python の名前から Excel の見出しを引ける。"""

    def test_header_returns_the_excel_header(self):
        assert Item.header("name") == "名前"
        assert Item.header("source") == "コピー元"

    def test_headers_are_in_declaration_order(self):
        assert Item.headers() == HEADERS

    def test_unknown_field_raises(self):
        with pytest.raises(KeyError):
            Item.header("存在しない")


class TestBlankPolicy:
    """空欄をどう扱うかは、既定値の有無で決まる。

    **既定値は「空欄でよい」という宣言。** 意味が反転する列（有効/無効）に既定値を
    付けると、書き忘れがそのまま「有効」になり、意図と逆の結果になる。
    """

    def test_blank_is_an_error_without_default(self, tmp_path):
        """既定値の無い列は、空欄なら止める（入力し忘れと区別が付かないため）。"""

        @dataclass(frozen=True, kw_only=True)
        class Strict(MasterRow):
            SHEET_NAME = "一覧"

            key: str = column("ID")
            enabled: bool = column("有効", choices=("○", "×"))  # 既定値を持たせない

        path = make_sheet(tmp_path / "一覧.xlsx", [["1001", None]], ["ID", "有効"])
        with pytest.raises(MasterRowValueError) as e:
            Strict.load(path)
        assert "有効" in str(e.value)

    def test_blank_uses_default_when_declared(self, tmp_path):
        """既定値があるなら、空欄は「そう書いた」とみなす。"""

        @dataclass(frozen=True, kw_only=True)
        class WithMemo(MasterRow):
            SHEET_NAME = "一覧"

            key: str = column("ID")
            memo: str = column("備考", default="")

        path = make_sheet(tmp_path / "一覧.xlsx", [["1001", None]], ["ID", "備考"])
        assert WithMemo.load(path)[0].memo == ""


class TestFormula:
    r"""セルに数式が入っていても、数式そのものが値として通らないこと。

    人は保存先などを数式で組み立てる（`=CONCATENATE(D2,"\input")`）。そのまま読むと
    **数式が文字列として通ってしまい**、エラーにもならず `=CONCATENATE(...)` という値で
    処理が進む。それが一番まずいので、計算結果を読む
    （計算結果がファイルに無いときだけ Excel を起動して計算させる）。
    """

    def _with_formula(self, path: Path) -> Path:
        make_sheet(path, [["1001", "受注一覧", "", "毎日", "○", r"\server\受注"]])
        book = load_workbook(path)
        book["PY_一覧"]["C2"] = r'=CONCATENATE(F2,"\data.csv")'
        book.save(path)
        book.close()
        return path

    def test_formula_is_never_taken_as_a_value(self, tmp_path):
        """数式が値として通らない。

        Excel がある PC では計算結果が読め、無い PC では
        ExcelApplicationNotAvailableError で止まる。**どちらでも
        '=CONCATENATE(...)' が値になることはない。**
        """
        path = self._with_formula(tmp_path / "数式.xlsx")
        try:
            items = Item.load(path)
        except ExcelApplicationNotAvailableError as e:
            assert "Excel" in str(e)  # Excel が無い PC。原因が分かる形で止まる
        else:
            assert "CONCATENATE" not in str(items[0].source)  # 計算結果が入っている
