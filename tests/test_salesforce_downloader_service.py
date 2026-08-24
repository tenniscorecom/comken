"""Salesforce レポートの集約取得（取りに行く側）を、Salesforce をモックして検証する。

管理表（Excel）と履歴（CSV）は tmp_path に本物を作り、実際に読み書きさせる。
Salesforce への通信だけを差し替える。

MASTER_PATH / HISTORY_PATH は `monkeypatch.setattr` で一時ディレクトリのパスへ
差し替える。利用側の API には管理表や履歴のパスを渡せない（設計判断）。
"""

from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from comken.core.table import Table
from comken.exceptions import (
    EmptyReportError,
    HistoryWriteError,
    InvalidReportURLError,
    MasterDuplicateValueError,
    MasterRowValueError,
    ReportDisabledError,
    ReportFolderNotFoundError,
    ReportNotRegisteredError,
    ScheduledDownloadFailedError,
)
from comken.services.salesforce_downloader import (
    ReportEntry,
    download_report,
    download_scheduled,
    history,
    load_master,
    shared_report_ids,
)
from comken.services.salesforce_downloader import provider as provider_module
from comken.services.salesforce_downloader import service as service_module
from comken.services.salesforce_downloader.cli import main as cli
from comken.services.salesforce_downloader.master import EXAMPLES
from comken.toolbox.csv import CSV
from comken.toolbox.excel import Excel

URL_A = "https://example--sandbox.sandbox.my.salesforce.com/lightning/r/Report/00O5g00000ABCDE/view"
URL_B = "https://example--sandbox.sandbox.my.salesforce.com/lightning/r/Report/00O5g00000FGHIJ/view"
ROWS = [{"名前": "山田", "金額": "100"}, {"名前": "鈴木", "金額": "200"}]

HEADERS = ["ID", "概要", "Salesforce URL", "実行方式", "保存先", "有効", "備考"]
# `0件あり` 列を足した見出し。列が無い管理表でも既定 `×` で読めることを確かめる
# ため、`paths` fixture は 7 列のままで固定する
HEADERS_WITH_ALLOW_EMPTY = [*HEADERS[:6], "0件あり", "備考"]


def make_master(path: Path, rows: list[list]) -> Path:
    """管理表（Excel）を作る。"""
    table_rows = [dict(zip(HEADERS, row, strict=True)) for row in rows]
    with Excel(path) as book:
        book.create_data_sheet("管理表").create_table("管理表", Table(HEADERS, table_rows))
    return path


def make_master_with_allow_empty(path: Path, rows: list[list]) -> Path:
    """`0件あり` 列を含む管理表（Excel）を作る。

    テスト1, 2, 3, 5, 6 で使う。テスト4（列が無くても読める）はあえて `make_master` の
    6 列版を使うので、ここでは7列版を別途用意する。
    """
    table_rows = [dict(zip(HEADERS_WITH_ALLOW_EMPTY, row, strict=True)) for row in rows]
    with Excel(path) as book:
        book.create_data_sheet("管理表").create_table(
            "管理表", Table(HEADERS_WITH_ALLOW_EMPTY, table_rows)
        )
    return path


@pytest.fixture
def paths(tmp_path, monkeypatch):
    """管理表・履歴・保存先をまとめて用意し、共有定数へ注入する。

    テスト関数側に `master_path=` / `history_path=` を渡さなくて済むように、
    `_paths.MASTER_PATH` / `_paths.HISTORY_PATH` を tmp_path 配下の値へ
    差し替える。`service.py` と `provider.py` は import 時に独自のローカル束縛を
    作るので、両方の属性も同期する。
    """
    folder = tmp_path / "保存先"
    folder.mkdir()
    master = make_master(
        tmp_path / "レポート管理表.xlsx",
        [
            ["1001", "顧客一覧", URL_A, "定期", str(folder), "○", ""],
            ["1002", "売上実績", URL_B, "個別", str(folder), "○", ""],
            ["1003", "停止中", URL_B, "定期", str(folder), "×", ""],
        ],
    )
    history_path = tmp_path / "ダウンロード履歴.csv"
    monkeypatch.setattr("comken.services.salesforce_downloader._paths.MASTER_PATH", master)
    monkeypatch.setattr("comken.services.salesforce_downloader._paths.HISTORY_PATH", history_path)
    # `service.MASTER_PATH` は `from ... import MASTER_PATH` で再束縛されているので、
    # `_paths` の差し替えだけでは反映されない。明示的に同じ値を入れる
    monkeypatch.setattr(service_module, "MASTER_PATH", master)
    monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)
    # `provider` も `_paths` から import で束縛しているので同期する
    monkeypatch.setattr(provider_module, "MASTER_PATH", master)
    return {
        "master_path": master,
        "history_path": history_path,
        "folder": folder,
    }


def fake_salesforce(rows: list[dict] | None = None) -> MagicMock:
    """report.run() が rows を返す Salesforce クライアント。"""
    client = MagicMock()
    client.__enter__.return_value.report.run.return_value = ROWS if rows is None else rows
    site = MagicMock(return_value=client)
    return site


class TestLoadMaster:
    """管理表の読み取りと検証。"""

    def test_reads_rows_and_extracts_report_id(self, paths):
        entries = load_master(paths["master_path"])
        assert list(entries) == ["1001", "1002", "1003"]
        # Salesforce のレポート ID は URL から取り出す（人には入力させない）
        assert entries["1001"].report_id == "00O5g00000ABCDE"
        assert entries["1001"].is_scheduled
        assert not entries["1002"].is_scheduled
        assert not entries["1003"].enabled

    def test_blank_rows_are_skipped(self, tmp_path):
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(tmp_path), "○", ""], [None] * 7],
        )
        assert list(load_master(master)) == ["1001"]

    def test_duplicate_key_raises(self, tmp_path):
        master = make_master(
            tmp_path / "管理表.xlsx",
            [
                ["1001", "顧客一覧", URL_A, "定期", str(tmp_path), "○", ""],
                ["1001", "別の名前", URL_B, "個別", str(tmp_path), "○", ""],
            ],
        )
        with pytest.raises(MasterDuplicateValueError):
            load_master(master)

    def test_url_without_report_id_raises_with_row_number(self, tmp_path):
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", "https://example.com/", "定期", str(tmp_path), "○", ""]],
        )
        with pytest.raises(InvalidReportURLError) as e:
            load_master(master)
        assert "1001" in str(e.value)  # 行番号ではなく管理番号で示す（空行があるとズレるため）

    def test_unknown_schedule_raises(self, tmp_path):
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "毎日", str(tmp_path), "○", ""]],
        )
        with pytest.raises(MasterRowValueError) as e:
            load_master(master)
        assert "実行方式" in str(e.value)

    def test_non_numeric_key_is_allowed(self, tmp_path):
        """str 型の `key` なので、文字列の ID もそのまま使える。"""
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["A001", "顧客一覧", URL_A, "定期", str(tmp_path), "○", ""]],
        )
        assert list(load_master(master)) == ["A001"]


class TestSharedReportIds:
    """同じ Salesforce レポートを複数の管理番号が指していることの検出。"""

    def test_detects_reports_used_by_multiple_keys(self, paths):
        entries = load_master(paths["master_path"])
        shared = shared_report_ids(entries)
        # "1002" と "1003" が同じ URL（＝同じレポート）を指している
        assert shared == {"00O5g00000FGHIJ": ["1002", "1003"]}

    def test_unique_reports_are_not_listed(self, tmp_path):
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(tmp_path), "○", ""]],
        )
        assert shared_report_ids(load_master(master)) == {}


class TestDownloadReport:
    """download_report() は必ず Salesforce へ取りに行く。"""

    def test_saves_file_and_returns_csv_reader(self, paths):
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            reader = download_report("1001", "案件集計")
        assert reader.path.is_file()
        assert reader.read().read() == ROWS

    def test_csv_reader_path_is_inherited_from_file_base(self, paths):
        reader = CSV(paths["history_path"])
        assert reader.path == paths["history_path"]

    def test_fetches_again_even_if_already_downloaded_today(self, paths):
        """今日すでに取っていても取り直す（明示的な最新取得なので）。"""
        site = fake_salesforce()
        with patch("comken.services.salesforce_downloader.service.site_for", return_value=site):
            download_report("1001")
            download_report("1001")
        saved = list(paths["folder"].glob("1001_*.csv"))
        assert len(saved) == 2
        assert saved[0].name != saved[1].name
        assert site.return_value.__enter__.return_value.report.run.call_count == 2

    def test_existing_collision_is_not_overwritten(self, paths, monkeypatch):
        collision = paths["folder"] / "1001_顧客一覧_fixed.csv"
        collision.write_text("既存", encoding="utf-8")
        monkeypatch.setattr(service_module, "file_path_of", lambda unused: collision)
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            result = download_report("1001")
        assert collision.read_text(encoding="utf-8") == "既存"
        assert result.path.name == "1001_顧客一覧_fixed_1.csv"

    def test_unregistered_key_raises(self, paths):
        with pytest.raises(ReportNotRegisteredError):
            download_report("9999")

    def test_disabled_report_raises(self, paths):
        with pytest.raises(ReportDisabledError):
            download_report("1003")

    def test_empty_report_raises_and_saves_nothing(self, paths):
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce([]),
            ),
            pytest.raises(EmptyReportError),
        ):
            download_report("1001")
        assert list(paths["folder"].glob("*.csv")) == []

    def test_missing_folder_raises_and_is_not_created(self, tmp_path, monkeypatch):
        missing = tmp_path / "無いフォルダ"
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(missing), "○", ""]],
        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            pytest.raises(ReportFolderNotFoundError),
        ):
            download_report("1001")
        assert not missing.exists()

    def test_no_temporary_file_is_left_behind(self, paths):
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            download_report("1001")
        assert list(paths["folder"].glob("~*")) == []


class TestHistory:
    """履歴には成否も、誰が要求したかも残る。"""

    def test_success_is_recorded_with_project_and_counts(self, paths):
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            download_report("1001", "案件集計")
        row = _history_rows(paths)[-1]
        assert row["管理番号"] == "1001"
        assert row["プロジェクト"] == "案件集計"
        assert row["実行方式"] == "個別"
        assert row["成否"] == "成功"
        assert row["Salesforce取得結果"] == "成功"
        assert row["保存結果"] == "成功"
        assert row["取得件数"] == "2"
        assert row["レポートID"] == "00O5g00000ABCDE"
        assert row["エラーコード"] == ""

    def test_failure_is_recorded(self, paths):
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce([]),
            ),
            pytest.raises(EmptyReportError),
        ):
            download_report("1001", "案件集計")
        row = _history_rows(paths)[-1]
        assert row["成否"] == "失敗"
        # 0 行でも Salesforce への問い合わせは成功している点が重要
        assert row["Salesforce取得結果"] == "成功"
        assert row["保存結果"] == ""
        assert row["エラーコード"] == "EmptyReportError"
        assert "0 行" in row["エラー内容"]

    def test_downloaded_today_only_counts_the_matching_trigger(self, paths):
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            download_report("1001")  # 個別として記録される
        # 個別に取っただけでは「本日の定期取得が済んだ」ことにはならない
        assert not history.downloaded_today(paths["history_path"], "1001")
        assert history.downloaded_today(paths["history_path"], "1001", history.TRIGGER_ON_DEMAND)

    def test_missing_history_file_is_not_downloaded(self, tmp_path):
        assert not history.downloaded_today(tmp_path / "無い.csv", "1001")

    # ── 履歴の5ケース（4. の表に対応する個別テスト）────────────────────
    def test_history_when_folder_is_missing(self, tmp_path, monkeypatch):
        """保存先フォルダが無い → 成否=失敗 / Salesforce取得結果=空 / 保存結果=空 /
        エラーコード=ReportFolderNotFoundError。"""
        missing = tmp_path / "無いフォルダ"
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(missing), "○", ""]],
        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            pytest.raises(ReportFolderNotFoundError),
        ):
            download_report("1001")
        rows = CSV(history_path).read().read()
        assert len(rows) == 1
        row = rows[0]
        assert row["成否"] == "失敗"
        assert row["Salesforce取得結果"] == ""
        assert row["保存結果"] == ""
        assert row["エラーコード"] == "ReportFolderNotFoundError"

    def test_history_when_salesforce_call_fails(self, tmp_path, monkeypatch):
        """Salesforce への問い合わせが失敗 → 成否=失敗 / Salesforce取得結果=失敗 /
        保存結果=空 / エラーコード=送出された例外クラス名。"""
        from comken.exceptions import SalesforceRequestError

        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(folder), "○", ""]],
        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        site = MagicMock()
        client = MagicMock()
        client.__enter__.return_value.report.run.side_effect = SalesforceRequestError(
            "GET", "/report", 500, "boom"
        )
        site.return_value = client

        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=site,
            ),
            pytest.raises(SalesforceRequestError),
        ):
            download_report("1001")
        rows = CSV(history_path).read().read()
        assert len(rows) == 1
        row = rows[0]
        assert row["成否"] == "失敗"
        assert row["Salesforce取得結果"] == "失敗"
        assert row["保存結果"] == ""
        assert row["エラーコード"] == "SalesforceRequestError"

    def test_history_when_report_is_empty(self, paths):
        """取得できたが 0 行だった → 成否=失敗 / Salesforce取得結果=成功 /
        保存結果=空 / エラーコード=EmptyReportError（通信は成功していて中身が空、という区別）。"""
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce([]),
            ),
            pytest.raises(EmptyReportError),
        ):
            download_report("1001")
        row = _history_rows(paths)[-1]
        assert row["成否"] == "失敗"
        assert row["Salesforce取得結果"] == "成功"
        assert row["保存結果"] == ""
        assert row["エラーコード"] == "EmptyReportError"

    def test_history_when_csv_write_fails(self, tmp_path, monkeypatch):
        """Salesforce 取得は成功したが CSV 書き込みが失敗 → 成否=失敗 /
        Salesforce取得結果=成功 / 保存結果=失敗 / エラーコード=送出された例外クラス名。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(folder), "○", ""]],
        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        write_error = OSError("書き込み失敗")

        def _raise_write(path, fieldnames, *args, **kwargs):
            raise write_error

        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            patch(
                "comken.services.salesforce_downloader.service.CSV._write",
                side_effect=write_error,
            ),
            pytest.raises(OSError),
        ):
            download_report("1001")
        rows = CSV(history_path).read().read()
        assert len(rows) == 1
        row = rows[0]
        assert row["成否"] == "失敗"
        assert row["Salesforce取得結果"] == "成功"
        assert row["保存結果"] == "失敗"
        assert row["エラーコード"] == "OSError"

    # ── 「原因区分」列（4区分 + 成功時の空文字）─────────────────────
    def test_cause_is_blank_on_success(self, paths):
        """成功時は原因区分が空文字。"""
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            download_report("1001")
        row = _history_rows(paths)[-1]
        assert row["成否"] == "成功"
        assert row["原因区分"] == ""

    def test_cause_is_config_when_folder_is_missing(self, tmp_path, monkeypatch):
        """保存先フォルダが無い → 「設定」（取得段階に入る前に落ちる）。"""
        missing = tmp_path / "無いフォルダ"
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(missing), "○", ""]],
        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            pytest.raises(ReportFolderNotFoundError),
        ):
            download_report("1001")
        row = CSV(history_path).read().read()[-1]
        assert row["原因区分"] == "設定"

    def test_cause_is_salesforce_when_request_fails(self, tmp_path, monkeypatch):
        """Salesforce への問い合わせが失敗 → 「Salesforce」。"""
        from comken.exceptions import SalesforceRequestError

        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(folder), "○", ""]],
        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        site = MagicMock()
        client = MagicMock()
        client.__enter__.return_value.report.run.side_effect = SalesforceRequestError(
            "GET", "/report", 500, "boom"
        )
        site.return_value = client

        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=site,
            ),
            pytest.raises(SalesforceRequestError),
        ):
            download_report("1001")
        row = CSV(history_path).read().read()[-1]
        assert row["原因区分"] == "Salesforce"

    def test_cause_is_empty_data_when_report_is_empty(self, paths):
        """取得できたが 0 行 → 「データなし」（取得は成功・保存未到達を一意に指す区分）。

        この管理表には `0件あり` 列が無いので `×` 既定扱いで `EmptyReportError` が送出される。
        段階は「取得成功 → 保存に進まず 0 行で失敗」になるため、4 区分だった頃の
        `Salesforce` から、新仕様の `データなし` に変わった。
        """
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce([]),
            ),
            pytest.raises(EmptyReportError),
        ):
            download_report("1001")
        row = _history_rows(paths)[-1]
        assert row["原因区分"] == "データなし"

    def test_cause_is_file_when_csv_write_fails(self, tmp_path, monkeypatch):
        """CSV 書き込みが OSError で失敗 → 「ファイル」（共有サーバー・権限）。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(folder), "○", ""]],
        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        write_error = OSError("書き込み失敗")
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            patch(
                "comken.services.salesforce_downloader.service.CSV._write",
                side_effect=write_error,
            ),
            pytest.raises(OSError),
        ):
            download_report("1001")
        row = CSV(history_path).read().read()[-1]
        assert row["原因区分"] == "ファイル"

    def test_cause_is_program_when_unexpected_error_raises(self, tmp_path, monkeypatch):
        """_fetch() が TypeError を投げる（comken 側のバグ想定）→ 「プログラム」。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(folder), "○", ""]],
        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        site = MagicMock()
        client = MagicMock()
        client.__enter__.return_value.report.run.side_effect = TypeError("想定外")
        site.return_value = client

        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=site,
            ),
            pytest.raises(TypeError),
        ):
            download_report("1001")
        row = CSV(history_path).read().read()[-1]
        assert row["原因区分"] == "プログラム"


class TestDownloadScheduled:
    """定期取得は「定期」かつ有効なものだけを対象にする。"""

    def test_only_enabled_scheduled_reports_are_downloaded(self, paths):
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            saved = download_scheduled()
        # "1001" だけ（"1002" は個別、"1003" は無効）
        assert [path.name.split("_")[0] for path in saved] == ["1001"]

    def test_one_failure_does_not_stop_the_rest(self, tmp_path, monkeypatch):
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [
                ["1001", "落ちる方", URL_A, "定期", str(tmp_path / "無い"), "○", ""],
                ["1002", "通る方", URL_B, "定期", str(folder), "○", ""],
            ],
        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()
        # "1001" で失敗しても "1002" は保存されている（続けたうえで最後に知らせる）
        assert [path.name.split("_")[0] for path in folder.glob("*.csv")] == ["1002", "1002"]

    def test_os_error_does_not_stop_the_rest(self, tmp_path, monkeypatch):
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [
                ["1001", "書込失敗", URL_A, "定期", str(folder), "○", ""],
                ["1002", "取得成功", URL_B, "定期", str(folder), "○", ""],
            ],
        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")
        original_write_csv = service_module._write_csv

        def fail_first_write(path, rows):
            if path.name.startswith("1001_"):
                raise OSError("共有サーバーへ書き込めません")
            original_write_csv(path, rows)

        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            patch.object(service_module, "_write_csv", side_effect=fail_first_write),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()

        assert [path.name.split("_")[0] for path in folder.glob("*.csv")] == ["1002", "1002"]

    def test_cache_update_failure_keeps_archive_and_marks_failure(self, paths):
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            patch.object(
                service_module,
                "_update_daily_cache",
                side_effect=OSError("キャッシュ更新失敗"),
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()

        assert len(list(paths["folder"].glob("1001_*.csv"))) == 1
        row = _history_rows(paths)[-1]
        assert row["成否"] == "失敗"
        assert row["保存結果"] == "失敗"

    def test_unexpected_error_stops_the_run_immediately(self, tmp_path, monkeypatch):
        """想定外（`TypeError` など）はその場で抜ける。`ScheduledDownloadFailedError` には
        変換しない（非エンジニアが「もう一度実行してみる」を繰り返すだけになるため）。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [
                ["1001", "想定外", URL_A, "定期", str(folder), "○", ""],
                ["1002", "通る方", URL_B, "定期", str(folder), "○", ""],
            ],
        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")

        site = MagicMock()
        client = MagicMock()
        # 1001 のレポートを取るときだけ想定外を投げる
        client.__enter__.return_value.report.run.side_effect = [
            TypeError("想定外"),
            ROWS,
        ]
        site.return_value = client

        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=site,
            ),
            pytest.raises(TypeError),
        ):
            download_scheduled()
        # 想定外で止めたので、2件目は保存されない
        # （続けた結果の ScheduledDownloadFailedError ではないことを確認）
        assert list(folder.glob("*.csv")) == []

    def test_records_the_trigger_as_scheduled(self, paths):
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            download_scheduled("定期実行")
        assert _history_rows(paths)[-1]["実行方式"] == "定期"


def _history_rows(paths: dict) -> list[dict]:
    return CSV(paths["history_path"]).read().read()


class TestRequiredHistory:
    """履歴が書けない場合は、取得結果だけを成功として返さない。"""

    def test_history_write_failure_stops_download(self, paths):
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            patch.object(history, "_append", side_effect=OSError("履歴書込み失敗")),
            pytest.raises(HistoryWriteError),
        ):
            download_report("1001")

    def test_original_failure_remains_in_history_error(self, paths):
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce([]),
            ),
            patch.object(history, "_append", side_effect=OSError("履歴書込み失敗")),
            pytest.raises(HistoryWriteError) as caught,
        ):
            download_report("1001")
        assert "0 行" in str(caught.value)
        assert caught.value.__cause__ is not None


# ── 0件あり / 0 行の扱い ─────────────────────────────────────────────
class TestAllowEmpty:
    """管理表の「0件あり」列で、0 行のとき失敗にするか正常終了にするかを選ぶ。"""

    def test_empty_report_with_allow_empty_no_raises_and_records_empty_data_cause(
        self, tmp_path, monkeypatch
    ):
        """1. `0件あり` が `×` で 0 行 → `EmptyReportError`、ファイルができない、
        履歴の `原因区分` が `データなし`。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master_with_allow_empty(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(folder), "○", "×", ""]],
        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce([]),
            ),
            pytest.raises(EmptyReportError),
        ):
            download_report("1001")

        # ファイルは作られない
        assert list(folder.glob("*.csv")) == []
        # 履歴には `データなし` が残る（取得成功・保存未到達の組合せのみ取り得る）
        row = CSV(history_path).read().read()[-1]
        assert row["成否"] == "失敗"
        assert row["Salesforce取得結果"] == "成功"
        assert row["保存結果"] == ""
        assert row["エラーコード"] == "EmptyReportError"
        assert row["原因区分"] == "データなし"

    def test_empty_report_with_allow_empty_yes_succeeds_and_writes_empty_file(
        self, tmp_path, monkeypatch
    ):
        """2. `0件あり` が `○` で 0 行 → 例外にならない、空のファイルができる、
        履歴が `成否=成功` / `取得件数=0` / `原因区分` が空。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master_with_allow_empty(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(folder), "○", "○", ""]],
        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        with patch(
            "comken.services.salesforce_downloader.service.site_for",
            return_value=fake_salesforce([]),
        ):
            reader = download_report("1001")  # 例外にならない

        # 空ファイルが作られる（ただし 0 バイトでよい）
        saved = list(folder.glob("*.csv"))
        assert len(saved) == 1
        assert saved[0].read_bytes() == b""
        assert reader.path == saved[0]

        # 履歴は成功・取得件数 0・原因区分 空
        row = CSV(history_path).read().read()[-1]
        assert row["成否"] == "成功"
        assert row["Salesforce取得結果"] == "成功"
        assert row["保存結果"] == "成功"
        assert row["取得件数"] == "0"
        assert row["原因区分"] == ""
        assert row["エラーコード"] == ""

    def test_empty_csv_can_be_read_with_no_rows(self, tmp_path, monkeypatch):
        """3. 2. で作られたファイルを CSV で読むと空の Table を返す。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master_with_allow_empty(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(folder), "○", "○", ""]],
        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        with patch(
            "comken.services.salesforce_downloader.service.site_for",
            return_value=fake_salesforce([]),
        ):
            reader = download_report("1001")

        # CSV は 0 バイトでも例外を出さず、空の行リストを返す
        assert reader.read().read() == []
        # ヘッダー名での索引も空（ヘッダー行が無いので当然）
        assert reader.read().read() == []

    def test_scheduled_empty_report_can_be_received(self, tmp_path, monkeypatch):
        """0件で成功した定期取得は、本日取得済みとして空のまま受け取れる。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master_with_allow_empty(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(folder), "○", "○", ""]],
        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr("comken.services.salesforce_downloader._paths.MASTER_PATH", master)
        monkeypatch.setattr(
            "comken.services.salesforce_downloader._paths.HISTORY_PATH", history_path
        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)
        monkeypatch.setattr(provider_module, "MASTER_PATH", master)

        with patch(
            "comken.services.salesforce_downloader.service.site_for",
            return_value=fake_salesforce([]),
        ):
            download_scheduled()

        from comken.services.salesforce_downloader import cached_report

        reader = cached_report("1001")
        assert reader.path.is_file()
        assert reader.read().read() == []

    def test_master_without_allow_empty_column_defaults_to_no(self, tmp_path, monkeypatch):
        """4. `0件あり` の列が無い管理表でも読める（既定 `×` として扱われる）。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        # 7 列のまま = `0件あり` 列が無い管理表
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(folder), "○", ""]],
        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        entries = load_master(master)
        assert entries["1001"].allow_empty is False  # 既定値

        # 0 行のときは従来どおり `EmptyReportError`（列が無い管理表を
        # 後付けで読む既存プロジェクトを壊さないため）
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce([]),
            ),
            pytest.raises(EmptyReportError),
        ):
            download_report("1001")
        assert list(folder.glob("*.csv")) == []

    def test_invalid_allow_empty_value_raises(self, tmp_path):
        """5. `0件あり` に `○` `×` 以外を書くとエラーになる（choices で弾く）。"""
        master = make_master_with_allow_empty(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", URL_A, "定期", str(tmp_path), "○", "△", ""]],
        )
        with pytest.raises(MasterRowValueError) as e:
            load_master(master)
        # 列名でユーザーが書き方を直せるよう、見出しをメッセージに残す
        assert "0件あり" in str(e.value)

    def test_download_scheduled_with_allow_empty_yes_does_not_raise_on_zero_rows(
        self, tmp_path, monkeypatch
    ):
        """6. `download_scheduled()` で `0件あり` が `○` のレポートが 0 行でも
        `ScheduledDownloadFailedError` にならない。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master_with_allow_empty(
            tmp_path / "管理表.xlsx",
            [
                ["1001", "空でもOK", URL_A, "定期", str(folder), "○", "○", ""],
                ["1002", "普通のレポート", URL_B, "定期", str(folder), "○", "×", ""],
            ],
        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        # "1001" は 0 行、"1002" は通常データを返す
        site = MagicMock()
        client = MagicMock()

        def _run(report_id):
            return [] if report_id == "00O5g00000ABCDE" else ROWS

        client.__enter__.return_value.report.run.side_effect = _run
        site.return_value = client

        with patch("comken.services.salesforce_downloader.service.site_for", return_value=site):
            saved = download_scheduled()  # 例外にならない

        # 両方とも保存される（"1001" は空ファイル、"1002" は通常の CSV）
        names = sorted(path.name.split("_")[0] for path in saved)
        assert names == ["1001", "1002"]
        # 履歴を確認: "1001" は成功・0件、"1002" も成功・2件
        rows = CSV(history_path).read().read()
        by_key = {row["管理番号"]: row for row in rows}
        assert by_key["1001"]["成否"] == "成功"
        assert by_key["1001"]["取得件数"] == "0"
        assert by_key["1001"]["原因区分"] == ""
        assert by_key["1002"]["成否"] == "成功"
        assert by_key["1002"]["取得件数"] == "2"


class TestTemplate:
    """管理表の雛形は、そのまま読み込める状態で作られる。"""

    def test_generated_template_can_be_loaded(self, tmp_path):
        """雛形の記入例が、そのまま load_master() を通る（列名の食い違いが起きない）。"""
        path = ReportEntry.create_template(tmp_path / "レポート管理表.xlsx", EXAMPLES)
        entries = load_master(path)
        assert list(entries) == ["1001", "1002"]
        assert entries["1001"].is_scheduled
        assert not entries["1002"].is_scheduled

    def test_examples_point_at_different_reports(self, tmp_path):
        """記入例が同じレポートを指していると、check が重複として報告してしまう。"""
        entries = load_master(ReportEntry.create_template(tmp_path / "管理表.xlsx", EXAMPLES))
        assert shared_report_ids(entries) == {}

    def test_guide_sheet_is_included(self, tmp_path):
        """非エンジニアが1枚で分かるよう、記入方法のシートを付ける。"""
        from openpyxl import load_workbook

        path = ReportEntry.create_template(tmp_path / "レポート管理表.xlsx", EXAMPLES)
        assert "記入方法" in load_workbook(path).sheetnames


class TestCommandLine:
    """保守用コマンド（雛形作成・検査）。"""

    def test_init_creates_the_template(self, tmp_path, capsys):
        path = tmp_path / "レポート管理表.xlsx"
        assert cli(["init", str(path)]) == 0
        assert path.is_file()

    def test_init_does_not_overwrite(self, tmp_path, capsys):
        """記入済みの管理表を消さない。"""
        path = tmp_path / "管理表.xlsx"
        cli(["init", str(path)])
        before = path.read_bytes()
        assert cli(["init", str(path)]) == 0
        assert path.read_bytes() == before
        assert "すでにある" in capsys.readouterr().err

    def test_check_reports_counts(self, paths, capsys):
        assert cli(["check", str(paths["master_path"])]) == 0
        out = capsys.readouterr().out
        assert "登録 3 件" in out
        assert "00O5g00000FGHIJ" in out  # 同じレポートを指している管理番号を知らせる

    def test_check_returns_failure_for_a_broken_master(self, tmp_path, capsys):
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "顧客一覧", "https://example.com/", "定期", str(tmp_path), "○", ""]],
        )
        assert cli(["check", str(master)]) == 1
        assert "エラー:" in capsys.readouterr().err
