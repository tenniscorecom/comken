"""comken/toolbox/excel/table.py — Excel データシートを操作する。"""

from datetime import datetime
from typing import TYPE_CHECKING, TypeAlias

from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from comken.core.table.model import Table
from comken.exceptions import (
    DuplicateHeaderCellError,
    EmptyExcelTableError,
    EmptyHeaderCellError,
    InvalidTableInputError,
    InvalidTableOperationError,
    TableColumnMismatchError,
    TableFormulaOverwriteError,
)

if TYPE_CHECKING:
    from comken.toolbox.excel.workbook import Excel

Value: TypeAlias = str | int | float | bool | datetime


class ExcelTable:
    """データシート全体を1つのテーブルとして操作する。

    Sheet の表示操作と分けることで、表データの読み書きがレイアウト変更へ
    意図せず影響されないようにしている。
    """

    def __init__(self, excel: "Excel", worksheet: Worksheet, name: str | None = None) -> None:
        self._excel = excel
        self._worksheet = worksheet
        self._name = name

    def read(self, *, force_com: bool = False) -> Table:
        """Excelテーブルの実際の定義範囲だけを読み、値を返す。

        シートの使用範囲ではなく Excel が保持する ``ref`` を使うため、表の外に
        ある無関係なセルを現在の Table に混ぜません。数式の計算結果が
        保存されていない場合だけ内部でCOMへ切り替えます。``force_com=True``
        はキャッシュを信頼できないブックをExcel実機で強制再計算します。
        """
        self._excel._ensure_open()
        if self._name is None:
            table_names = list(self._worksheet.tables)
            if len(table_names) != 1:
                raise InvalidTableOperationError("対象テーブルを一意に決められません。")
            self._name = table_names[0]
        excel_table = self._worksheet.tables[self._name]
        min_col, min_row, max_col, max_row = _table_boundaries(excel_table.ref)
        formula_cells = [
            cell
            for row in self._worksheet.iter_rows(
                min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col
            )
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        if force_com or (formula_cells and self._excel._is_dirty):
            rows = self._excel._read_range_with_com(
                self._worksheet.title, min_col, min_row, max_col, max_row
            )
        elif formula_cells:
            rows, needs_com = self._excel._cached_range(
                self._worksheet.title, min_col, min_row, max_col, max_row
            )
            if needs_com:
                rows = self._excel._read_range_with_com(
                    self._worksheet.title, min_col, min_row, max_col, max_row
                )
        else:
            # 数式がない表は現在のメモリ上の値を読む。新規ブックはまだ
            # 作業ファイルが無いため、毎回キャッシュ用ブックを開いてはいけない。
            rows = [
                tuple(cell.value for cell in row)
                for row in self._worksheet.iter_rows(
                    min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
                )
            ]
        if not rows:
            raise EmptyExcelTableError(
                self._worksheet.title,
                "テーブル範囲を読み取れませんでした",
            )
        # ref が定義する列数をそのまま使う。末尾の空見出しを切り捨てると、
        # テーブル定義の壊れを見逃し、後ろの列データを失うため。
        headers = list(rows[0])
        empty_columns = [index for index, header in enumerate(headers, 1) if header is None]
        if empty_columns:
            raise EmptyHeaderCellError(empty_columns)
        duplicate_headers = [
            header for header in dict.fromkeys(headers) if headers.count(header) > 1
        ]
        if duplicate_headers:
            raise DuplicateHeaderCellError(duplicate_headers)
        result = [
            {
                str(header): (
                    ""
                    if value is None
                    else value
                    if str(header) in self._excel._types
                    else str(value)
                )
                for header, value in zip(headers, row, strict=True)
            }
            for row in rows[1:]
            if any(value is not None for value in row)
        ]
        return Table([str(header) for header in headers], result, types=self._excel._types)

    def replace(
        self,
        rows: list[dict[str, Value]] | Table,
        *,
        allow_formula_overwrite: bool = False,
    ) -> None:
        """データシート全体を置き換える。

        既存データ部に人が入れた数式があると、既定では ``TableFormulaOverwriteError``
        で止める。数式を値で潰すと依存セルや集計式が壊れたことに遅れて気づくため。
        意図的に上書きしてよいときだけ ``allow_formula_overwrite=True`` を渡す。

        渡された ``Table`` が **既存の数式列を含まない** 場合、その列はそのまま
        保持される。行が増えたぶんは、既存の数式を
        ``openpyxl.formula.translate.Translator`` で下方向へずらして埋める。
        行が減ったぶんは、数式セルの値を消す。

        見出しの列は **既存の見出しと名前で対応付ける**。既存の見出しに無い
        列名が含まれていた場合は ``TableColumnMismatchError``。
        """
        self._excel._ensure_writable("replace")
        if self._name is None:
            names = list(self._worksheet.tables)
            if len(names) != 1:
                raise InvalidTableOperationError("書き込み対象テーブルを一意に決められません。")
            self._name = names[0]
        excel_table = self._worksheet.tables[self._name]
        min_col, min_row, max_col, max_row = _table_boundaries(excel_table.ref)

        existing_headers = [
            str(self._worksheet.cell(row=min_row, column=col).value or "")
            for col in range(min_col, max_col + 1)
        ]
        formula_columns = self._detect_formula_columns(
            min_col=min_col,
            min_row=min_row,
            max_col=max_col,
            max_row=max_row,
            headers=existing_headers,
        )

        table = rows if isinstance(rows, Table) else Table(list(rows[0]) if rows else [], rows)
        rows_list = table.read_rows()
        passed_columns = [str(c) for c in table.columns]

        if not any(passed_columns):
            raise InvalidTableOperationError("列のないTableはExcelテーブルにできません。")

        # 既存の見出しに無い列名はエラー（黙って無視しない）
        missing_in_existing = [c for c in passed_columns if c not in existing_headers]
        if missing_in_existing:
            raise TableColumnMismatchError(self._name, missing_in_existing)

        # 既存テーブルから省かれた列は、すべて数式列である必要がある
        omitted = [c for c in existing_headers if c not in passed_columns]
        non_formula_omitted = [c for c in omitted if c not in formula_columns]
        if non_formula_omitted:
            raise TableColumnMismatchError(self._name, non_formula_omitted)

        # 渡された Table に数式列が含まれているならエラー
        if not allow_formula_overwrite:
            formulas_in_passed = [c for c in passed_columns if c in formula_columns]
            if formulas_in_passed:
                formula_locations = [
                    cell.coordinate
                    for row in self._worksheet.iter_rows(
                        min_row=min_row + 1,
                        max_row=max_row,
                        min_col=min_col,
                        max_col=max_col,
                    )
                    for cell in row
                    if isinstance(cell.value, str) and cell.value.startswith("=")
                ]
                raise TableFormulaOverwriteError(self._name, formula_locations)

        existing_col_by_name = {
            name: idx for idx, name in enumerate(existing_headers, min_col)
        }

        # 見出し行は渡された列を既存位置に書き込む（位置ではなく名前で対応付ける）
        for col_name in passed_columns:
            col_num = existing_col_by_name[col_name]
            self._worksheet.cell(row=min_row, column=col_num, value=col_name)

        # 行データの書き込み（数式列には触らない）
        for row_number, row in enumerate(rows_list, min_row + 1):
            for col_name in passed_columns:
                col_num = existing_col_by_name[col_name]
                self._worksheet.cell(
                    row=row_number, column=col_num, value=row.get(col_name, "")
                )

        # 数式列の埋め込み・クリア
        self._fill_formula_columns(
            min_col=min_col,
            min_row=min_row,
            old_max_row=max_row,
            new_row_count=len(rows_list),
            headers=existing_headers,
            formula_columns=formula_columns,
        )

        # 行が減った場合：旧データ行の全セルをクリア
        new_max_row = min_row + max(len(rows_list), 1)
        self._clear_removed_cells(
            min_col=min_col,
            old_max_col=max_col,
            old_max_row=max_row,
            new_max_row=new_max_row,
        )

        # ref を更新（列幅は変えない）
        last_cell = self._worksheet.cell(new_max_row, max_col).coordinate
        excel_table.ref = (
            f"{self._worksheet.cell(min_row, min_col).coordinate}:{last_cell}"
        )
        self._excel._mark_dirty()

    def append(
        self,
        rows: list[dict[str, Value]] | dict[str, Value] | Table,
        *,
        allow_formula_overwrite: bool = False,
    ) -> None:
        """Table、1行、または行リストを既存テーブルの末尾へ追加する。

        既存テーブルに数式列があっても、その列は保持される。渡された行に
        数式列が含まれている場合は ``TableFormulaOverwriteError``
        （``allow_formula_overwrite=True`` で上書き可能）。
        """
        self._excel._ensure_writable("append")
        if self._name is None:
            names = list(self._worksheet.tables)
            if len(names) != 1:
                raise InvalidTableOperationError("対象テーブルを一意に決められません。")
            self._name = names[0]
        excel_table = self._worksheet.tables[self._name]
        min_col, min_row, max_col, max_row = _table_boundaries(excel_table.ref)

        existing_headers = [
            str(self._worksheet.cell(row=min_row, column=col).value or "")
            for col in range(min_col, max_col + 1)
        ]
        formula_columns = self._detect_formula_columns(
            min_col=min_col,
            min_row=min_row,
            max_col=max_col,
            max_row=max_row,
            headers=existing_headers,
        )

        if isinstance(rows, Table):
            additions = rows.read_rows()
            additions_columns = [str(c) for c in rows.columns]
        elif isinstance(rows, dict):
            additions = [rows]
            additions_columns = list(rows.keys())
        elif isinstance(rows, list):
            additions = rows
            additions_columns = list(rows[0].keys()) if rows else []
        else:
            raise InvalidTableInputError(
                "ExcelTable の追記には Table、1行、または行リストを指定してください。"
            )

        formula_in_additions = [c for c in additions_columns if c in formula_columns]

        # 数式列を追加しようとしていたらエラー
        if formula_in_additions and not allow_formula_overwrite:
            formula_locations = [
                cell.coordinate
                for row in self._worksheet.iter_rows(
                    min_row=min_row + 1,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                )
                for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            ]
            raise TableFormulaOverwriteError(self._name, formula_locations)

        # 数式列を含めて既存行を読む（COM 経由の read() を避け、ワークシートから直接読む）
        current_rows = self._read_worksheet_rows(
            min_col=min_col,
            min_row=min_row,
            max_col=max_col,
            max_row=max_row,
            headers=existing_headers,
        )

        if formula_in_additions:
            # 数式列の値を上書きする: current に additions を結合し、replace に渡す
            current = Table(existing_headers, current_rows)
            current.append(additions)
            self.replace(current, allow_formula_overwrite=True)
            return

        # 数式列は保持する: current と additions から数式列を除外して replace に渡す
        non_formula_columns = [c for c in existing_headers if c not in formula_columns]
        if not non_formula_columns:
            # すべての既存列が数式列。書き込める列が無いので何もしない。
            return
        filtered_current_rows = [
            {header: row[header] for header in non_formula_columns}
            for row in current_rows
        ]
        filtered_additions = [
            {c: row[c] for c in non_formula_columns if c in row} for row in additions
        ]
        filtered_table = Table(non_formula_columns, filtered_current_rows)
        filtered_table.append(filtered_additions)
        self.replace(filtered_table)

    def count(self) -> int:
        """データ行数を返す。"""
        return len(self.read())

    def _detect_formula_columns(
        self,
        *,
        min_col: int,
        min_row: int,
        max_col: int,
        max_row: int,
        headers: list[str],
    ) -> set[str]:
        """既存データ部の数式セルを持つ列（見出し名）の集合を返す。

        見出し行は通常文字列なので対象外。
        """
        formula_columns: set[str] = set()
        for row in self._worksheet.iter_rows(
            min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            for cell, header in zip(row, headers):
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_columns.add(header)
        return formula_columns

    def _read_worksheet_rows(
        self,
        *,
        min_col: int,
        min_row: int,
        max_col: int,
        max_row: int,
        headers: list[str],
    ) -> list[dict[str, object]]:
        """ワークシートから既存テーブルの行を読む（COM を使わない）。

        ``read()`` は数式の計算結果を再計算するために Excel 実機 (COM) を呼ぶ
        ことがある。``append()`` で数式列を保持したい場合は数式セルを読み捨てる
        ので、ここでは COM を経由せずワークシートの生の値をそのまま使う。
        """
        result: list[dict[str, object]] = []
        for row in self._worksheet.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            values = [cell.value for cell in row]
            if all(value is None for value in values):
                continue
            result.append(
                {
                    str(header): ("" if value is None else value)
                    for header, value in zip(headers, values, strict=True)
                }
            )
        # 先頭行はヘッダーなので除く
        return result[1:] if result else []

    def _fill_formula_columns(
        self,
        *,
        min_col: int,
        min_row: int,
        old_max_row: int,
        new_row_count: int,
        headers: list[str],
        formula_columns: set[str],
    ) -> None:
        """数式列のセルを Translator で下方向へコピー、または余った行をクリアする。

        行が増えた場合は既存最終データ行の数式を起点に下方向へ翻訳、
        行が減った場合は余った行の式セルをクリアする。
        数式列が無いテーブルでは何もしない。
        """
        if not formula_columns:
            return
        new_max_row = min_row + max(new_row_count, 1)
        for col_index, header in enumerate(headers):
            col_num = min_col + col_index
            if header not in formula_columns:
                continue
            # 既存のデータ行のうち、直近の数式セルをコピー元にする
            source_cell = None
            for r in range(old_max_row, min_row, -1):
                cell = self._worksheet.cell(row=r, column=col_num)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    source_cell = cell
                    break
            if source_cell is None:
                # 既存データ行に数式が無い → 翻訳元が無いので何もしない
                continue
            if new_max_row > old_max_row:
                # 行が増えた: Translator で下方向へコピー
                translator = Translator(source_cell.value, origin=source_cell.coordinate)
                for new_row in range(old_max_row + 1, new_max_row + 1):
                    target_ref = self._worksheet.cell(
                        row=new_row, column=col_num
                    ).coordinate
                    self._worksheet.cell(
                        row=new_row,
                        column=col_num,
                        value=translator.translate_formula(target_ref),
                    )

    def _clear_removed_cells(
        self,
        *,
        min_col: int,
        old_max_col: int,
        old_max_row: int,
        new_max_row: int,
    ) -> None:
        """行が減った場合のみ、置換後の実テーブル範囲から外れる旧セルの値を空にする。

        列幅は変えないため、列方向のクリアは行わない。
        """
        if new_max_row >= old_max_row:
            return
        for row_number in range(new_max_row + 1, old_max_row + 1):
            for column in range(min_col, old_max_col + 1):
                self._worksheet.cell(row=row_number, column=column).value = None


def _table_boundaries(excel_table_ref: str) -> tuple[int, int, int, int]:
    """Excel テーブル ref を int 4 要素 ``(min_col, min_row, max_col, max_row)`` へ正規化する。

    openpyxl の ``range_boundaries()`` は返り値を ``tuple[int | None, ...]`` と推論するが、
    Excel テーブル ref は validated 済みの範囲なので通常 ``None`` にはならない。
    呼び出し側で None を意識せずに済むよう、ここで ``int`` へ揃えて返す。

    もし ``None`` が現れた場合のフォールバックは 0 とする。空テーブルを置換した結果として
    ref の各値が 0 に丸まっても、その直後の ``cell(...)`` 呼び出しが「先頭セル」として
    動くので、利用者に見える不整合は出ない（発注元の判断）。
    """
    min_col, min_row, max_col, max_row = range_boundaries(excel_table_ref)
    return (
        0 if min_col is None else min_col,
        0 if min_row is None else min_row,
        0 if max_col is None else max_col,
        0 if max_row is None else max_row,
    )