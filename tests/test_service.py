"""Salesforce レポートの集約取得（取りに行く側）を、Salesforce をモックして検証する。

管理表（Excel）と履歴（CSV）は tmp_path に本物を作り、実際に読み書きさせる。
Salesforce への通信だけを差し替える。

MASTER_PATH / HISTORY_PATH は `monkeypatch.setattr` で一時ディレクトリのパスへ
差し替える。利用側の API には管理表や履歴のパスを渡せない（設計判断）。
"""

import datetime as dt
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

# paths fixture が monkeypatch.setattr に直接渡せるよう、_paths モジュールを import しておく
import comken.services.salesforce_downloader._paths as _paths_module
from comken.core.table import Table
from comken.exceptions import (
    HistoryWriteError,
    InvalidReportURLError,
    MasterDuplicateValueError,
    MasterRowValueError,
    ScheduledDownloadFailedError,
)
from comken.services.salesforce_downloader import (
    ReportEntry,
    cached_report_path,
    download_scheduled,
    history,
    load_master,
    shared_report_ids,
)
from comken.services.salesforce_downloader import provider as provider_module
from comken.services.salesforce_downloader import service as service_module
from comken.services.salesforce_downloader.cli import main as cli
from comken.services.salesforce_downloader.master import EXAMPLES
from comken.toolbox.csv import CSV
from comken.toolbox.excel import Excel

URL_A = "https://example--sandbox.sandbox.my.salesforce.com/lightning/r/Report/00O5g00000ABCDE/view"
URL_B = "https://example--sandbox.sandbox.my.salesforce.com/lightning/r/Report/00O5g00000FGHIJ/view"
ROWS = [{"名前": "山田", "金額": "100"}, {"名前": "鈴木", "金額": "200"}]

HEADERS = [
    "ID",
    "グループ名",
    "担当者",
    "概要",
    "Salesforce URL",
    "出力形式",
    "保存先",
    "有効",
    "備考",
]
# `0件あり` 列を足した見出し。列が無い管理表でも既定 `×` で読めることを確かめる
# ため、`paths` fixture は 9 列のままで固定する
HEADERS_WITH_ALLOW_EMPTY = [
    "ID",
    "グループ名",
    "担当者",
    "概要",
    "Salesforce URL",
    "出力形式",
    "保存先",
    "有効",
    "0件あり",
    "備考",
]


def make_master(path: Path, rows: list[list]) -> Path:
    """管理表（Excel）を作る。"""
    table_rows = [dict(zip(HEADERS, row, strict=True)) for row in rows]
    with Excel(path) as book:
        book.create_data_sheet("管理表").create_table("管理表", Table(HEADERS, table_rows))
    return path


def make_master_with_allow_empty(path: Path, rows: list[list]) -> Path:
    """`0件あり` 列を含む管理表（Excel）を作る。

    テスト1, 2, 3, 5, 6 で使う。テスト4（列が無くても読める）はあえて `make_master` の
    6 列版を使うので、ここでは7列版を別途用意する。
    """
    table_rows = [dict(zip(HEADERS_WITH_ALLOW_EMPTY, row, strict=True)) for row in rows]
    with Excel(path) as book:
        book.create_data_sheet("管理表").create_table(
            "管理表", Table(HEADERS_WITH_ALLOW_EMPTY, table_rows)
        )
    return path


@pytest.fixture
def paths(tmp_path, monkeypatch):
    """管理表・履歴・保存先をまとめて用意し、共有定数へ注入する。

    テスト関数側に `master_path=` / `history_path=` を渡さなくて済むように、
    `_paths.MASTER_PATH` / `_paths.HISTORY_PATH` を tmp_path 配下の値へ
    差し替える。`service.py` と `provider.py` は import 時に独自のローカル束縛を
    作るので、両方の属性も同期する。

    `1001`（=CSV・有効）が `download_scheduled()` の唯一の有効件。
    `1002`（=Excel・有効）は Excel 出力テストが直接管理表を作る側で検証するため、
    ここでは**無効**にして対象外にしている。`1003` も無効。
    """
    folder = tmp_path / "保存先"
    folder.mkdir()
    master = make_master(
        tmp_path / "レポート管理表.xlsx",
        [
            [
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "",
            ],

            [
                "1002",
                "経理グループ",
                "佐藤",
                "売上実績",
                URL_B,
                "Excel",
                str(folder),
                "×",
                "",
            ],

            [
                "1003",
                "営業事務グループ",
                "山田",
                "停止中",
                URL_B,
                "CSV",
                str(folder),
                "×",
                "",
            ],

        ],
    )
    history_path = tmp_path / "ダウンロード履歴.csv"
    latest_status = tmp_path / "最新ステータス.xlsx"
    monkeypatch.setattr(_paths_module, "MASTER_PATH", master)
    monkeypatch.setattr(_paths_module, "HISTORY_PATH", history_path)
    monkeypatch.setattr(_paths_module, "LATEST_STATUS_PATH", latest_status)
    # `service.MASTER_PATH` は `from ... import MASTER_PATH` で再束縛されているので、
    # `_paths` の差し替えだけでは反映されない。明示的に同じ値を入れる
    monkeypatch.setattr(service_module, "MASTER_PATH", master)
    monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)
    monkeypatch.setattr(service_module, "LATEST_STATUS_PATH", latest_status)
    # `provider` も `_paths` から import で束縛しているので同期する
    monkeypatch.setattr(provider_module, "MASTER_PATH", master)
    return {
        "master_path": master,
        "history_path": history_path,
        "latest_status_path": latest_status,
        "folder": folder,
    }


def fake_salesforce(rows: list[dict] | None = None) -> MagicMock:
    """report.get() が Table を返す Salesforce クライアント。"""
    values = ROWS if rows is None else rows
    client = MagicMock()
    client.__enter__.return_value.report.get.return_value = Table(["名前", "金額"], values)
    site = MagicMock(return_value=client)
    return site


class TestLoadMaster:
    """管理表の読み取りと検証。"""

    def test_reads_rows_and_extracts_report_id(self, paths):
        entries = load_master(paths["master_path"])
        assert list(entries) == ["1001", "1002", "1003"]
        # Salesforce のレポート ID は URL から取り出す（人には入力させない）
        assert entries["1001"].report_id == "00O5g00000ABCDE"
        # `1001` だけが `download_scheduled()` の対象（=有効）
        assert entries["1001"].enabled
        assert not entries["1002"].enabled
        assert not entries["1003"].enabled

    def test_blank_rows_are_skipped(self, tmp_path):
        master = make_master(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(tmp_path),
                "○",
                "",
            ], [None] * 9],

        )
        assert list(load_master(master)) == ["1001"]

    def test_duplicate_key_raises(self, tmp_path):
        master = make_master(
            tmp_path / "管理表.xlsx",
            [
                [
                    "1001",
                    "営業事務グループ",
                    "山田",
                    "顧客一覧",
                    URL_A,
                    "CSV",
                    str(tmp_path),
                    "○",
                    "",
                ],

                [
                    "1001",
                    "別の部署",
                    "別の担当",
                    "別の名前",
                    URL_B,
                    "Excel",
                    str(tmp_path),
                    "○",
                    "",
                ],

            ],
        )
        with pytest.raises(MasterDuplicateValueError):
            load_master(master)

    def test_url_without_report_id_raises_with_row_number(self, tmp_path):
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "営業事務グループ", "山田", "顧客一覧", "https://example.com/", "CSV", str(tmp_path), "○", ""]],  # noqa: E501
        )
        with pytest.raises(InvalidReportURLError) as e:
            load_master(master)
        assert "1001" in str(e.value)  # 行番号ではなく管理番号で示す（空行があるとズレるため）

    def test_unknown_output_format_raises(self, tmp_path):
        """`出力形式`に想定外の値を入れるとエラー。"""
        master = make_master(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "PDF",
                str(tmp_path),
                "○",
                "",
            ]],

        )
        with pytest.raises(MasterRowValueError) as e:
            load_master(master)
        assert "出力形式" in str(e.value)

    def test_blank_output_format_raises(self, tmp_path):
        """`出力形式`は既定値なし。空欄のまま読み込むとエラー。"""
        master = make_master(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "",
                str(tmp_path),
                "○",
                "",
            ]],

        )
        with pytest.raises(MasterRowValueError) as e:
            load_master(master)
        assert "出力形式" in str(e.value)

    def test_non_numeric_key_is_allowed(self, tmp_path):
        """str 型の `key` なので、文字列の ID もそのまま使える。"""
        master = make_master(
            tmp_path / "管理表.xlsx",
            [[
                "A001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(tmp_path),
                "○",
                "",
            ]],

        )
        assert list(load_master(master)) == ["A001"]


class TestSharedReportIds:
    """同じ Salesforce レポートを複数の管理番号が指していることの検出。"""

    def test_detects_reports_used_by_multiple_keys(self, paths):
        entries = load_master(paths["master_path"])
        shared = shared_report_ids(entries)
        # "1002" と "1003" が同じ URL（＝同じレポート）を指している
        assert shared == {"00O5g00000FGHIJ": ["1002", "1003"]}

    def test_unique_reports_are_not_listed(self, tmp_path):
        master = make_master(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(tmp_path),
                "○",
                "",
            ]],

        )
        assert shared_report_ids(load_master(master)) == {}


class TestDownloadScheduledRecord:
    """`download_scheduled()` で 1 件のレポートを取得すると、ファイルと履歴が残る。"""

    def test_saves_file_with_csv_extension_for_csv_output_format(self, paths, monkeypatch):
        """`出力形式=CSV` のレポートは `.csv` で保存される。

        `download_scheduled()` は管理表の全有効件を処理するので、`paths` fixture の
        中では `1001`（=CSV）と `1002`（=Excel）が同時に取れる。CSV 側だけ
        検証するため、`.csv` のファイルが含まれていることを確認する。
        """
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            saved = download_scheduled("案件集計")
        csv_paths = [path for path in saved if path.suffix == ".csv"]
        assert len(csv_paths) == 1
        assert csv_paths[0].is_file()
        # CSV として読み戻せる
        with CSV(csv_paths[0], read_only=True) as csv_file:
            assert csv_file.read().read_rows() == ROWS

    def test_saves_file_with_xlsx_extension_for_excel_output_format(
        self, paths, monkeypatch, tmp_path
    ):
        """`出力形式=Excel` のレポートは `.xlsx` で保存される。

        `paths` fixture では `1002` を無効にしてある（`download_scheduled()` の
        対象を `1001` のみに絞るため）。Excel 出力の検証は、専用の管理表を
        作って別途検証する。
        """
        from comken.toolbox.excel import Excel as ExcelForMaster

        folder = tmp_path / "Excel保存先"
        folder.mkdir()
        master = tmp_path / "Excel管理表.xlsx"
        rows = [[
            "1002",
            "経理グループ",
            "佐藤",
            "売上実績",
            URL_B,
            "Excel",
            str(folder),
            "○",
            "",
        ]]

        with ExcelForMaster(master) as book:
            book.create_data_sheet("管理表").create_table(
                "管理表", Table(HEADERS, [dict(zip(HEADERS, row, strict=True)) for row in rows])
            )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")

        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            saved = download_scheduled()
        assert len(saved) == 1
        assert saved[0].suffix == ".xlsx"
        assert saved[0].is_file()
        # Excel として読み戻せる
        from openpyxl import load_workbook

        book = load_workbook(saved[0])
        sheet = book["PY_Report"]
        rows = list(sheet.iter_rows(values_only=True))
        # 1 行目は見出し、2 行目以降にデータ
        assert rows[0] == ("名前", "金額")
        assert rows[1] == ("山田", "100")

    def test_csv_path_is_accessible_after_construction(self, paths):
        with CSV(paths["history_path"]) as csv_file:
            assert csv_file.path == paths["history_path"]

    def test_fetches_again_even_if_already_downloaded_today(self, paths):
        """`paths` fixture の管理表には「スケジュール」シートが無い。

        スケジュール行が無いレポートは ``downloaded_today()`` ベースで
        1 日 1 回までに制限される（後方互換）。よって 2 回目はスキップされ、
        Salesforce へ問い合わせない。
        """
        site = fake_salesforce()
        with patch("comken.services.salesforce_downloader.service.site_for", return_value=site):
            download_scheduled()
            download_scheduled()
        # 1 回目だけ取得される。「時刻付き保管ファイル」1 件 + 「日次キャッシュ」
        # 1 件で合計 2 件
        saved = list(paths["folder"].glob("1001_*.csv"))
        assert len(saved) == 2
        # `report.get()` は 1 回しか呼ばれない（2 回目はスキップ）
        assert site.return_value.__enter__.return_value.report.get.call_count == 1

    def test_existing_collision_is_not_overwritten(self, paths, monkeypatch):
        collision = paths["folder"] / "1001_顧客一覧_fixed.csv"
        collision.write_text("既存", encoding="utf-8")
        monkeypatch.setattr(service_module, "file_path_of", lambda unused: collision)
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            download_scheduled()
        assert collision.read_text(encoding="utf-8") == "既存"
        # もう1つ作られたファイル = 衝突回避で _1 が付いたファイル
        archive = paths["folder"] / "1001_顧客一覧_fixed_1.csv"
        assert archive.exists()

    def test_reserve_path_raises_when_all_sequential_names_are_taken(self, paths, monkeypatch):
        """連番の上限に達したら ``ReportReservePathLimitError`` を内側で出し、
        `download_scheduled()` がそれを捕捉して ``ScheduledDownloadFailedError`` に変換する。

        共有サーバーの権限・同期の異常で ``FileExistsError`` が返り続けると
        既存実装では無限ループになる。 上限を設けて、運用側に気付ける
        メッセージを伴った例外で抜ける。
        """
        # テスト時間短縮のため、上限を小さい値に下げる
        monkeypatch.setattr(service_module, "RESERVE_PATH_LIMIT", 5)
        # ファイル名に日付・時刻が入るので、 file_path_of をモックして固定名にする
        base = paths["folder"] / "1001_顧客一覧.csv"
        monkeypatch.setattr(service_module, "file_path_of", lambda unused: base)

        # ベース名と ``_1`` 〜 ``_4`` までの連番を全部作っておく（計5ファイル）。
        # ``_reserve_path`` は base と ``_1`` 〜 ``_4`` を試して全部 FileExistsError
        # になると、上限に達して例外を上げる
        for sequence in range(5):
            candidate = base if sequence == 0 else paths["folder"] / f"1001_顧客一覧_{sequence}.csv"

            candidate.write_text("埋まり", encoding="utf-8")

        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            pytest.raises(ScheduledDownloadFailedError) as caught,
        ):
            download_scheduled()

        # メッセージに管理番号・上限値・保存先候補のどれかが含まれていれば、 運用側が
        # 「どこで何が起きているか」を追える
        assert "1001" in str(caught.value)

    def test_unregistered_key_does_not_affect_run(self, paths, monkeypatch, tmp_path):
        """管理表に無い管理番号は download_scheduled() の対象外（=直接指定の概念なし）。
        ここでは「1001 だけ取れて、空の 9999 を要求しても何も起きない」ことを確認するため、
        1001 が取得できることをもって対象外になっていることを示す。
        """
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            saved = download_scheduled()
        assert [path.name.split("_")[0] for path in saved] == ["1001"]

    def test_disabled_report_is_skipped(self, paths):
        """`1003` は「無効」なので取得対象外（periodic から除外される）。"""
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            saved = download_scheduled()
        assert all(path.name.split("_")[0] != "1003" for path in saved)

    def test_empty_report_raises_and_saves_nothing(self, paths):
        """0 件あり=× のレポートで 0 行だと、`ScheduledDownloadFailedError` で全体が失敗する。"""
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce([]),
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()
        assert list(paths["folder"].glob("*.csv")) == []

    def test_missing_folder_raises_and_is_not_created(self, tmp_path, monkeypatch):
        missing = tmp_path / "無いフォルダ"
        master = make_master(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(missing),
                "○",
                "",
            ]],

        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()
        assert not missing.exists()

    def test_no_temporary_file_is_left_behind(self, paths):
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            download_scheduled()
        assert list(paths["folder"].glob("~*")) == []


class TestHistory:
    """履歴には成否も、誰が要求したかも残る。"""

    def test_success_is_recorded_with_project_and_counts(self, paths):
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            download_scheduled("案件集計")
        row = _history_rows(paths)[-1]
        assert row["管理番号"] == "1001"
        assert row["プロジェクト"] == "案件集計"
        assert row["成否"] == "成功"
        assert row["Salesforce取得結果"] == "成功"
        assert row["保存結果"] == "成功"
        assert row["取得件数"] == "2"
        assert row["レポートID"] == "00O5g00000ABCDE"
        assert row["エラーコード"] == ""

    def test_failure_is_recorded(self, paths):
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce([]),
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled("案件集計")
        row = _history_rows(paths)[-1]
        assert row["成否"] == "失敗"
        # 0 行でも Salesforce への問い合わせは成功している点が重要
        assert row["Salesforce取得結果"] == "成功"
        assert row["保存結果"] == ""
        assert row["エラーコード"] == "EmptyReportError"
        assert "0 行" in row["エラー内容"]

    def test_downloaded_today_counts_after_scheduled_run(self, paths):
        """`download_scheduled()` で取った記録は `downloaded_today()` で拾える。"""
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            download_scheduled()
        assert history.downloaded_today(paths["history_path"], "1001")

    def test_missing_history_file_is_not_downloaded(self, tmp_path):
        assert not history.downloaded_today(tmp_path / "無い.csv", "1001")

    # ── 履歴の5ケース（4. の表に対応する個別テスト）────────────────────
    def test_history_when_folder_is_missing(self, tmp_path, monkeypatch):
        """保存先フォルダが無い → 成否=失敗 / Salesforce取得結果=空 / 保存結果=空 /
        エラーコード=ReportFolderNotFoundError。"""
        missing = tmp_path / "無いフォルダ"
        master = make_master(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(missing),
                "○",
                "",
            ]],

        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()
        with CSV(history_path) as csv_file:
            rows = csv_file.read()
            assert len(rows) == 1
            row = rows[0]
            assert row["成否"] == "失敗"
            assert row["Salesforce取得結果"] == ""
            assert row["保存結果"] == ""
            assert row["エラーコード"] == "ReportFolderNotFoundError"

    def test_history_when_salesforce_call_fails(self, tmp_path, monkeypatch):
        """Salesforce への問い合わせが失敗 → 成否=失敗 / Salesforce取得結果=失敗 /
        保存結果=空 / エラーコード=送出された例外クラス名。

        `download_scheduled()` は `ComkenError` を捕捉して `ScheduledDownloadFailedError`
        に変換するため、テストではそちらを期待する。履歴には元の例外クラス名が残る。
        """
        from comken.exceptions import SalesforceRequestError

        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "",
            ]],

        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        site = MagicMock()
        client = MagicMock()
        client.__enter__.return_value.report.get.side_effect = SalesforceRequestError(
            "GET", "/report", 500, "boom"
        )
        site.return_value = client

        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=site,
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()
        with CSV(history_path) as csv_file:
            rows = csv_file.read()
            assert len(rows) == 1
            row = rows[0]
            assert row["成否"] == "失敗"
            assert row["Salesforce取得結果"] == "失敗"
            assert row["保存結果"] == ""
            assert row["エラーコード"] == "SalesforceRequestError"

    def test_history_when_report_is_empty(self, paths):
        """取得できたが 0 行だった → 成否=失敗 / Salesforce取得結果=成功 /
        保存結果=空 / エラーコード=EmptyReportError（通信は成功していて中身が空、という区別）。"""
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce([]),
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()
        row = _history_rows(paths)[-1]
        assert row["成否"] == "失敗"
        assert row["Salesforce取得結果"] == "成功"
        assert row["保存結果"] == ""
        assert row["エラーコード"] == "EmptyReportError"

    def test_history_when_csv_write_fails(self, tmp_path, monkeypatch):
        """Salesforce 取得は成功したが CSV 書き込みが失敗 → 成否=失敗 /
        Salesforce取得結果=成功 / 保存結果=失敗 / エラーコード=送出された例外クラス名。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "",
            ]],

        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        write_error = OSError("書き込み失敗")

        def _raise_write(path, fieldnames, *args, **kwargs):
            raise write_error

        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            patch(
                "comken.services.salesforce_downloader.service.CSV._write",
                side_effect=write_error,
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()
        with CSV(history_path) as csv_file:
            rows = csv_file.read()
            assert len(rows) == 1
            row = rows[0]
            assert row["成否"] == "失敗"
            assert row["Salesforce取得結果"] == "成功"
            assert row["保存結果"] == "失敗"
            assert row["エラーコード"] == "OSError"

    # ── 「原因区分」列（4区分 + 成功時の空文字）─────────────────────
    def test_cause_is_blank_on_success(self, paths):
        """成功時は原因区分が空文字。"""
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            download_scheduled()
        row = _history_rows(paths)[-1]
        assert row["成否"] == "成功"
        assert row["原因区分"] == ""

    def test_cause_is_config_when_folder_is_missing(self, tmp_path, monkeypatch):
        """保存先フォルダが無い → 「設定」（取得段階に入る前に落ちる）。"""
        missing = tmp_path / "無いフォルダ"
        master = make_master(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(missing),
                "○",
                "",
            ]],

        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()
        with CSV(history_path) as csv_file:
            row = csv_file.read()[-1]
            assert row["原因区分"] == "設定"

    def test_cause_is_salesforce_when_request_fails(self, tmp_path, monkeypatch):
        """Salesforce への問い合わせが失敗 → 「Salesforce」。"""
        from comken.exceptions import SalesforceRequestError

        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "",
            ]],

        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        site = MagicMock()
        client = MagicMock()
        client.__enter__.return_value.report.get.side_effect = SalesforceRequestError(
            "GET", "/report", 500, "boom"
        )
        site.return_value = client

        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=site,
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()
        with CSV(history_path) as csv_file:
            row = csv_file.read()[-1]
            assert row["原因区分"] == "Salesforce"

    def test_cause_is_empty_data_when_report_is_empty(self, paths):
        """取得できたが 0 行 → 「データなし」（取得は成功・保存未到達を一意に指す区分）。

        この管理表には `0件あり` 列が無いので `×` 既定扱いで `EmptyReportError` が送出される。
        段階は「取得成功 → 保存に進まず 0 行で失敗」になるため、4 区分だった頃の
        `Salesforce` から、新仕様の `データなし` に変わった。
        """
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce([]),
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()
        row = _history_rows(paths)[-1]
        assert row["原因区分"] == "データなし"

    def test_cause_is_file_when_csv_write_fails(self, tmp_path, monkeypatch):
        """CSV 書き込みが OSError で失敗 → 「ファイル」（共有サーバー・権限）。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "",
            ]],

        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        write_error = OSError("書き込み失敗")
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            patch(
                "comken.services.salesforce_downloader.service.CSV._write",
                side_effect=write_error,
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()
        with CSV(history_path) as csv_file:
            row = csv_file.read()[-1]
            assert row["原因区分"] == "ファイル"

    def test_cause_is_program_when_unexpected_error_raises(self, tmp_path, monkeypatch):
        """_fetch() が TypeError を投げる（comken 側のバグ想定）→ 「プログラム」。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "",
            ]],

        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        site = MagicMock()
        client = MagicMock()
        client.__enter__.return_value.report.get.side_effect = TypeError("想定外")
        site.return_value = client

        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=site,
            ),
            pytest.raises(TypeError),
        ):
            download_scheduled()
        with CSV(history_path) as csv_file:
            row = csv_file.read()[-1]
            assert row["原因区分"] == "プログラム"


class TestDownloadScheduled:
    """定期取得は「有効」なものだけを対象にする。"""

    def test_only_enabled_reports_are_downloaded(self, paths):
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            saved = download_scheduled()
        # `paths` fixture では `1001`（=CSV・有効）だけが対象
        keys = sorted(path.name.split("_")[0] for path in saved)
        assert keys == ["1001"]

    def test_one_failure_does_not_stop_the_rest(self, tmp_path, monkeypatch):
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [
                [
                    "1001",
                    "営業事務グループ",
                    "山田",
                    "落ちる方",
                    URL_A,
                    "CSV",
                    str(tmp_path / "無い"),
                    "○",
                    "",
                ],

                [
                    "1002",
                    "経理グループ",
                    "佐藤",
                    "通る方",
                    URL_B,
                    "CSV",
                    str(folder),
                    "○",
                    "",
                ],

            ],
        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()
        # "1001" で失敗しても "1002" は保存されている（続けたうえで最後に知らせる）。
        # `*.csv` には時刻付き保管ファイル（1件）と日次キャッシュ（1件）があるので
        # `1002` が2件拾われる
        keys = sorted(path.name.split("_")[0] for path in folder.glob("*.csv"))
        assert keys == ["1002", "1002"]

    def test_os_error_does_not_stop_the_rest(self, tmp_path, monkeypatch):
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [
                [
                    "1001",
                    "営業事務グループ",
                    "山田",
                    "書込失敗",
                    URL_A,
                    "CSV",
                    str(folder),
                    "○",
                    "",
                ],

                [
                    "1002",
                    "経理グループ",
                    "佐藤",
                    "取得成功",
                    URL_B,
                    "CSV",
                    str(folder),
                    "○",
                    "",
                ],

            ],
        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")
        original_write_csv = service_module._write_csv

        def fail_first_write(path, rows):
            if path.name.startswith("1001_"):
                raise OSError("共有サーバーへ書き込めません")
            original_write_csv(path, rows)

        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            patch.object(service_module, "_write_csv", side_effect=fail_first_write),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()

        keys = sorted(path.name.split("_")[0] for path in folder.glob("*.csv"))
        assert keys == ["1002", "1002"]

    def test_cache_update_failure_keeps_archive_and_marks_failure(self, paths):
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            patch.object(
                service_module,
                "_update_daily_cache",
                side_effect=OSError("キャッシュ更新失敗"),
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()

        assert len(list(paths["folder"].glob("1001_*.csv"))) == 1
        row = _history_rows(paths)[-1]
        assert row["成否"] == "失敗"
        assert row["保存結果"] == "失敗"

    def test_unexpected_error_stops_the_run_immediately(self, tmp_path, monkeypatch):
        """想定外（`TypeError` など）はその場で抜ける。`ScheduledDownloadFailedError` には
        変換しない（非エンジニアが「もう一度実行してみる」を繰り返すだけになるため）。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [
                [
                    "1001",
                    "営業事務グループ",
                    "山田",
                    "想定外",
                    URL_A,
                    "CSV",
                    str(folder),
                    "○",
                    "",
                ],

                [
                    "1002",
                    "経理グループ",
                    "佐藤",
                    "通る方",
                    URL_B,
                    "CSV",
                    str(folder),
                    "○",
                    "",
                ],

            ],
        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")

        site = MagicMock()
        client = MagicMock()
        # 1001 のレポートを取るときだけ想定外を投げる
        client.__enter__.return_value.report.get.side_effect = [
            TypeError("想定外"),
            ROWS,
        ]
        site.return_value = client

        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=site,
            ),
            pytest.raises(TypeError),
        ):
            download_scheduled()
        # 想定外で止めたので、2件目は保存されない
        # （続けた結果の ScheduledDownloadFailedError ではないことを確認）
        assert list(folder.glob("*.csv")) == []

    def test_records_the_trigger_as_scheduled(self, paths):
        """`download_scheduled()` で取った記録は履歴に `プロジェクト` 名で残る。"""
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            download_scheduled("定期実行")
        assert _history_rows(paths)[-1]["プロジェクト"] == "定期実行"

    def test_writes_latest_status_after_scheduled_run(self, paths):
        """定期取得のあとに、管理表×履歴の最新行をまとめた Excel を作る。"""
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            download_scheduled()
        assert paths["latest_status_path"].is_file()

    def test_latest_status_failure_does_not_affect_outcome(self, paths, monkeypatch):
        """write_latest_status() が失敗しても、定期取得の成功／失敗判定に影響させない。"""
        from comken.services.salesforce_downloader import service as service_module_inner

        def boom(*args, **kwargs):
            raise OSError("最新ステータスを書けません")

        # `service.py` は `from latest_status import write_latest_status` で
        # ローカル束縛しているので、呼び出し側を直接差し替える
        monkeypatch.setattr(service_module_inner, "write_latest_status", boom)
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            saved = download_scheduled()  # 例外にならず、保存できたファイルが返る
        assert [path.name.split("_")[0] for path in saved] == ["1001"]
        # 失敗しても本体結果（成功件数・失敗件数）には反映されない
        assert not paths["latest_status_path"].exists()

    # ── スケジュール管理表と組み合わせた判定 ──────────────────────────────
    def test_schedule_sheet_missing_keeps_backward_compatible_behavior(
        self, tmp_path, monkeypatch
    ):
        """「スケジュール」シートが無い管理表は、これまで通り毎回対象になる（後方互換）。

        この機能追加を境に既存のレポートが突然取得されなくなる事故を防ぐ。
        """
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "",
            ]],

        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")
        # 「今」を月曜 09:00 に固定してもスケジュール判定には影響しない
        fixed_now = dt.datetime(2026, 1, 5, 9, 0)  # noqa: DTZ001 — テスト用に意図的に固定した tz-naive な datetime
        monkeypatch.setattr(service_module, "clock_now", lambda: fixed_now)

        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            saved = download_scheduled()
        assert [path.name.split("_")[0] for path in saved] == ["1001"]

    def test_schedule_rule_not_due_excludes_report(self, tmp_path, monkeypatch):
        """スケジュール行が「今は要らない (False)」を返したレポートは対象から外れる。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        # 水曜 12:00 固定 → 「毎週・月曜・09:00」は曜日不一致で False
        fixed_now = dt.datetime(2026, 1, 7, 12, 0)  # noqa: DTZ001 — テスト用に意図的に固定した tz-naive な datetime  # noqa: DTZ001 — テスト用に意図的に固定した tz-naive な datetime
        # 「スケジュール」シートを足した管理表を作る（既存の make_master を使うため、
        # 直接 Excel を組み立てる）
        master = make_master_with_schedule(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "曜日外し",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "",
            ]],

            schedule_rows=[
                ["S001", "1001", "毎週", "09:00", "月", "取得しない", "○"],
            ],
        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")
        monkeypatch.setattr(service_module, "clock_now", lambda: fixed_now)
        _patch_default_calendar(monkeypatch, holidays=set())

        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            saved = download_scheduled()
        # スケジュールの判定で外れるので、保存されない
        assert saved == []
        assert list(folder.glob("*.csv")) == []

    def test_schedule_rule_due_includes_report(self, tmp_path, monkeypatch):
        """スケジュール行が「今は要る (True)」を返したレポートは対象に入る。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        # 水曜 12:00 固定 → 「毎週・水曜・09:00」は `now.time() >= run_time` で True
        fixed_now = dt.datetime(2026, 1, 7, 12, 0)  # noqa: DTZ001 — テスト用に意図的に固定した tz-naive な datetime
        master = make_master_with_schedule(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "曜日一致",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "",
            ]],

            schedule_rows=[
                ["S001", "1001", "毎週", "09:00", "水", "取得しない", "○"],
            ],
        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")
        monkeypatch.setattr(service_module, "clock_now", lambda: fixed_now)
        _patch_default_calendar(monkeypatch, holidays=set())

        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            saved = download_scheduled()
        assert [path.name.split("_")[0] for path in saved] == ["1001"]

    def test_multiple_schedule_rules_use_or_semantics(self, tmp_path, monkeypatch):
        """同じレポートに複数行があり、どれか1つでも is_due() なら対象に入る（OR 条件）。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        # 水曜 12:00 固定 → 1行目（月曜）は曜日外れ、2行目（水曜）は一致
        fixed_now = dt.datetime(2026, 1, 7, 12, 0)  # noqa: DTZ001 — テスト用に意図的に固定した tz-naive な datetime
        master = make_master_with_schedule(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "OR判定",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "",
            ]],

            schedule_rows=[
                ["S001", "1001", "毎週", "09:00", "月", "取得しない", "○"],
                ["S002", "1001", "毎週", "09:00", "水", "取得しない", "○"],
            ],
        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")
        monkeypatch.setattr(service_module, "clock_now", lambda: fixed_now)
        _patch_default_calendar(monkeypatch, holidays=set())

        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=fake_salesforce()
        ):
            saved = download_scheduled()
        assert [path.name.split("_")[0] for path in saved] == ["1001"]


# ── スケジュール単位の重複実行防止（今回の機能の核心）───────────────────
class TestScheduleDedup:
    """同じスケジュール行を同日に 2 回 `download_scheduled()` で呼んでも、
    2 回目はスキップされる（履歴ベースのdedup）。"""

    def test_second_call_for_same_schedule_key_is_skipped(
        self, tmp_path, monkeypatch
    ):
        """同じスケジュール行（=同じ schedule_key）に紐付くレポートは、当日中に
        成功履歴があれば 2 回目の `download_scheduled()` で再取得されない。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        fixed_now = dt.datetime(2026, 1, 7, 12, 0)  # noqa: DTZ001 — テスト用に意図的に固定した tz-naive な datetime
        master = make_master_with_schedule(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "曜日一致",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "",
            ]],

            schedule_rows=[
                ["S001", "1001", "毎週", "09:00", "水", "取得しない", "○"],
            ],
        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")
        monkeypatch.setattr(service_module, "clock_now", lambda: fixed_now)
        _patch_default_calendar(monkeypatch, holidays=set())

        site = fake_salesforce()
        with patch(
            "comken.services.salesforce_downloader.service.site_for", return_value=site
        ):
            download_scheduled()
            download_scheduled()
        # 1 回目だけ Salesforce へ問い合わせる（2 回目は履歴を見てスキップ）
        assert site.return_value.__enter__.return_value.report.get.call_count == 1
        # 1 回目だけ取得されるので、保存ファイルは「時刻付き保管」+「日次キャッシュ」
        # の 2 件だけ
        saved = list(folder.glob("1001_*.csv"))
        assert len(saved) == 2
        # 履歴の「スケジュールキー」列に、根拠のキーが記録されている
        with CSV(tmp_path / "履歴.csv") as csv_file:
            rows = csv_file.read()
        assert len(rows) == 1
        assert rows[0]["管理番号"] == "1001"
        assert rows[0]["スケジュールキー"] == "S001"
        assert rows[0]["成否"] == "成功"

    def test_second_call_does_not_skip_when_first_failed(
        self, tmp_path, monkeypatch
    ):
        """1 回目が失敗（保存失敗など）した場合は、成功履歴が残らないため、
        2 回目は再取得される（失敗を繰り返さない運用にするため）。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        fixed_now = dt.datetime(2026, 1, 7, 12, 0)  # noqa: DTZ001 — テスト用に意図的に固定した tz-naive な datetime
        master = make_master_with_schedule(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "曜日一致",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "",
            ]],

            schedule_rows=[
                ["S001", "1001", "毎週", "09:00", "水", "取得しない", "○"],
            ],
        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")
        monkeypatch.setattr(service_module, "clock_now", lambda: fixed_now)
        _patch_default_calendar(monkeypatch, holidays=set())

        # 1 回目は OSError で保存失敗、2 回目は成功する
        original_write_csv = service_module._write_csv
        attempts = {"count": 0}

        def fail_first(path, table):
            attempts["count"] += 1
            if attempts["count"] == 1:
                raise OSError("共有サーバーへ書き込めません")
            original_write_csv(path, table)

        site = fake_salesforce()
        first_site = fake_salesforce()
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=first_site,
            ),
            patch.object(
                service_module, "_write_csv", side_effect=fail_first
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()
        # 1 回目は失敗 → 2 回目を呼ぶ
        with patch(
            "comken.services.salesforce_downloader.service.site_for",
            return_value=site,
        ):
            download_scheduled()
        # 失敗のあと再呼び出ししたので、Salesforce への問い合わせは 2 回
        assert first_site.return_value.__enter__.return_value.report.get.call_count == 1
        assert site.return_value.__enter__.return_value.report.get.call_count == 1

    def test_failure_row_records_schedule_key(self, tmp_path, monkeypatch):
        """失敗時の履歴行にも `schedule_key` が記録される（成功時と非対称に
        空文字のままだと、後から「このスケジュール行がいつ失敗したか」を
        追えなくなるため）。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        fixed_now = dt.datetime(2026, 1, 7, 12, 0)  # noqa: DTZ001 — テスト用に意図的に固定した tz-naive な datetime
        master = make_master_with_schedule(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "曜日一致",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "",
            ]],

            schedule_rows=[
                ["S001", "1001", "毎週", "09:00", "水", "取得しない", "○"],
            ],
        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")
        monkeypatch.setattr(service_module, "clock_now", lambda: fixed_now)
        _patch_default_calendar(monkeypatch, holidays=set())

        # 保存段階で OSError を起こして失敗させる（_Attempt.record_failure() の経路を通る）
        def fail(path, table):
            raise OSError("共有サーバーへ書き込めません")

        site = fake_salesforce()
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=site,
            ),
            patch.object(service_module, "_write_csv", side_effect=fail),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()

        # 失敗時の履歴行の「スケジュールキー」列に、空でない根拠キーが入っている
        with CSV(tmp_path / "履歴.csv") as csv_file:
            rows = csv_file.read()
        assert len(rows) == 1
        assert rows[0]["管理番号"] == "1001"
        assert rows[0]["成否"] == "失敗"
        assert rows[0]["スケジュールキー"] == "S001"

    def test_dedup_is_per_schedule_key_not_per_report(
        self, tmp_path, monkeypatch
    ):
        """同じレポートでも別スケジュールキーが当たって成功した場合、新キーで
        is_due=True になれば重複として防がれる。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        fixed_now = dt.datetime(2026, 1, 7, 12, 0)  # noqa: DTZ001 — テスト用に意図的に固定した tz-naive な datetime
        master = make_master_with_schedule(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "複数スケジュール",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "",
            ]],

            schedule_rows=[
                # 月曜 09:00 と水曜 09:00 の 2 行で `1001` を取得する設定
                ["S_MON", "1001", "毎週", "09:00", "月", "取得しない", "○"],
                ["S_WED", "1001", "毎週", "09:00", "水", "取得しない", "○"],
            ],
        )
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", tmp_path / "履歴.csv")
        monkeypatch.setattr(service_module, "clock_now", lambda: fixed_now)
        _patch_default_calendar(monkeypatch, holidays=set())

        # 1 回目: 水曜 12:00 → S_WED が is_due=True、ただし履歴に何もないので取得
        first_site = fake_salesforce()
        with patch(
            "comken.services.salesforce_downloader.service.site_for",
            return_value=first_site,
        ):
            download_scheduled()
        # 履歴には S_WED が記録されている
        with CSV(tmp_path / "履歴.csv") as csv_file:
            first_rows = csv_file.read()
        assert len(first_rows) == 1
        assert first_rows[0]["スケジュールキー"] == "S_WED"

        # 時刻を進めて月曜 12:00 にする → S_MON が is_due=True になる。
        # スケジュールキーが違うので S_MON の履歴は無いため、S_MON で取得される
        # （=重複ではなく別キー）
        monday_now = dt.datetime(2026, 1, 5, 12, 0)  # noqa: DTZ001 — テスト用に意図的に固定した tz-naive な datetime
        monkeypatch.setattr(service_module, "clock_now", lambda: monday_now)
        second_site = fake_salesforce()
        with patch(
            "comken.services.salesforce_downloader.service.site_for",
            return_value=second_site,
        ):
            saved2 = download_scheduled()
        assert [path.name.split("_")[0] for path in saved2] == ["1001"]
        with CSV(tmp_path / "履歴.csv") as csv_file:
            rows = csv_file.read()
        # S_MON の成功履歴が増える
        keys = sorted(row["スケジュールキー"] for row in rows)
        assert keys == ["S_MON", "S_WED"]


def _patch_default_calendar(monkeypatch: pytest.MonkeyPatch, *, holidays: set) -> None:
    """``service.default_calendar()`` を固定の祝日セットを持つ偽物に差し替える。"""

    class _FakeCalendar:
        def is_holiday(self, date: object) -> bool:
            return date in holidays

    monkeypatch.setattr(service_module, "default_calendar", lambda: _FakeCalendar())


def make_master_with_schedule(
    path: Path,
    master_rows: list[list],
    *,
    schedule_rows: list[list],
) -> Path:
    """レポート管理表 + スケジュールシートの2シート構成のブックを作る。

    ``tests/test_schedule_load.py`` 側でも同じ関数が必要だが、fixture として
    共有するのが煩雑なので、サービス側のテストはローカルに持つ。
    """
    master_headers = [
        "ID",
        "グループ名",
        "担当者",
        "概要",
        "Salesforce URL",
        "出力形式",
        "保存先",
        "有効",
        "備考",
    ]
    schedule_headers = [
        "スケジュールキー",
        "レポートキー",
        "取得頻度",
        "取得時刻",
        "曜日",
        "祝日対応",
        "有効",
    ]
    master_table_rows = [dict(zip(master_headers, row, strict=True)) for row in master_rows]
    schedule_table_rows = [
        dict(zip(schedule_headers, row, strict=True)) for row in schedule_rows
    ]
    with Excel(path) as book:
        book.create_data_sheet("管理表").create_table(
            "管理表", Table(master_headers, master_table_rows)
        )
        book.create_data_sheet("スケジュール").create_table(
            "スケジュール", Table(schedule_headers, schedule_table_rows)
        )
    return path


def _history_rows(paths: dict) -> list[dict]:
    with CSV(paths["history_path"]) as csv_file:
        return csv_file.read()


class TestRequiredHistory:
    """履歴が書けない場合は、取得結果だけを成功として返さない。"""

    def test_history_write_failure_stops_download(self, paths):
        """`HistoryWriteError` は `ScheduledDownloadFailedError` に変換されて返る。"""
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce(),
            ),
            patch.object(history, "_append", side_effect=OSError("履歴書込み失敗")),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()

    def test_original_failure_remains_in_history_error(self, paths):
        """`EmptyReportError` が起きて履歴も書けなかった場合、エラーは
        `ScheduledDownloadFailedError` で伝搬し、メッセージに元の失敗が含まれる。"""
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce([]),
            ),
            patch.object(history, "_append", side_effect=OSError("履歴書込み失敗")),
            pytest.raises(ScheduledDownloadFailedError) as caught,
        ):
            download_scheduled()
        # `ScheduledDownloadFailedError` の元例外 (`__cause__`) が
        # `HistoryWriteError` で、その中に元の `EmptyReportError` が連鎖している
        assert isinstance(caught.value.__cause__, HistoryWriteError)
        original = caught.value.__cause__.__cause__
        assert original is not None
        assert "0 行" in str(original)


# ── 0件あり / 0 行の扱い ─────────────────────────────────────────────
class TestAllowEmpty:
    """管理表の「0件あり」列で、0 行のとき失敗にするか正常終了にするかを選ぶ。"""

    def test_empty_report_with_allow_empty_no_raises_and_records_empty_data_cause(
        self, tmp_path, monkeypatch
    ):
        """1. `0件あり` が `×` で 0 行 → `EmptyReportError`、ファイルができない、
        履歴の `原因区分` が `データなし`。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master_with_allow_empty(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "×",
                "",
            ]],

        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce([]),
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()

        # ファイルは作られない
        assert list(folder.glob("*.csv")) == []
        # 履歴には `データなし` が残る（取得成功・保存未到達の組合せのみ取り得る）
        with CSV(history_path) as csv_file:
            row = csv_file.read()[-1]
            assert row["成否"] == "失敗"
            assert row["Salesforce取得結果"] == "成功"
            assert row["保存結果"] == ""
            assert row["エラーコード"] == "EmptyReportError"
            assert row["原因区分"] == "データなし"

    def test_empty_report_with_allow_empty_yes_succeeds_and_writes_empty_file(
        self, tmp_path, monkeypatch
    ):
        """2. `0件あり` が `○` で 0 行 → 例外にならない、空のファイルができる、
        履歴が `成否=成功` / `取得件数=0` / `原因区分` が空。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master_with_allow_empty(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "○",
                "",
            ]],

        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        with patch(
            "comken.services.salesforce_downloader.service.site_for",
            return_value=fake_salesforce([]),
        ):
            download_scheduled()  # 例外にならない

        # 0 行でも Salesforce の列情報を持つ CSV が作られる。`*.csv` には
        # 時刻付き保管ファイルと日次キャッシュの 2 件が `1001` で作られる
        saved = list(folder.glob("1001_*.csv"))
        assert len(saved) == 2
        with CSV(saved[0], read_only=True) as csv_file:
            assert csv_file.read().columns == ["名前", "金額"]

        # 履歴は成功・取得件数 0・原因区分 空
        with CSV(history_path) as csv_file:
            row = csv_file.read()[-1]
            assert row["成否"] == "成功"
            assert row["Salesforce取得結果"] == "成功"
            assert row["保存結果"] == "成功"
            assert row["取得件数"] == "0"
            assert row["原因区分"] == ""
            assert row["エラーコード"] == ""

    def test_scheduled_empty_report_can_be_received(self, tmp_path, monkeypatch):
        """0件で成功した定期取得は、本日取得済みとして空のまま受け取れる。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master_with_allow_empty(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "○",
                "",
            ]],

        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(_paths_module, "MASTER_PATH", master)
        monkeypatch.setattr(_paths_module, "HISTORY_PATH", history_path)
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)
        monkeypatch.setattr(provider_module, "MASTER_PATH", master)

        with patch(
            "comken.services.salesforce_downloader.service.site_for",
            return_value=fake_salesforce([]),
        ):
            download_scheduled()

        from comken.services.salesforce_downloader import cached_report

        reader = cached_report("1001")
        assert cached_report_path("1001").is_file()
        assert reader.read_rows() == []

    def test_master_without_allow_empty_column_defaults_to_no(self, tmp_path, monkeypatch):
        """4. `0件あり` の列が無い管理表でも読める（既定 `×` として扱われる）。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        # 7 列のまま = `0件あり` 列が無い管理表
        master = make_master(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(folder),
                "○",
                "",
            ]],

        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        entries = load_master(master)
        assert entries["1001"].allow_empty is False  # 既定値

        # 0 行のときは従来どおり `EmptyReportError`（列が無い管理表を
        # 後付けで読む既存プロジェクトを壊さないため）
        with (
            patch(
                "comken.services.salesforce_downloader.service.site_for",
                return_value=fake_salesforce([]),
            ),
            pytest.raises(ScheduledDownloadFailedError),
        ):
            download_scheduled()
        assert list(folder.glob("*.csv")) == []

    def test_invalid_allow_empty_value_raises(self, tmp_path):
        """5. `0件あり` に `○` `×` 以外を書くとエラーになる（choices で弾く）。"""
        master = make_master_with_allow_empty(
            tmp_path / "管理表.xlsx",
            [[
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                URL_A,
                "CSV",
                str(tmp_path),
                "○",
                "△",
                "",
            ]],

        )
        with pytest.raises(MasterRowValueError) as e:
            load_master(master)
        # 列名でユーザーが書き方を直せるよう、見出しをメッセージに残す
        assert "0件あり" in str(e.value)

    def test_download_scheduled_with_allow_empty_yes_does_not_raise_on_zero_rows(
        self, tmp_path, monkeypatch
    ):
        """6. `download_scheduled()` で `0件あり` が `○` のレポートが 0 行でも
        `ScheduledDownloadFailedError` にならない。"""
        folder = tmp_path / "保存先"
        folder.mkdir()
        master = make_master_with_allow_empty(
            tmp_path / "管理表.xlsx",
            [
                [
                    "1001",
                    "営業事務グループ",
                    "山田",
                    "空でもOK",
                    URL_A,
                    "CSV",
                    str(folder),
                    "○",
                    "○",
                    "",
                ],

                [
                    "1002",
                    "経理グループ",
                    "佐藤",
                    "普通のレポート",
                    URL_B,
                    "CSV",
                    str(folder),
                    "○",
                    "×",
                    "",
                ],

            ],
        )
        history_path = tmp_path / "履歴.csv"
        monkeypatch.setattr(service_module, "MASTER_PATH", master)
        monkeypatch.setattr(service_module, "HISTORY_PATH", history_path)

        # "1001" は 0 行、"1002" は通常データを返す
        site = MagicMock()
        client = MagicMock()

        def _run(report_id):
            rows = [] if report_id == "00O5g00000ABCDE" else ROWS
            return Table(["名前", "金額"], rows)

        client.__enter__.return_value.report.get.side_effect = _run
        site.return_value = client

        with patch("comken.services.salesforce_downloader.service.site_for", return_value=site):
            saved = download_scheduled()  # 例外にならない

        # 両方とも保存される（"1001" は空ファイル、"1002" は通常の CSV）
        names = sorted(path.name.split("_")[0] for path in saved)
        assert names == ["1001", "1002"]
        # 履歴を確認: "1001" は成功・0件、"1002" も成功・2件
        with CSV(history_path) as csv_file:
            rows = csv_file.read()
            by_key = {row["管理番号"]: row for row in rows}
            assert by_key["1001"]["成否"] == "成功"
            assert by_key["1001"]["取得件数"] == "0"
            assert by_key["1001"]["原因区分"] == ""
            assert by_key["1002"]["成否"] == "成功"
            assert by_key["1002"]["取得件数"] == "2"


class TestTemplate:
    """管理表の雛形は、そのまま読み込める状態で作られる。"""

    def test_generated_template_can_be_loaded(self, tmp_path):
        """雛形の記入例が、そのまま load_master() を通る（列名の食い違いが起きない）。"""
        path = ReportEntry.create_template(tmp_path / "レポート管理表.xlsx", EXAMPLES)
        entries = load_master(path)
        assert list(entries) == ["1001", "1002"]
        # 「実行方式」列は廃止し、`出力形式` 列が新設計。雛形 1 行目は CSV 例、
        # 2 行目は Excel 例として ``EXAMPLES`` に書き込まれている
        assert entries["1001"].output_format == "CSV"
        assert entries["1002"].output_format == "Excel"

    def test_examples_point_at_different_reports(self, tmp_path):
        """記入例が同じレポートを指していると、check が重複として報告してしまう。"""
        entries = load_master(ReportEntry.create_template(tmp_path / "管理表.xlsx", EXAMPLES))
        assert shared_report_ids(entries) == {}

    def test_guide_sheet_is_included(self, tmp_path):
        """非エンジニアが1枚で分かるよう、記入方法のシートを付ける。"""
        from openpyxl import load_workbook

        path = ReportEntry.create_template(tmp_path / "レポート管理表.xlsx", EXAMPLES)
        assert "記入方法" in load_workbook(path).sheetnames


class TestCommandLine:
    """保守用コマンド（検査）。"""

    def test_check_reports_counts(self, paths, capsys):
        assert cli(["check", str(paths["master_path"])]) == 0
        out = capsys.readouterr().out
        assert "登録 3 件" in out
        assert "00O5g00000FGHIJ" in out  # 同じレポートを指している管理番号を知らせる

    def test_check_returns_failure_for_a_broken_master(self, tmp_path, capsys):
        master = make_master(
            tmp_path / "管理表.xlsx",
            [["1001", "営業事務グループ", "山田", "顧客一覧", "https://example.com/", "CSV", str(tmp_path), "○", ""]],  # noqa: E501
        )
        assert cli(["check", str(master)]) == 1
        assert "エラー:" in capsys.readouterr().err

