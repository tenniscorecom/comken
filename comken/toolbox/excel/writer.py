"""comken/toolbox/excel/writer.py — Excel の書き込み・書式設定・保存を行う入口。"""

# 定義中の ExcelWriter を戻り値の型注釈に使うため、注釈の評価を遅延する。
from __future__ import annotations

import logging
import tempfile
from pathlib import Path
from typing import Self

from openpyxl import Workbook

from comken.core.files.base import FileBase
from comken.core.timer import measure
from comken.exceptions import (
    ExcelSaveNotCompletedError,
    LastSheetDeletionError,
    SheetAlreadyExistsError,
    SheetNotFoundError,
)
from comken.runtime import dry_run_log, is_dry_run
from comken.toolbox.excel.base import ExcelBase
from comken.toolbox.excel.sheet import Sheet

logger = logging.getLogger(__name__)


class ExcelWriter(ExcelBase):
    """Excel ブックの読み取り・書き込み・保存を行うクラス。

    読み取りメソッドも継承しているため、データを読んでから Sheet で
    書き換える処理を1つのブックで完結できる。

    """

    def __init__(
        self,
        path: str | Path,
        data_only: bool = False,
        local_copy_threshold_mb: float = 10,
        headers: list[str] | None = None,
    ) -> None:
        """
        Args:
            path: Excel ファイルのパス。
            data_only: True にすると数式セルのキャッシュ値を読む（read_computed_rows 推奨）。
            local_copy_threshold_mb: この MB 以上のファイルはローカルにコピーしてから開く。
                NAS・ネットワークドライブのファイルが遅い・不安定な場合に有効。
                0 を指定するとローカルコピーを無効化できる。
            headers: ヘッダー行がない Excel の場合に、列名のリストをここで付ける。
                指定すると read_rows_as_dicts() は全行をデータとして読む。
        """
        super().__init__(
            path,
            data_only=data_only,
            read_only=False,
            local_copy_threshold_mb=local_copy_threshold_mb,
            headers=headers,
        )

    def sheet(self, name: str) -> Sheet:
        """シートの高レベルラッパーを返す（シート単位でセル・行を書き込む）。

        Args:
            name: シート名。

        Raises:
            SheetNotFoundError: 指定したシートが存在しない場合。
        """
        return Sheet(self._sheet(name))

    def add_sheet(self, name: str, index: int | None = None) -> Sheet:
        """シートを追加し、そのまま書き込める Sheet を返す。

        Args:
            name: 追加するシート名。
            index: 挿入位置（0始まり）。省略時は末尾。

        Raises:
            SheetAlreadyExistsError: 同名のシートが既に存在する場合。
        """
        if name in self._wb.sheetnames:
            raise SheetAlreadyExistsError(name)
        return Sheet(self._wb.create_sheet(title=name, index=index))

    def rename_sheet(self, old_name: str, new_name: str) -> None:
        """シート名を変更する。"""
        if old_name not in self._wb.sheetnames:
            raise SheetNotFoundError(old_name, self._wb.sheetnames)
        if new_name in self._wb.sheetnames and new_name != old_name:
            raise SheetAlreadyExistsError(new_name)
        self._wb[old_name].title = new_name

    def delete_sheet(self, name: str) -> None:
        """シートを削除する。

        シートを削除すると、そのシートを参照している数式が ``#REF!`` になる。
        削除する前に、他のシートから参照されていないか確認すること。
        """
        if name not in self._wb.sheetnames:
            raise SheetNotFoundError(name, self._wb.sheetnames)
        if len(self._wb.sheetnames) == 1:
            raise LastSheetDeletionError(name)
        self._wb.remove(self._wb[name])

    @classmethod
    def create(cls, path: str | Path, sheet_name: str = "Sheet1") -> Self:
        """新規ブックを作る（ファイルはまだ作られず、save() で path に保存される）。
        Args:
            path: save() で保存されるパス。親フォルダがなければ保存時に自動作成される。
            sheet_name: 最初のシートの名前（デフォルト: "Sheet1"）。
        """
        instance = cls.__new__(cls)
        FileBase.__init__(instance, path)
        instance._original_path = instance._path
        instance._tmp = None
        instance._working_path = instance._original_path
        instance._headers = None
        instance._wb = Workbook()
        # 新規 Workbook() は必ず既定で1枚のアクティブシートを持つため None にはならない
        active = instance._wb.active
        if active is not None:
            active.title = sheet_name
        return instance

    @measure
    def save(
        self,
        path: str | Path | None = None,
        read_pw: str = "",
        write_pw: str = "",
    ) -> None:
        """ファイルを保存する。

        ローカルコピーで開いている場合も、省略時の保存先は元のファイル
        （一時コピーに保存すると close() でコピーごと消えてしまうため）。

        read_pw / write_pw のどちらかを指定すると、パスワード保護付きの
        ファイルとして保存する。openpyxl だけではパスワード保存ができないため、
        Excel COM を経由する。Excel が入っていない環境ではパスワード付き保存は
        できないため、入っている PC で実行すること。

        Args:
            path: 保存先のパス。省略すると開いた元のファイルに上書き保存する。
            read_pw: 読み取りパスワード。指定するとパスワード保護された
                     ファイルとして保存する（COM が必要な操作）。
                     設定した拡張子のファイル形式と中身の形式が一致しないと
                     Excel 側で警告が出る。
            write_pw: 書き込みパスワード。指定すると開くときに書き込みパスワード
                     を要求する形で保存する（COM が必要な操作）。

        Raises:
            ExcelSaveNotCompletedError: 保存後にファイルが見当たらない場合。
            ExcelApplicationNotAvailableError: パスワード付きで保存しようとした
                が、Excel が見つからない環境で実行した場合。
        """
        save_path = Path(path) if path else self._original_path
        logger.debug("Excel書き込み開始: %s", save_path)
        if is_dry_run():
            # パスワードは秘匿値なのでログには出さない（save_path のみ）
            dry_run_log("Excel を保存: %s", save_path)
            return
        save_path.parent.mkdir(parents=True, exist_ok=True)
        if not read_pw and not write_pw:
            # パスワードなし — openpyxl のみで完結する既存の経路。
            # os.replace は同一ドライブ内で使うため、保存先と同じフォルダに一時ファイルを作る。
            # NOTE: openpyxl がパスへ保存できるよう、名前を確保して即座に閉じる。
            tmp = tempfile.NamedTemporaryFile(  # noqa: SIM115
                dir=save_path.parent, prefix=f".{save_path.name}.", suffix=".tmp", delete=False
            )
            tmp_path = Path(tmp.name)
            tmp.close()
            try:
                self._wb.save(tmp_path)
                tmp_path.replace(save_path)
            finally:
                tmp_path.unlink(missing_ok=True)
        else:
            # パスワードあり — Excel COM 経由で保存する。COM は拡張子で形式を判定するため、
            # 一時ファイルの拡張子は save_path と同じにしないと（.xlsm を .tmp で保存すると）
            # FileFormatMismatchError で失敗する。一時ファイルは COM が読んだ後に消す。
            # NOTE: 一時ファイルを delete_file() で消さない理由 — delete_file() は dry-run を
            # 尊重してログだけ出すため、確実に消す必要がある場面では直接 unlink する。
            tmp = tempfile.NamedTemporaryFile(  # noqa: SIM115
                dir=save_path.parent,
                prefix=f".{save_path.name}.",
                suffix=save_path.suffix,
                delete=False,
            )
            tmp_path = Path(tmp.name)
            tmp.close()
            try:
                self._wb.save(tmp_path)
                from comken.toolbox.windows.handler import ExcelComHandler

                with ExcelComHandler(tmp_path) as com:
                    com.save_as(save_path, read_pw=read_pw, write_pw=write_pw)
            finally:
                tmp_path.unlink(missing_ok=True)
        # openpyxl と Excel COM は DisplayAlerts=False のとき、保存失敗を静かに通すことがある。
        # 呼び出し側が「保存できた」と誤認して元ファイルを消す事故を防ぐため、ファイルの存在で
        # もう一度確かめる。パスワードの有無に関わらず行う。
        if not save_path.exists():
            raise ExcelSaveNotCompletedError(save_path)
        logger.debug("Excel書き込み完了: %s", save_path)

    def run_macro(self, macro_name: str, save: bool = True) -> None:
        """VBA マクロを実行する。内部で win32com（pywin32）を使用する。

        COM は保存せずに閉じる仕様のため、save=True（デフォルト）で実行後に
        元ファイルへ保存する。マクロがブックを変更しても保存しないと結果は破棄される。

        WARNING: このメソッドは COM で元ファイルを直接編集する。openpyxl 側
            （Sheet で行った書き込み等）の未保存の変更とは独立で、run_macro の後に f.save() を
            呼ぶと openpyxl の内容で上書きされマクロの結果が消える。
            マクロと openpyxl 書き込みを混在させないこと。

        Args:
            macro_name: 実行するマクロ名。"モジュール名.プロシージャ名" の形式で指定する。
                        例: "Module1.UpdateData"
            save: True（デフォルト）ならマクロ実行後に元ファイルへ保存する。
        """
        from comken.toolbox.windows.handler import ExcelComHandler

        # local_copy の一時コピーではなく元ファイルに対して実行・保存する
        with ExcelComHandler(self._original_path) as com:
            com.run_macro(macro_name)
            if save:
                com.save()
