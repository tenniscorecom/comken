"""comken/toolbox/excel/workbook.py — Excel ブックとデータ領域を操作する。"""

import builtins
import hashlib
import os
import shutil
import tempfile
import warnings
from collections.abc import Callable, Mapping
from datetime import datetime
from pathlib import Path
from types import TracebackType
from typing import TYPE_CHECKING, Any, Literal, Self, TypeAlias, cast
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.table import Table as OpenPyXLTable
from openpyxl.worksheet.table import TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from comken.core.files import atomic_write, copy_to_local_if_large
from comken.core.table.model import Table
from comken.core.timer import measure
from comken.exceptions import (
    DuplicateHeaderCellError,
    EmptyHeaderCellError,
    ExcelFileNotFoundError,
    ExcelMacroPreservationError,
    ExcelReadOnlyOperationError,
    ExcelSaveValidationError,
    InvalidTableInputError,
    InvalidTableOperationError,
    SheetAlreadyExistsError,
    SheetNameError,
    SheetNotFoundError,
    TableAlreadyExistsError,
    TableNotOpenError,
    UnsupportedFileSuffixError,
)
from comken.runtime import dry_run_log, is_dry_run
from comken.toolbox.excel.sheet import Sheet
from comken.toolbox.excel.table import ExcelTable

if TYPE_CHECKING:
    from comken.toolbox.windows.handler import ExcelCOMHandler

Engine = Literal["openpyxl", "com"]

Value: TypeAlias = str | int | float | bool | datetime
_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def _force_local_copy(path: Path) -> tuple[Path, Path]:
    """UNCパス上のブックを作業コピーへ複製し、``(working_path, tmp_path)`` を返す。

    ``copy_to_local_if_large`` はサイズでコピーを判定するが、UNCパス指定時
    （書き込み経路）はサイズに関わらず作業コピーを使うのが要件。保存時に
    元へ書き戻す関係上、tmp のライフサイクルは呼び出し側が持つ。
    """
    fd, name = tempfile.mkstemp(suffix=path.suffix)
    os.close(fd)
    tmp_path = Path(name)
    shutil.copy2(path, tmp_path)
    return tmp_path, tmp_path


class Excel:
    """Excel ワークブックを開き、シート単位の操作を提供する。"""

    PY_PREFIX = "PY_"

    def __init__(
        self,
        source: str | Path,
        *,
        types: Mapping[str, Callable[[Any], Any]] | None = None,
        read_only: bool = False,
        local_copy: bool | None = None,
        engine: Engine = "openpyxl",
    ) -> None:
        """設定を保持する。**ブックは開かない**。

        ``with`` の中で ``__enter__`` が呼ばれたとき、はじめてブックを開く。
        読み取り専用で開くか、書き込み用に開くかは引数 ``read_only`` で
        切り替える。利用者がエンジンを選ぶ必要はない。通常操作はOpenPyXLを使い、
        未計算の数式値の読取りとVBA実行だけ、一時的にExcel COMへ昇格する。
        ``local_copy=None`` の既定では、書き込み時にUNCパスのブックだけ
        一時作業コピーを使う（読み取り専用では UNC でもコピーしない。保存が無いため
        「作業中だけローカルを使い、保存時に元へ戻す」契約を適用する場面がない）。
        ``local_copy=True`` で強制、``local_copy=False`` で無効化でき、保存先は常に
        元ファイルになる。

        ``engine`` でブックを開く実装を選べる。``"openpyxl"``（既定）は openpyxl を
        使う通常経路、``"com"`` は pywin32 経由で Excel COM に接続する。
        ``"com"`` は約31シート+ピボット十数個のような openpyxl で重いブック、
        または数式・マクロ・COM 固有 API を直接利用したいときに使う。
        ``local_copy`` の扱いはエンジンごとに分かれる（``engine='com'`` のときは
        ``ExcelCOMHandler`` の ``local_copy_threshold_mb`` に変換する。未指定なら
        ``__enter__`` で一度だけ警告する）。

        ``read_only``、dry-run、またはwithブロックが例外で終わった場合は保存しない。
        """
        self.path = Path(source)
        if self.path.suffix.casefold() not in _EXCEL_SUFFIXES:
            raise UnsupportedFileSuffixError(self.path, tuple(sorted(_EXCEL_SUFFIXES)))
        self._types = dict(types or {})
        self._read_only = read_only
        self._engine = engine
        # ``_local_copy_required`` は「作業コピーが必要か」を bool で持つ openpyxl 側の内部表現。
        # engine='com' の経路では openpyxl を使わないのでこの値は参照しない。
        # ``_local_copy`` は呼び出し時の ``local_copy`` の値を ``bool | None`` でそのまま残し、
        # engine='com' 経路で「ユーザーが明示したか」を警告判定に使う。
        self._local_copy: bool | None = local_copy
        if local_copy is None:
            # 読み取り専用は保存が無いため作業コピー（=後で元へ書き戻す中間ファイル）が
            # 要らない。UNC でもネットワーク越しに直接読む（openpyxl の read_only
            # ストリームはUNC でも安定して動作する）。
            if read_only:
                self._local_copy_required = False
            else:
                # 書き込み時は UNC でのみ作業コピーを挟む。ローカルドライブは
                # そもそもネットワーク問題の起きない素の読み書きで十分。
                self._local_copy_required = self._is_unc_path(source)
        else:
            self._local_copy_required = local_copy
        self._local_copy_path: Path | None = None
        self._working_path: Path = self.path
        self._working_copy_is_stale = False
        # 通常 Workbook は遅延。read_only で stream だけで完結する経路
        # (``_read_computed_rows``) では開かれない。書き込みや Excel テーブル API
        # が要求されたときだけ遅延オープンする。
        self._workbook: Workbook | None = None
        # 数式判定用の read_only ストリーム Workbook。``_read_computed_rows`` 系の
        # ホットパスから zip を 1 度しか読まないようキャッシュする。
        self._stream_workbook: Workbook | None = None
        self._stream_workbook_data_only: Workbook | None = None
        # engine='com' 用に保持する ExcelCOMHandler。openpyxl 経路では触らない。
        self._com_handler: ExcelCOMHandler | None = None
        # engine='com' で local_copy 未指定の警告を ``__enter__`` で1度だけ出す。
        self._local_copy_warned = False
        self._is_open = False
        self._is_closed = False
        self._is_dirty = False

    def __enter__(self) -> Self:
        # 既に開いている状態で再度入ってきた場合は、二重にブックを開かない。
        if self._is_open and not self._is_closed:
            return self
        if self._engine == "com":
            self._enter_com_engine()
            return self
        self._enter_openpyxl_engine()
        return self

    def _enter_com_engine(self) -> None:
        """``engine='com'`` でブックを開く（``__enter__`` の一部）。"""
        # pywin32 が無い環境では ExcelCOMHandler を import できない。関数内 import に
        # することで openpyxl 経路だけを使う利用側-PC では import 失敗を避ける
        # （既存 ``_read_range_with_com`` と同じ設計）。
        from comken.toolbox.windows.handler import ExcelCOMHandler

        if not self._working_path.exists():
            # COM は既存ファイルを開くための経路なので、新規ファイル作成は openpyxl に任せる
            raise ExcelFileNotFoundError(self.path)
        threshold = self._resolve_com_local_copy_threshold()
        self._com_handler = ExcelCOMHandler(self._working_path, local_copy_threshold_mb=threshold)
        self._is_open = True
        self._is_closed = False

    def _resolve_com_local_copy_threshold(self) -> float:
        """``engine='com'`` 用の ``local_copy_threshold_mb`` を ``_local_copy`` から算出する。"""
        if self._local_copy is True:
            return 0
        if self._local_copy is False:
            # ``copy_to_local_if_large`` は ``threshold_mb=0`` を「コピー無効」と
            # 扱うので、``float('inf')`` を渡しても同じ結果になる（stat() 比較が
            # 常に True になり、元パスをそのまま返す）。
            return float("inf")
        if not self._local_copy_warned:
            warnings.warn(
                "engine='com' で local_copy を明示していません。"
                "UNC パスでは不安定になることがあるため、"
                "local_copy=True か local_copy=False を明示してください。",
                stacklevel=2,
            )
            self._local_copy_warned = True
        return 10  # ExcelCOMHandler の既定

    def _enter_openpyxl_engine(self) -> None:
        """``engine='openpyxl'`` でブックを開く（``__enter__`` の一部）。"""
        if self._local_copy_required and self.path.exists():
            # NAS・ネットワークドライブ上のブックを直接扱うと OpenPyXL/Excel の I/O が
            # 不安定になることがある。作業中だけローカルを使い、保存時に元のパスへ戻す。
            # 仕様書 4.5 では「大きなブック」と書かれているが、UNC ではサイズに関わら
            # ず全件コピーする（社内で扱うブックは小さく個別閾値を設ける実務的意義が薄い）。
            self._working_path, self._local_copy_path = _force_local_copy(self.path)
        else:
            self._working_path = self.path
        if not self._read_only:
            # 書き込み用ブックは __enter__ で必ず開く。``_read_computed_rows`` のように
            # メモリ上のセル値を参照する経路があるため。
            if self._working_path.exists():
                self._workbook = load_workbook(
                    self._working_path,
                    read_only=False,
                    keep_vba=self.path.suffix.casefold() in {".xlsm", ".xltm"},
                )
            else:
                self._workbook = Workbook()
        else:
            # 読み取り専用では Workbook を遅延オープンする。``Sheet.table()`` や
            # ``excel.sheet(...)._worksheet[...]`` のように通常Worksheetを要求する
            # 経路は、その時点で ``_ensure_normal_workbook()`` が開く。
            if not self._working_path.exists():
                raise ExcelFileNotFoundError(self.path)
        self._is_open = True
        self._is_closed = False

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_value: BaseException | None,
        traceback: TracebackType | None,
    ) -> None:
        # 例外時に保存すると、途中までの変更で元ファイルを壊す可能性がある。
        # read_only と dry-run の扱いは close() 側へ集約している。
        self.close(save=exc_type is None)

    def sheet(self, name: str | None = None) -> "Sheet":
        """名前でシートを取得する。未存在の新規ブックでは最初のシートを改名する。"""
        # engine='com' は Worksheet を返さない設計（``Sheet`` 系 API は openpyxl 前提）。
        # 共通 API（``list_sheets`` / ``last_row`` / ``exists_sheet`` など）か
        # ``excel.com_handler`` 経由で操作する。
        if self._engine == "com":
            self._ensure_open()
            raise InvalidTableOperationError(
                "engine='com' では excel.sheet() は使えません。"
                "Sheet 系の API は openpyxl 経路（engine='openpyxl'）で開いてください。"
            )
        self._ensure_normal_workbook()
        if name is None:
            display_sheets = [
                sheet for sheet in self._workbook.sheetnames if not self._is_data_sheet_name(sheet)
            ]
            if len(display_sheets) != 1:
                raise SheetNotFoundError("省略", display_sheets)
            name = display_sheets[0]
        if name not in self._workbook.sheetnames:
            if not self.path.exists() and self._is_pristine_workbook():
                cast(Worksheet, self._workbook.active).title = name
                self._is_dirty = True
            else:
                raise SheetNotFoundError(name, self._workbook.sheetnames)
        return Sheet(self, self._workbook[name])

    def find_sheet(self, *candidates: str) -> str:
        """候補を順に試し、最初に見つかったシートの名前を返す。

        「古いファイルと新しいファイルでシート名が違う」「テンプレ更新で
        シート名が変わった」のように、**業務上よくある候補の違い**を 1 行で
        吸収する。 ``Config`` 側で ``SHEET_NAME = [Sheet1, 一覧]`` のように
        候補リストを持っておき、その順番に試したいときに使う。

        戻り値は **シート名（``str``）**。``Sheet`` インスタンスが要るときは
        戻ってきた名前を ``self.sheet(name)`` に渡す。

        候補が全て見つからないときは、最後の試行の名前で ``SheetNotFoundError``
        を送出する（メッセージにブックに実在するシート名一覧が入るので、
        利用者が config を直せる）。 候補を 1 つも渡さなかったときも、
        同じ例外（候補名が空文字・実在シート一覧入り）で止める。

        ``self.sheet(name)`` を経由せず ``sheetnames`` の所属判定で済ませる。
        ``sheet()`` は未存在の新規ブックで **自動でリネーム**する仕様なので、
        候補違いのときに知らぬ間にブックが変わる事故を防ぐ。
        """
        if self._engine == "com":
            self._ensure_open()
            raise InvalidTableOperationError(
                "engine='com' では excel.find_sheet() は使えません。"
                "exists_sheet() / list_sheets() を使ってください。"
            )
        self._ensure_normal_workbook()
        last_error = SheetNotFoundError(
            "" if not candidates else candidates[-1], self._workbook.sheetnames
        )
        for name in candidates:
            if name in self._workbook.sheetnames:
                return name
            last_error = SheetNotFoundError(name, self._workbook.sheetnames)
        raise last_error

    def data_sheet(self, name: str | None = None) -> "Sheet":
        """データシートを取得する。名前を省略できるのは1枚のときだけ。"""
        if self._engine == "com":
            self._ensure_open()
            raise InvalidTableOperationError(
                "engine='com' では excel.data_sheet() は使えません。"
                "Sheet 系の API は openpyxl 経路で開いてください。"
            )
        self._ensure_normal_workbook()
        names = self.list_data_sheets()
        if name is None:
            if len(names) != 1:
                raise SheetNotFoundError("データシート省略", names)
            name = names[0]
        return self.sheet(self._with_python_prefix(name))

    def create_data_sheet(self, name: str) -> "Sheet":
        """指定名の空のデータシートを作成する。"""
        if self._engine == "com":
            self._ensure_open()
            raise InvalidTableOperationError(
                "engine='com' では excel.create_data_sheet() は使えません。"
                "Sheet 系の API は openpyxl 経路で開いてください。"
            )
        self._ensure_writable("create_data_sheet")
        full_name = self._with_python_prefix(name)
        if full_name in self._workbook.sheetnames:
            raise SheetAlreadyExistsError(full_name)
        worksheet = self._workbook.create_sheet(full_name)
        self._mark_dirty()
        return Sheet(self, worksheet)

    def create_sheet(self, name: str) -> "Sheet":
        """指定名の空の表示用シートを作成する。

        ``create_data_sheet`` は ``PY_`` プレフィックスを補ってデータシート専用
        にするのに対し、こちらは入力名をそのまま使い、表示用の自由配置として
        読み書きする。書式や自由セル配置が要る帳票は ``create_sheet``、
        構造化テーブルとして読み書きするなら ``create_data_sheet``。
        """
        if self._engine == "com":
            self._ensure_open()
            raise InvalidTableOperationError(
                "engine='com' では excel.create_sheet() は使えません。"
                "Sheet 系の API は openpyxl 経路で開いてください。"
            )
        self._ensure_writable("create_sheet")
        if self._is_data_sheet_name(name):
            raise SheetNameError(name)
        if name in self._workbook.sheetnames:
            raise SheetAlreadyExistsError(name)
        worksheet = self._workbook.create_sheet(name)
        self._mark_dirty()
        return Sheet(self, worksheet)

    def list_data_sheets(self) -> list[str]:
        """データシート名をブック内の順序で返す。"""
        if self._engine == "com":
            self._ensure_open()
            raise InvalidTableOperationError(
                "engine='com' では excel.list_data_sheets() は使えません。"
                "list_sheets() を使ってください。"
            )
        self._ensure_normal_workbook()
        return [name for name in self._workbook.sheetnames if self._is_data_sheet_name(name)]

    def _read_range_with_com(
        self, sheet_name: str, min_col: int, min_row: int, max_col: int, max_row: int
    ) -> list[tuple[Any, ...]]:
        """実テーブル範囲だけをCOMの計算値で読む。"""
        if self._is_dirty:
            self._prepare_com_working_copy()
        self._sync_working_file()
        from comken.toolbox.windows.handler import ExcelCOMHandler

        with ExcelCOMHandler(self._working_path, local_copy_threshold_mb=0) as excel_com:
            return excel_com.read_block(sheet_name, min_col, min_row, max_col, max_row)

    @property
    def com_handler(self) -> "ExcelCOMHandler":
        """engine='com' で開いている内部の ``ExcelCOMHandler`` を返す。

        ``run_macro`` / ``save_as``（パスワード付き保存）など、Phase 1 で
        共通化しない COM 機能への直接アクセス用。``engine='openpyxl'`` の
        インスタンスで触ると ``InvalidTableOperationError`` で止める。
        """
        self._ensure_open()
        if self._com_handler is None:
            raise InvalidTableOperationError(
                "engine='openpyxl' では com_handler にアクセスできません。"
                "engine='com' で開いた Excel インスタンスで使ってください。"
            )
        return self._com_handler

    @measure
    def list_sheets(self) -> list[str]:
        """シート名一覧（表示用・データ用を含む全シート）を返す。

        ``engine`` によらず同じシグネチャで使える薄い API。``PY_`` プレフィックスの
        データシートもそのまま含める。「``list_``」の動詞は ``list_data_sheets``
        と同じ命名で、シート名一覧を順序通り返す操作のため例外的に許容する。
        """
        if self._engine == "com":
            self._ensure_open()
            com_handler = self._com_handler
            assert com_handler is not None  # _ensure_open で開いた後は必ず設定済み
            return [sheet.Name for sheet in com_handler._wb.Sheets]
        self._ensure_normal_workbook()
        workbook = self._workbook
        assert workbook is not None  # _ensure_normal_workbook の後は必ず Workbook が開く
        return list(workbook.sheetnames)

    @measure
    def count_sheets(self) -> int:
        """シート数を返す。"""
        if self._engine == "com":
            self._ensure_open()
            com_handler = self._com_handler
            assert com_handler is not None  # _ensure_open で開いた後は必ず設定済み
            return int(com_handler._wb.Sheets.Count)
        self._ensure_normal_workbook()
        workbook = self._workbook
        assert workbook is not None  # _ensure_normal_workbook の後は必ず Workbook が開く
        return len(workbook.sheetnames)

    @measure
    def last_row(self, sheet_name: str) -> int:
        """指定シートの最終データ行の番号を返す（1 始まり）。

        openpyxl は ``Worksheet.max_row`` を使う。COM は ``ExcelCOMHandler.last_row``
        の ``UsedRange`` 計算を使う（数式が入ったセルも含む最終行）。
        """
        if self._engine == "com":
            self._ensure_open()
            com_handler = self._com_handler
            assert com_handler is not None  # _ensure_open で開いた後は必ず設定済み
            return com_handler.last_row(sheet_name)
        self._ensure_normal_workbook()
        workbook = self._workbook
        assert workbook is not None  # _ensure_normal_workbook の後は必ず Workbook が開く
        if sheet_name not in workbook.sheetnames:
            raise SheetNotFoundError(sheet_name, workbook.sheetnames)
        worksheet = workbook[sheet_name]
        return int(worksheet.max_row)

    @measure
    def exists_sheet(self, name: str) -> bool:
        """指定した名前のシートが存在するか返す。"""
        if self._engine == "com":
            self._ensure_open()
            com_handler = self._com_handler
            assert com_handler is not None  # _ensure_open で開いた後は必ず設定済み
            return any(sheet.Name == name for sheet in com_handler._wb.Sheets)
        self._ensure_normal_workbook()
        workbook = self._workbook
        assert workbook is not None  # _ensure_normal_workbook の後は必ず Workbook が開く
        return name in workbook.sheetnames

    def convert_range_to_table(
        self,
        sheet_name: str,
        *,
        range: str,
        table_name: str,
        header_row: int | None = None,
    ) -> ExcelTable:
        """既存のセル範囲を Excel テーブルに変換する。安全に判定できない場合はエラーで止める。

        既存のセル値をそのままテーブル化する。``Sheet.create_table`` が「新規に
        ``Table`` を書き込んで作る」のに対し、こちらは「既に書き込まれている表を
        Excel テーブルとして登録する」操作。表示用シート・データシートどちらでも
        利用可能で、``PY_T_`` プレフィックスは補わない（指定された名前をそのまま使う）。

        自動判定は限定的にする: ``header_row`` 未指定のとき ``range`` の先頭行に
        結合セルがあれば次の行を見出しとする（A2 ルール）。それ以外の推定
        （フォントサイズ差・空白判定など）は行わない。``header_row`` を明示する
        ほうが事故が少ない。

        Args:
            sheet_name: 対象のシート名。
            range: テーブル化するセル範囲（例: ``"A1:E100"``）。必須。
            table_name: 登録するテーブル名。Excel の命名規則に従う必要がある。
            header_row: 見出し行の番号（1 始まり）。``None`` のとき A2 ルールで推定。

        Returns:
            作成された ``ExcelTable``。

        Raises:
            NotImplementedError: ``engine='com'`` で開いたインスタンスで呼ばれたとき。
            InvalidTableInputError: 範囲・結合・空データ行のいずれかが条件違反のとき。
            EmptyHeaderCellError: 見出し行に空セルがあるとき。
            DuplicateHeaderCellError: 見出し行に同じ名前が複数あるとき。
            InvalidTableNameError: ``table_name`` が Excel の命名規則に合わないとき。
            TableAlreadyExistsError: 指定したテーブル名が既に存在するとき。
        """
        self._ensure_open()
        if self._engine != "openpyxl":
            # engine='com' では Worksheet を保持しないため、openpyxl の Table オブジェクトを
            # 作成するこの API は対応しない。COM 経路でテーブル化したい場合は openpyxl で
            # 開いたブックで実行してから COM で読む、という流れにする。
            raise NotImplementedError(
                "convert_range_to_table は openpyxl で開いたブックでのみ対応しています。"
                "engine='openpyxl' で開いてください。"
            )
        self._ensure_writable("convert_range_to_table")
        self._ensure_normal_workbook()
        workbook = self._workbook
        assert workbook is not None  # _ensure_normal_workbook の後は必ず Workbook が開く
        if sheet_name not in workbook.sheetnames:
            raise SheetNotFoundError(sheet_name, workbook.sheetnames)
        worksheet = workbook[sheet_name]
        _min_col, _min_row, range_max_col, range_max_row, header_row, _header_cells = (
            self._validate_range_for_table(worksheet, range, header_row)
        )
        # ``table_name`` が Excel の命名規則を満たすこと。
        Sheet._validate_table_name(table_name)
        # 変換先のテーブル名が既存テーブルと衝突しないこと。
        # ``Sheet.create_table`` と異なり ``PY_T_`` プレフィックスは補わない
        # （表示用シートの既存表をそのままテーブル化するため）。
        if table_name in worksheet.tables:
            raise TableAlreadyExistsError(table_name)
        # テーブル化: 既存値はそのままで ``Table`` 定義だけを書き加える。
        ref = (
            f"{get_column_letter(_min_col)}{header_row}:"
            f"{get_column_letter(range_max_col)}{range_max_row}"
        )
        excel_table = OpenPyXLTable(displayName=table_name, ref=ref)
        excel_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(excel_table)
        self._mark_dirty()
        return ExcelTable(self, worksheet, table_name)

    def _validate_range_for_table(
        self,
        worksheet: Worksheet,
        range_spec: str,
        explicit_header_row: int | None,
    ) -> tuple[int, int, int, int, int, list[Any]]:
        """``convert_range_to_table`` の安全性判定を集約する。

        戻り値は ``(min_col, min_row, max_col, max_row, header_row, header_cells)``。
        各判定に違反したら対応する既存例外で止める。
        """
        boundaries = range_boundaries(range_spec)
        # ``range_boundaries`` は型ヒント上 ``int | None`` を返すが、利用者が ``range_spec``
        # を ``range_boundaries`` を通せる形で渡しているため通常 ``None`` にはならない。
        # ``None`` が現れた場合のフォールバックは 0 とする（``cell(...)`` 呼び出しが
        # 「先頭セル」として動くため、見える不整合は出ない）。
        range_min_col = 0 if boundaries[0] is None else boundaries[0]
        range_min_row = 0 if boundaries[1] is None else boundaries[1]
        range_max_col = 0 if boundaries[2] is None else boundaries[2]
        range_max_row = 0 if boundaries[3] is None else boundaries[3]
        self._check_range_inside_dimensions(
            worksheet, range_spec, range_min_col, range_min_row, range_max_col, range_max_row
        )
        header_row = self._resolve_header_row(
            worksheet, explicit_header_row, range_min_row, range_min_col, range_max_col
        )
        # 引数 ``range`` が組み込み ``range`` を覆い隠すため ``builtins.range`` を直接呼ぶ
        header_cells = [
            worksheet.cell(row=header_row, column=column).value
            for column in builtins.range(range_min_col, range_max_col + 1)
        ]
        self._check_header_not_empty(header_cells)
        self._check_no_merged_cells_in_range(
            worksheet,
            range_spec,
            range_min_row,
            range_max_row,
            range_min_col,
            range_max_col,
            header_row,
        )
        self._check_no_blank_data_rows(
            worksheet, header_row, range_max_row, range_min_col, range_max_col
        )
        self._check_no_duplicate_headers(header_cells)
        return range_min_col, range_min_row, range_max_col, range_max_row, header_row, header_cells

    @staticmethod
    def _check_range_inside_dimensions(
        worksheet: Worksheet,
        range_spec: str,
        range_min_col: int,
        range_min_row: int,
        range_max_col: int,
        range_max_row: int,
    ) -> None:
        """``range`` がシートの使用範囲 (``dimensions``) 内に収まっているか検証する。"""
        dim = worksheet.dimensions
        # ``dimensions`` が空（新規シート等）のときは ``range`` をそのまま許容する。
        if not dim or dim == "A1:A1":
            return
        # ``range_boundaries`` は型ヒント上 ``int | None`` を返すが、既存の
        # ``_table_boundaries`` と同じく「dimensions から取った値は通常 None にならない」
        # 前提で 0 フォールバックを ``int`` に揃える。
        boundaries = range_boundaries(dim)
        dim_min_col = 0 if boundaries[0] is None else boundaries[0]
        dim_min_row = 0 if boundaries[1] is None else boundaries[1]
        dim_max_col = 0 if boundaries[2] is None else boundaries[2]
        dim_max_row = 0 if boundaries[3] is None else boundaries[3]
        if (
            range_min_col < dim_min_col
            or range_min_row < dim_min_row
            or range_max_col > dim_max_col
            or range_max_row > dim_max_row
        ):
            raise InvalidTableInputError(
                f"指定した範囲 {range_spec} がシート「{worksheet.title}」の使用範囲 "
                f"({dim}) を超えています。"
            )

    @staticmethod
    def _resolve_header_row(
        worksheet: Worksheet,
        explicit_header_row: int | None,
        range_min_row: int,
        range_min_col: int,
        range_max_col: int,
    ) -> int:
        """見出し行の番号を決定する（明示 > A2 ルール > 先頭行）。"""
        if explicit_header_row is not None:
            return explicit_header_row
        # A2 ルール: 先頭行に結合セルがあるなら次行を見出し行とする（タイトル行の許容）。
        for merged_range in worksheet.merged_cells.ranges:
            if (
                merged_range.min_row == range_min_row
                and merged_range.min_col >= range_min_col
                and merged_range.max_col <= range_max_col
            ):
                return range_min_row + 1
        return range_min_row

    @staticmethod
    def _check_header_not_empty(header_cells: list[Any]) -> None:
        """見出し行のセルが全て埋まっているか検証する。"""
        empty_columns = [
            column
            for column, value in enumerate(header_cells, start=1)
            if value is None or value == ""
        ]
        if empty_columns:
            raise EmptyHeaderCellError(empty_columns)

    @staticmethod
    def _check_no_merged_cells_in_range(
        worksheet: Worksheet,
        range_spec: str,
        range_min_row: int,
        range_max_row: int,
        range_min_col: int,
        range_max_col: int,
        header_row: int,
    ) -> None:
        """``range`` 内に結合セルが無いことを検証する。

        A2 が発火した「見出し行より前の行」にある結合はタイトル行として許容する。
        """
        merged_in_range: list[str] = []
        for merged_range in worksheet.merged_cells.ranges:
            if (
                merged_range.min_row < range_min_row
                or merged_range.max_row > range_max_row
                or merged_range.min_col < range_min_col
                or merged_range.max_col > range_max_col
            ):
                # 完全に範囲外
                continue
            # ``header_row`` 以下の行に完全に収まる結合は「タイトル領域」として除外する。
            if merged_range.max_row < header_row:
                continue
            merged_in_range.append(str(merged_range))
        if merged_in_range:
            raise InvalidTableInputError(
                f"指定した範囲 {range_spec} に結合セルがあります: "
                f"{', '.join(merged_in_range)}。"
                "convert_range_to_table は結合セルを含まない範囲をテーブル化します。"
                "タイトル行として許容されるのは見出し行より上にある結合だけです。"
            )

    @staticmethod
    def _check_no_blank_data_rows(
        worksheet: Worksheet,
        header_row: int,
        range_max_row: int,
        range_min_col: int,
        range_max_col: int,
    ) -> None:
        """データ行（見出し行より下、range 内）に「全セル空」の行が無いことを検証する。"""
        blank_data_rows: list[int] = []
        for row_index in builtins.range(header_row + 1, range_max_row + 1):
            row_values = [
                worksheet.cell(row=row_index, column=column).value
                for column in builtins.range(range_min_col, range_max_col + 1)
            ]
            if all(value is None or value == "" for value in row_values):
                blank_data_rows.append(row_index)
        if blank_data_rows:
            raise InvalidTableInputError(
                f"指定した範囲のデータ行に空行があります: 行 {blank_data_rows}。"
                "データは連続している必要があります。"
            )

    @staticmethod
    def _check_no_duplicate_headers(header_cells: list[Any]) -> None:
        """見出し行の値に重複が無いことを検証する。"""
        header_strings = ["" if value is None else str(value) for value in header_cells]
        duplicates: list[str] = []
        seen: set[str] = set()
        for header in header_strings:
            if header in seen:
                duplicates.append(header)
            else:
                seen.add(header)
        if duplicates:
            raise DuplicateHeaderCellError(duplicates)

    def close(self, *, save: bool = True) -> None:
        """ブックを閉じる。通常はwithの正常終了時に変更を自動保存する。"""
        if self._is_closed or not self._is_open:
            return
        try:
            if save and self._is_dirty:
                self.save()
        finally:
            if self._workbook is not None:
                self._workbook.close()
            self._close_stream_workbook(self._stream_workbook)
            self._close_stream_workbook(self._stream_workbook_data_only)
            self._stream_workbook = None
            self._stream_workbook_data_only = None
            # engine='com' で開いていた場合は ExcelCOMHandler を閉じる。
            # ExcelCOMHandler.close() は2回呼んでも安全な実装なので、ここで再代入だけ行う。
            if self._com_handler is not None:
                self._com_handler.close()
                self._com_handler = None
            self._is_closed = True
            self._is_open = False
            if self._local_copy_path is not None:
                self._local_copy_path.unlink(missing_ok=True)

    @staticmethod
    def _close_stream_workbook(workbook: Workbook | None) -> None:
        """read_only ストリームの Workbook をクローズする。None なら何もしない。"""
        if workbook is not None:
            workbook.close()

    @staticmethod
    def _row_is_blank(values: Any) -> bool:
        """行の全セルが「空」（``None`` または空文字 ``""``）かを返す。

        Excel の ``dimension`` は書式の残ったセルにも広がるため、宣言された
        範囲をそのまま信じると思わぬ数の空行が返る（症状: 2000 行のブックから
        30 万行返る）。 ``0`` や ``False`` は値として残す（数値の 0 を落とすと
        集計が狂うため）。共通判定をここに集約する。
        """
        return all(value is None or value == "" for value in values)

    @measure
    def save(self) -> None:
        """変更を元ファイルへ保存する。

        長い処理の途中で確定したいときに使う。``with`` を分けて閉じ開きすると、
        共有サーバー上のファイルではロックや同期の問題を自分で作り出すことになるため、
        この経路を残している。``save()`` の後に変更がなければ ``with`` 終了時に
        再保存はしない（``_is_dirty`` で判定）。
        """
        self._ensure_open()
        if self._engine == "com":
            # engine='com' 経路では変更しない設計のため保存経路は用意しない。
            # 保存したい場合は ``excel.com_handler.save()`` を使う。
            raise InvalidTableOperationError(
                "engine='com' では excel.save() は使えません。"
                "保存は excel.com_handler.save() を利用してください。"
            )
        self._ensure_normal_workbook()
        if self._read_only or not self._is_dirty:
            return
        if is_dry_run():
            dry_run_log("Excel を保存: %s", self.path)
            return
        self.path.parent.mkdir(parents=True, exist_ok=True)
        original_vba = self._vba_digest(self.path)
        with atomic_write(self.path) as temporary_path:
            self._workbook.save(temporary_path)
            verification = None
            try:
                verification = load_workbook(
                    temporary_path,
                    read_only=True,
                    keep_vba=self.path.suffix.casefold() in {".xlsm", ".xltm"},
                )
            except (KeyboardInterrupt, SystemExit):
                raise
            except Exception as error:
                raise ExcelSaveValidationError(self.path, error) from error
            finally:
                if verification is not None:
                    verification.close()
            if original_vba is not None and self._vba_digest(temporary_path) != original_vba:
                raise ExcelMacroPreservationError(self.path)
        self._is_dirty = False
        # 明示 save() 後も with 内でCOM操作を続けられる。次のCOM利用時に
        # 保存後の原本からローカル作業コピーを同期し、古い値を開かない。
        self._working_copy_is_stale = self._working_path != self.path

    @measure
    def run_macro(self, macro_name: str) -> None:
        """Excel COMへ一時的に昇格してVBAマクロを実行する。

        COMには元ファイルではなく作業ファイルを渡す。ローカルコピー利用時にも、
        例外終了なら元ファイルを変更しないというwithの契約を守るためである。
        """
        if self._engine == "com":
            self._ensure_open()
            # engine='com' のブックは既に COM で開かれている。COM 専用 API
            # （``excel.com_handler.run_macro``）を使ってもらう。
            raise InvalidTableOperationError(
                "engine='com' では excel.run_macro() は使えません。"
                "excel.com_handler.run_macro() を利用してください。"
            )
        self._ensure_writable("run_macro")
        if is_dry_run():
            dry_run_log("Excel マクロを実行: %s (%s)", macro_name, self.path)
            return
        self._prepare_com_working_copy()
        self._sync_working_file()
        from comken.toolbox.windows.handler import ExcelCOMHandler

        with ExcelCOMHandler(self._working_path, local_copy_threshold_mb=0) as excel_com:
            excel_com.run_macro(macro_name)
            excel_com.save()
        self._reload_workbook()
        self._is_dirty = True

    @measure
    def _read_computed_rows(self, sheet_name: str, min_row: int = 2) -> list[tuple[Any, ...]]:
        """数式の計算結果を行単位で読む。未計算の数式がある場合だけCOMへ昇格する。

        公開 API の ``read()`` が内部で使う薄いヘルパー。下流のテストや COM 経由
        の ``read_row_values()`` との橋渡し用。``read()`` ではなく行タプルを
        直接欲しいときだけ利用する想定（通常は ``read()`` を使う）。
        """
        self._ensure_open()
        rows, needs_com = self._cached_rows(sheet_name, min_row)
        if not needs_com:
            return rows
        # 書込み後の計算読取りでも、後続処理が例外なら元ファイルを変えない。
        # COMへ同期する先を一時コピーへ切り替えてから保存する。
        if self._is_dirty:
            self._prepare_com_working_copy()
        self._sync_working_file()
        from comken.toolbox.windows.handler import ExcelCOMHandler

        with ExcelCOMHandler(self._working_path, local_copy_threshold_mb=0) as excel_com:
            return excel_com.read_row_values(sheet_name, min_row)

    def read(
        self,
        sheet_name: str,
        *,
        header_row: int = 1,
        force_com: bool = False,
    ) -> Table:
        """見出し行をキーに計算結果を読み ``Table`` で返す。未計算時だけCOMへ昇格する。

        ``header_row`` で見出し行の番号を指定する。既定は 1（先頭行）。
        ``force_com=True`` でキャッシュを無視して Excel 実機で強制再計算させる。
        数式の列は ``Table`` 化されない（文字列のまま入るので、必要なら ``Table`` の
        ``types`` で ``int`` / ``float`` などに変換すること）。

        engine='com' で開いたブックは、内部で保持している ``ExcelCOMHandler`` に
        そのまま委譲する（COM で再計算済みの値を読む）。

        Args:
            sheet_name: シート名。
            header_row: 見出し行の番号（1 始まり）。既定は 1。
            force_com: ``True`` でキャッシュを無視して Excel 実機で強制再計算。

        Returns:
            シートの内容を表す ``Table``。全セルが空の行は除外される。
        """
        self._ensure_open()
        if self._engine == "com":
            # engine='com' は「COM 専用で読む」前提なので ``force_com`` は実質無効。
            # 受け取って無視しても害はないため、エラーにはせず握りつぶす。
            assert self._com_handler is not None  # _ensure_open の後に必ず設定済み
            return self._com_handler.read(sheet_name, header_row=header_row)
        rows = self._read_computed_rows(sheet_name, header_row)
        if not rows:
            return Table([], [])
        headers = list(rows[0])
        empty_columns = [index for index, header in enumerate(headers, start=1) if header is None]
        if empty_columns:
            raise EmptyHeaderCellError(empty_columns)
        data_rows = [dict(zip(headers, row, strict=False)) for row in rows[1:]]
        return Table(headers, data_rows)

    def _ensure_open(self) -> None:
        if not self._is_open or self._is_closed:
            raise TableNotOpenError("Excel")

    def _ensure_normal_workbook(self) -> None:
        """通常モード Workbook がまだなら遅延オープンする。

        書き込み用 ``Excel`` では ``__enter__`` で既に開いている。``read_only`` の
        Excel では ``Sheet.table()`` / セル属性アクセスなど Worksheet を必要とする
        経路で呼ばれたときだけ開く。``read_only`` 経路の ``_read_computed_rows`` は
        このメソッドを呼ばないため、通常 Workbook を開かない。
        """
        self._ensure_open()
        if self._workbook is not None:
            return
        if self._working_path.exists():
            self._workbook = load_workbook(
                self._working_path,
                read_only=False,
                keep_vba=self.path.suffix.casefold() in {".xlsm", ".xltm"},
            )
        else:
            if self._read_only:
                raise ExcelFileNotFoundError(self.path)
            self._workbook = Workbook()

    def _mark_dirty(self) -> None:
        self._ensure_writable("書き込み")
        self._is_dirty = True

    def _ensure_writable(self, operation: str) -> None:
        self._ensure_normal_workbook()
        if self._read_only:
            raise ExcelReadOnlyOperationError(operation)

    def _sync_working_file(self) -> None:
        """COMへ渡す前に現在状態を作業ファイルへ同期する。"""
        if self._working_copy_is_stale and self.path.exists():
            shutil.copy2(self.path, self._working_path)
            self._working_copy_is_stale = False
        working_file_is_empty = (
            not self._working_path.exists() or self._working_path.stat().st_size == 0
        )
        if self._is_dirty or working_file_is_empty:
            self._workbook.save(self._working_path)
            self._working_copy_is_stale = False

    def _prepare_com_working_copy(self) -> None:
        """マクロの変更を正常終了まで元ファイルから隔離する作業コピーを用意する。"""
        if self._local_copy_path is not None:
            return
        if self.path.exists():
            # 書き込み経路では UNC かどうかに関係なく作業コピーを用意する（既に使って
            # いるケースに加え、読み取り専用 Excel でマクロを呼ぶ非常経路もここに来る）。
            # 仕様書 4.5 と同じくサイズ無制限でコピーする。
            self._working_path, self._local_copy_path = copy_to_local_if_large(
                self.path, threshold_mb=0
            )
            # ``copy_to_local_if_large(threshold_mb=0)`` は ``(src, None)`` を返す
            # 設計のため、UNC で確実にコピーするために ``_force_local_copy`` で
            # 取り直す。
            if self._local_copy_path is None:
                self._working_path, self._local_copy_path = _force_local_copy(self.path)
            self._working_copy_is_stale = False
            return
        fd, name = tempfile.mkstemp(suffix=self.path.suffix or ".xlsx")
        os.close(fd)
        self._working_path = Path(name)
        self._local_copy_path = self._working_path
        self._working_copy_is_stale = False

    def _reload_workbook(self) -> None:
        """通常 Workbook と stream Workbook を再読込する（COM でのマクロ実行後など）。"""
        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None
        self._close_stream_workbook(self._stream_workbook)
        self._close_stream_workbook(self._stream_workbook_data_only)
        self._stream_workbook = None
        self._stream_workbook_data_only = None
        self._workbook = load_workbook(
            self._working_path,
            keep_vba=self.path.suffix.casefold() in {".xlsm", ".xltm"},
        )

    def _open_stream_workbook(self, *, data_only: bool) -> Workbook:
        """openpyxl の read_only ストリーム Workbook を遅延オープンして返す。

        zip を一度しか読まないよう Workbook をキャッシュし、``close()`` で閉じる。
        ``data_only`` の違いで別インスタンスを保持する。``_reload_workbook``
        のタイミングで破棄される。
        """
        cache = self._stream_workbook_data_only if data_only else self._stream_workbook
        if cache is not None:
            return cache
        workbook = load_workbook(
            self._working_path,
            read_only=True,
            data_only=data_only,
            keep_vba=self.path.suffix.casefold() in {".xlsm", ".xltm"},
        )
        if data_only:
            self._stream_workbook_data_only = workbook
        else:
            self._stream_workbook = workbook
        return workbook

    def _cached_rows(self, sheet_name: str, min_row: int) -> tuple[list[tuple[Any, ...]], bool]:
        """キャッシュ値と数式を並べ、値がない数式だけをCOM昇格対象にする。

        ``read_only`` の Excel で ``_read_computed_rows`` だけを呼ぶ経路では
        通常 Workbook を開かず、openpyxl のストリーム読みだけで値を取り出す。
        数式が1つもないブックでは ``data_only=False`` の Workbook を開かない。
        書き込み後の dirty 時 (``_is_dirty=True``) や作業ファイルがまだ無い
        新規ブックではメモリ上の Workbook を読む（ストリームは古い値しか
        持っていないため）。
        """
        if self._is_dirty or not self._working_path.exists():
            return self._cached_rows_from_memory(sheet_name, min_row)
        return self._cached_rows_from_stream(sheet_name, min_row)

    def _cached_rows_from_memory(
        self, sheet_name: str, min_row: int
    ) -> tuple[list[tuple[Any, ...]], bool]:
        """書き込み直後／新規ブック用: メモリ上の Workbook を読んで値と数式を返す。

        ``Worksheet.iter_rows(values_only=True)`` で行ごとに値を流し、セル単位の
        ``cell(row, column)`` 呼び出しを避けてオブジェクト生成コストを抑える。
        空行（全セルが ``None`` または空文字）はストリーム段階で落とす。
        """
        self._ensure_normal_workbook()
        formula_sheet = self._workbook[sheet_name]
        rows: list[tuple[Any, ...]] = []
        needs_com = False
        for formula_row in formula_sheet.iter_rows(min_row=min_row, values_only=True):
            row_tuple = tuple(formula_row)
            if Excel._row_is_blank(row_tuple):
                # 空行は tuple 化せずにスキップ
                continue
            rows.append(row_tuple)
            for value in row_tuple:
                # OpenPyXL で何かを書いた後は既存キャッシュが残っていても現在の値に
                # 対応する保証がないので、Excel で再計算する。
                if (
                    isinstance(value, str)
                    and value.startswith("=")
                    and (value is None or self._is_dirty)
                ):
                    needs_com = True
        return rows, needs_com

    def _cached_rows_from_stream(
        self, sheet_name: str, min_row: int
    ) -> tuple[list[tuple[Any, ...]], bool]:
        """通常時（read_only 含む）はストリームで読む。

        ``ReadOnlyWorksheet.iter_rows(values_only=True)`` で行ごとに値を流す。
        ``cell(row, column)`` でセルを1つずつ指すと ReadOnlyWorksheet が毎回
        先頭から走査し直すため、行 × 列で二乗になる。イテレータなら前から1度
        だけ流すので線形で済む。

        セル値が ``None`` のときだけ ``data_only=False`` の Workbook を遅延
        オープンして数式かどうかを判定するため、**数式が無いブックでは zip を
        1 度しか読まない**。判定後に改めて数式側も ``iter_rows`` で同じ行から
        流し、同じ位置のセルを突き合わせる。
        """
        cached_workbook = self._open_stream_workbook(data_only=True)
        formula_workbook: Workbook | None = None
        try:
            cached_sheet = cached_workbook[sheet_name]
            rows, any_none = self._collect_cached_rows(cached_sheet, min_row)
            if not any_none:
                return rows, False
            formula_workbook = self._open_stream_workbook(data_only=False)
            formula_sheet = formula_workbook[sheet_name]
            new_rows, needs_com = self._mark_uncalculated_formulas(rows, formula_sheet, min_row)
            return new_rows, needs_com
        finally:
            cached_workbook.close()
            if formula_workbook is not None:
                formula_workbook.close()

    @staticmethod
    def _collect_cached_rows(
        cached_sheet: Worksheet, min_row: int
    ) -> tuple[list[tuple[Any, ...]], bool]:
        """``data_only=True`` の値を ``min_row`` から流し、空行は捨てる。

        「ストリーム段階」で落とす: ``iter_rows`` から yield された行を
        tuple 化して空行（全セルが ``None`` または空文字）か判定し、空なら
        メモリに積まずにスキップする。 ``0`` / ``False`` は値として残す。
        Excel の ``dimension`` が膨らんだブックでも、不要な tuple や dict を
        残さずに線形時間で返せる。
        """
        rows: list[tuple[Any, ...]] = []
        any_none = False
        for cached_row in cached_sheet.iter_rows(min_row=min_row, values_only=True):
            row_tuple = tuple(cached_row)
            if Excel._row_is_blank(row_tuple):
                # 空行はメモリに積まずにスキップ
                continue
            rows.append(row_tuple)
            if not any_none:
                for value in row_tuple:
                    if value is None:
                        any_none = True
                        break
        return rows, any_none

    @staticmethod
    def _mark_uncalculated_formulas(
        rows: list[tuple[Any, ...]],
        formula_sheet: Worksheet,
        min_row: int,
    ) -> tuple[list[tuple[Any, ...]], bool]:
        """数式側ストリームを流し、``rows`` の None セルのうち数式を ``needs_com`` に積む。"""
        new_rows: list[tuple[Any, ...]] = []
        needs_com = False
        for cached_row, formula_row in zip(
            rows, formula_sheet.iter_rows(min_row=min_row, values_only=True), strict=False
        ):
            formula_tuple = tuple(formula_row)
            # cached 側と formula 側で行長が違う場合（末尾の空セル等）に
            # 備えて cached 側に合わせる。
            if len(formula_tuple) < len(cached_row):
                formula_tuple = formula_tuple + (None,) * (len(cached_row) - len(formula_tuple))
            new_row: list[Any] = []
            for cached_value, formula_value in zip(cached_row, formula_tuple, strict=False):
                if (
                    cached_value is None
                    and isinstance(formula_value, str)
                    and formula_value.startswith("=")
                ):
                    # 未計算の数式セル: COM で再計算する
                    needs_com = True
                new_row.append(cached_value)
            new_rows.append(tuple(new_row))
        return new_rows, needs_com

    def _cached_range(
        self, sheet_name: str, min_col: int, min_row: int, max_col: int, max_row: int
    ) -> tuple[list[tuple[Any, ...]], bool]:
        """実テーブル範囲の保存済み計算値と、COM再計算の要否を返す。"""
        if self._is_dirty or not self._working_path.exists():
            self._ensure_normal_workbook()
            formula_sheet = self._workbook[sheet_name]
            rows = [
                tuple(
                    formula_sheet.cell(row=row_number, column=column).value
                    for column in range(min_col, max_col + 1)
                )
                for row_number in range(min_row, max_row + 1)
            ]
            needs_com = any(
                isinstance(value, str) and value.startswith("=") for row in rows for value in row
            )
            return rows, needs_com
        cached_workbook = self._open_stream_workbook(data_only=True)
        formula_workbook: Workbook | None = None
        try:
            cached_sheet = cached_workbook[sheet_name]
            rows = []
            needs_com = False
            for row_number in range(min_row, max_row + 1):
                row_values = []
                for column in range(min_col, max_col + 1):
                    cached_value = cached_sheet.cell(row=row_number, column=column).value
                    if cached_value is None:
                        if formula_workbook is None:
                            formula_workbook = self._open_stream_workbook(data_only=False)
                        formula_value = (
                            formula_workbook[sheet_name].cell(row=row_number, column=column).value
                        )
                        if isinstance(formula_value, str) and formula_value.startswith("="):
                            needs_com = True
                    row_values.append(cached_value)
                rows.append(tuple(row_values))
            return rows, needs_com
        finally:
            cached_workbook.close()
            if formula_workbook is not None:
                formula_workbook.close()

    @staticmethod
    def _is_unc_path(path: str | Path) -> bool:
        """WindowsのUNC表記かを、ファイルの存在確認なしで判定する。"""
        return str(path).replace("/", "\\").startswith("\\\\")

    @staticmethod
    def _vba_digest(path: Path) -> bytes | None:
        """マクロ付き形式に含まれるVBAバイナリのハッシュを返す。"""
        if not path.exists() or path.suffix.casefold() not in {".xlsm", ".xltm"}:
            return None
        try:
            with ZipFile(path) as archive:
                return hashlib.sha256(archive.read("xl/vbaProject.bin")).digest()
        except KeyError:
            return None

    def _is_pristine_workbook(self) -> bool:
        worksheet = cast(Worksheet, self._workbook.active)
        return len(self._workbook.worksheets) == 1 and worksheet["A1"].value is None

    def _is_data_sheet_name(self, name: str) -> bool:
        return name.startswith(self.PY_PREFIX)

    def _with_python_prefix(self, name: str) -> str:
        """利用者が短い名前を書いたとき、Python管理用の名前を補う。"""
        if self._is_data_sheet_name(name):
            return name
        return f"{self.PY_PREFIX}{name}"
