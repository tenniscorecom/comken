"""``comken.services.salesforce_downloader.schedule.create_schedule_template`` を検証する。

雛形の役割は「管理表ブックに ``PY_スケジュール`` シートを追加する」こと。
往復テスト（``load_schedule`` で読み戻せる）、エラーパス、シート属性
（ドロップダウン・フォント）の3点を確かめる。
"""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from comken.core.table import Table
from comken.exceptions import ExcelFileNotFoundError, SheetAlreadyExistsError
from comken.services.salesforce_downloader.schedule import (
    FREQUENCY_HOURLY,
    FREQUENCY_MONTHLY,
    FREQUENCY_WEEKLY,
    SCHEDULE_HEADERS_FULL,
    SCHEDULE_SHEET_NAME,
    create_schedule_template,
    load_schedule,
)
from comken.toolbox.excel import Excel


def _make_master_only(path: Path) -> Path:
    """``PY_管理表`` だけを持つ既存の管理表ブックを作る。"""
    master_headers = [
        "ID",
        "グループ名",
        "担当者",
        "概要",
        "Salesforce URL",
        "保存先",
        "有効",
        "備考",
    ]
    master_row = dict(
        zip(
            master_headers,
            [
                "1001",
                "営業事務グループ",
                "山田",
                "顧客一覧",
                "https://example.com/a/view",
                str(path.parent),
                "○",
                "",
            ],
            strict=True,
        )
    )
    with Excel(path) as book:
        book.create_data_sheet("管理表").create_table("管理表", Table(master_headers, [
            master_row,
        ]))

    return path


class TestAddScheduleSheet:
    """既存の管理表ブックにスケジュールシートを追加する。"""

    def test_round_trip_via_load_schedule(self, tmp_path):
        """雛形の記入例が ``load_schedule`` で ``ScheduleRule`` として読めること。"""
        path = _make_master_only(tmp_path / "管理表.xlsx")
        create_schedule_template(path)

        rules = load_schedule(path)
        # デフォルト記入例は2行
        assert len(rules) == 2
        # 取得頻度の異なる2パターンが入っている（「毎週」と「1時間ごと」）
        frequencies = {rule.frequency for rule in rules}
        assert frequencies == {FREQUENCY_WEEKLY, FREQUENCY_HOURLY}
        # 「毎週」のほうは曜日=月・取得時刻=09:00 が入っているはず
        weekly = next(rule for rule in rules if rule.frequency == FREQUENCY_WEEKLY)
        assert weekly.weekday == 0  # 月曜
        assert weekly.run_time is not None and weekly.run_time.hour == 9
        # 「1時間ごと」のほうは間隔（60分）と開始時刻（取得時刻列=09:00）が埋まる
        hourly = next(rule for rule in rules if rule.frequency == FREQUENCY_HOURLY)
        assert hourly.interval_minutes == 60
        assert hourly.run_time is not None and hourly.run_time.hour == 9

    def test_schedule_sheet_has_full_headers(self, tmp_path):
        """雛形は ``SCHEDULE_HEADERS_FULL`` の全列を持つ（不足列を足させる運用を避ける）。"""
        path = _make_master_only(tmp_path / "管理表.xlsx")
        create_schedule_template(path)

        book = load_workbook(path)
        sheet = book[f"PY_{SCHEDULE_SHEET_NAME}"]
        headers = [cell.value for cell in sheet[1]]
        assert headers == list(SCHEDULE_HEADERS_FULL)

    def test_excel_table_is_created(self, tmp_path):
        """構造化ファイルとして ``PY_T_スケジュール`` が登録されている。"""
        path = _make_master_only(tmp_path / "管理表.xlsx")
        create_schedule_template(path)

        book = load_workbook(path)
        sheet = book[f"PY_{SCHEDULE_SHEET_NAME}"]
        assert "PY_T_スケジュール" in sheet.tables

    def test_font_is_noto_sans_jp(self, tmp_path):
        """スケジュール表の全セル（見出し・記入例）に Noto Sans JP が当たること。"""
        path = _make_master_only(tmp_path / "管理表.xlsx")
        create_schedule_template(path)

        book = load_workbook(path)
        sheet = book[f"PY_{SCHEDULE_SHEET_NAME}"]
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                assert cell.font.name == "Noto Sans JP"

    def test_freeze_panes_is_set(self, tmp_path):
        """見出し行（1行目）がスクロール固定される。"""
        path = _make_master_only(tmp_path / "管理表.xlsx")
        create_schedule_template(path)

        book = load_workbook(path)
        sheet = book[f"PY_{SCHEDULE_SHEET_NAME}"]
        assert sheet.freeze_panes == "A2"

    def test_choice_columns_have_dropdown(self, tmp_path):
        """``choices`` 列には Excel のドロップダウン（入力規則）が付くこと。

        検証対象: 取得頻度・曜日・有効（3列）。``月末指定`` 列が無くなったので
        ドロップダウン対象も 4 列 → 3 列に減る。
        対象外: 祝日対応（自由記述のためドロップダウンを敢えて付けない）。
        """
        path = _make_master_only(tmp_path / "管理表.xlsx")
        create_schedule_template(path)

        book = load_workbook(path)
        sheet = book[f"PY_{SCHEDULE_SHEET_NAME}"]
        validations = list(sheet.data_validations.dataValidation)
        # 3列分のドロップダウンが付く（末尾の 1003 は ``_DATA_VALIDATION_ROWS`` ぶん
        # の余裕）。 ``1行目の見出し``は含めずデータ行から開始する
        formulas = {str(v.sqref): v.formula1 for v in validations}
        assert formulas["C2:C1003"] == '"1時間ごと,毎日,毎週,毎月"'
        assert formulas["D2:D1003"] == '"月,火,水,木,金,土,日"'
        assert formulas["I2:I1003"] == '"○,×"'
        # 祝日対応（H 列）には付かないことを確認
        for v in validations:
            assert not str(v.sqref).startswith("H")

    def test_weekday_and_date_columns_have_conditional_formatting(self, tmp_path):
        """「曜日」「日付」列に、取得頻度に応じたグレーアウトの条件付き書式が付くこと。

        「曜日」列（D 列）は「毎週」でないとき、「日付」列（G 列）は「毎月」でない
        ときにグレーアウトするルールが、それぞれ 1 つずつ存在する。
        動的な入力禁止ではなく見た目だけのヒントなので、データ検証や
        ``protection`` には触らない（``Cell.is_locked`` も変えない）。
        """
        path = _make_master_only(tmp_path / "管理表.xlsx")
        create_schedule_template(path)

        book = load_workbook(path)
        sheet = book[f"PY_{SCHEDULE_SHEET_NAME}"]
        # ``ConditionalFormattingList`` は ``__iter__`` で ``ConditionalFormatting``
        # を1個ずつ yield し、それぞれ ``.rules`` で ``Rule`` リストが取れる
        cf_by_range: dict[str, list] = {}
        for cf in sheet.conditional_formatting:
            cf_by_range[str(cf.sqref)] = list(cf.rules)

        weekday_range = "D2:D1003"
        date_range = "G2:G1003"
        assert weekday_range in cf_by_range, (
            f"「曜日」列に条件付き書式がありません（登録済: {sorted(cf_by_range)}）"
        )
        assert date_range in cf_by_range, (
            f"「日付」列に条件付き書式がありません（登録済: {sorted(cf_by_range)}）"
        )

        weekday_rules = cf_by_range[weekday_range]
        date_rules = cf_by_range[date_range]
        # 「毎週」「毎月」以外ではグレー塗りにしたいので、塗り色が設定された
        # ``FormulaRule`` が1つ以上存在することを確認する
        assert any(_is_gray_fill_formula(rule, FREQUENCY_WEEKLY) for rule in weekday_rules), (
            f"「曜日」列に『{FREQUENCY_WEEKLY}』不一致条件のグレー塗りルールが"
            f"見つかりません（ルール: {weekday_rules}）"
        )
        assert any(_is_gray_fill_formula(rule, FREQUENCY_MONTHLY) for rule in date_rules), (
            f"「日付」列に『{FREQUENCY_MONTHLY}』不一致条件のグレー塗りルールが"
            f"見つかりません（ルール: {date_rules}）"
        )

    def test_missing_master_file_raises(self, tmp_path):
        """存在しないファイルを指定すると ``ExcelFileNotFoundError``（書き込みモードで
        空ブックを新規作成しない＝既存ファイルを保護する）。
        """
        missing = tmp_path / "無い.xlsx"
        assert not missing.exists()
        with pytest.raises(ExcelFileNotFoundError):
            create_schedule_template(missing)
        assert not missing.exists()  # 空ブックも作らない

    def test_existing_schedule_sheet_raises(self, tmp_path):
        """既に ``スケジュール`` シートがあるブックに呼ぶとエラー（上書き防止）。

        ``SheetAlreadyExistsError`` の ``PY_`` プレフィックス付き名がメッセージに
        入っていれば、既存のデータを消そうとしていると気づける。
        """
        path = _make_master_only(tmp_path / "管理表.xlsx")
        # 1回目: 成功
        create_schedule_template(path)
        # 2回目: 失敗
        with pytest.raises(SheetAlreadyExistsError) as e:
            create_schedule_template(path)
        assert "PY_スケジュール" in str(e.value)


class TestGuideSheet:
    """``記入方法`` シートへの追記挙動。"""

    def test_creates_guide_sheet_when_missing(self, tmp_path):
        """``記入方法`` シートが無いブックを渡された場合はエラーにせず新規作成する。"""
        path = _make_master_only(tmp_path / "管理表.xlsx")
        create_schedule_template(path)

        book = load_workbook(path)
        assert "記入方法" in book.sheetnames
        guide = book["記入方法"]
        # スケジュールセクションの見出しが書かれている
        assert guide["A1"].value == "列"
        # スケジュール列の見出し（少なくとも「スケジュール」が見出しセルに）
        # が出現することだけ確認（詳細な行番号は実装に依存しない）
        text = "\n".join(str(cell.value) for row in guide.iter_rows() for cell in row)
        assert "スケジュール" in text
        assert "スケジュールキー" in text
        assert "取得頻度" in text

    def test_appends_to_existing_guide_sheet(self, tmp_path):
        """``記入方法`` シートが既にあるブックでは、既存の内容を壊さず末尾に追記する。"""
        path = _make_master_only(tmp_path / "管理表.xlsx")
        # ``_make_master_only`` で作られたブックに「記入方法」シートを追加する。
        # ``Excel(path) as book:`` の中で ``book.create_sheet(...)`` を経由すれば
        # ``_is_dirty`` が立って保存される。``_workbook.create_sheet`` を直接
        # 叩くと保存されないので避ける
        with Excel(path) as book:
            book.create_sheet("記入方法").write_value("A1", "管理表の案内")

        # 残りは ``openpyxl`` で直接書き足す（Excel を save させてから reopen）
        wb_pre = load_workbook(path)
        guide_pre = wb_pre["記入方法"]
        guide_pre["A3"] = "列"
        guide_pre["B3"] = "何を書くか"
        guide_pre["A4"] = "ID"
        guide_pre["B4"] = "管理番号"
        guide_pre["A5"] = "注意"
        guide_pre["B5"] = "見出しを変えないでください"
        wb_pre.save(path)
        wb_pre.close()

        create_schedule_template(path)

        book = load_workbook(path)
        guide = book["記入方法"]
        text = "\n".join(str(cell.value) for row in guide.iter_rows() for cell in row)
        # 既存の内容は残っている
        assert "管理表の案内" in text
        assert "管理番号" in text
        assert "見出しを変えないでください" in text
        # スケジュールセクションの内容も追記されている
        assert "スケジュール" in text
        assert "スケジュールキー" in text


def _is_gray_fill_formula(rule: object, target_frequency: str) -> bool:
    """``rule`` が「``target_frequency`` でないときグレー塗り」ルールか判定する。

    openpyxl の ``FormulaRule`` は ``formula`` に数式文字列の ``tuple`` を持ち、
    塗り色は ``rule.dxf``（``DifferentialStyle``）の中の ``fill`` に格納される
    （``Rule`` 自体は ``fill`` 属性を持たない）。ここでは数式が ``<>"<頻度>"``
    形式（``<>`` 演算子でターゲット頻度を不一致比較）を含み、かつ
    ``DifferentialStyle`` に塗り色が設定されていることを確認する。具体的な
    色コードまでは見ない（``BFBFBF`` を将来変えてもテストが壊れないように
    するため）。
    """
    formula = getattr(rule, "formula", None)
    if not formula:
        return False
    needle = f'<>"{target_frequency}"'
    if not any(needle in str(text) for text in formula):
        return False
    dxf = getattr(rule, "dxf", None)
    return getattr(dxf, "fill", None) is not None
