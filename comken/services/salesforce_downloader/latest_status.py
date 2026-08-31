"""comken/services/salesforce_downloader/latest_status.py — 全レポートの最新実行結果（Excel）。

`download_scheduled()` のあとに、管理表の全レポートについて履歴 CSV から
最新（実行日時が最大）の行を引いて、1 ファイルへ上書き生成する。
失敗したレポートは `Color.PINK` で塗りつぶす。

**人が読む用の帳票で、プログラムだけが上書きする。** 管理表（Excel）・
履歴（CSV）とは別のファイルにする。理由は `history.py` のモジュール
docstring を参照（書く主体が違うものは分けないと、人が開いている間に
プログラムが保存できず履歴が飛ぶ）。
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from comken.constants import Color
from comken.core.table.model import Table
from comken.core.timer import measure
from comken.services.salesforce_downloader import _paths as _paths_module
from comken.services.salesforce_downloader import history
from comken.services.salesforce_downloader.master import load_master
from comken.toolbox.excel import Excel

# 出力シートの列。順序はこの順でしか読まれない（手で編集しないファイルだが、
# 将来の読み返しを考え `COLUMNS` と同じく先頭で固定する）
_COLUMNS = (
    "管理番号",
    "概要",
    "最新実行日時",
    "成否",
    "原因区分",
    "エラー内容",
)

# 管理表にあるが履歴に1行も無い管理番号の「成否」セルに使う文字列
_NOT_RUN = "未実行"


@measure
def write_latest_status(
    master_path: Path | None = None,
    history_path: Path | None = None,
    output_path: Path | None = None,
) -> None:
    """管理表の全エントリについて、履歴から最新の実行結果を 1 シートへ書き出す。

    管理番号ごとに履歴 CSV の最新行（実行日時が最大）を採用する。**履歴が無い
    管理番号は「未実行」として 1 行出す**——対象から落とすと「一覧にあるのに
    結果が載っていない」という見落としが起きるため。

    失敗した行は `Color.PINK` でセルを塗りつぶす。`report_master.py` の記入例と
    同じやり方（``Excel`` で表を作って保存したあと、``openpyxl.load_workbook``
    で開き直して直接スタイルを当てる）を踏襲する。``Sheet`` はラッパーの
    private 属性を外から触らせる作りになっていないため、スタイル適用は
    ``Excel`` のコンテキストを閉じたあとに行う。

    ``None`` を渡すと `_paths.MASTER_PATH` / `_paths.HISTORY_PATH` /
    `_paths.LATEST_STATUS_PATH` の現在値を使う。**関数定義時のデフォルト値で
    はなく呼び出し時の値を読む**ので、テストが ``_paths`` を monkeypatch
    して既定の保存先を確認できる。

    Args:
        master_path: レポート管理表（Excel）のパス。
        history_path: ダウンロード履歴（CSV）のパス。
        output_path: 生成先の Excel パス。

    Raises:
        ExcelFileNotFoundError: master_path が無い場合（`load_master` から伝播）。
        HistoryHeaderMismatchError: 履歴の見出しが想定と合わない場合。
    """
    master_path = Path(master_path) if master_path is not None else _paths_module.MASTER_PATH
    history_path = Path(history_path) if history_path is not None else _paths_module.HISTORY_PATH
    output_path = Path(output_path) if output_path is not None else _paths_module.LATEST_STATUS_PATH
    entries = load_master(master_path)
    latest_by_key = _latest_rows_by_key(history_path)

    body: list[dict[str, str]] = []
    failure_row_numbers: list[int] = []  # Excel の行番号（見出しが 1 行目）
    for key, entry in entries.items():
        latest = latest_by_key.get(key)
        if latest is None:
            body.append(
                {
                    "管理番号": key,
                    "概要": entry.summary,
                    "最新実行日時": "",
                    "成否": _NOT_RUN,
                    "原因区分": "",
                    "エラー内容": "",
                }
            )
            continue
        succeeded = latest["成否"] != history.FAILURE
        body.append(
            {
                "管理番号": key,
                "概要": entry.summary,
                "最新実行日時": latest["実行日時"],
                "成否": latest["成否"],
                "原因区分": latest["原因区分"],
                "エラー内容": latest["エラー内容"],
            }
        )
        if not succeeded:
            # 失敗の塗りつぶし対象。Excel 行番号は見出しが 1 行目、データは 2 行目から
            failure_row_numbers.append(len(body) + 1)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    table = Table(list(_COLUMNS), body)
    with Excel(output_path) as book:
        sheet = book.create_data_sheet("最新ステータス")
        sheet.create_table("最新ステータス", table)

    if failure_row_numbers:
        # `create_data_sheet` は名前に `PY_` を補う（`Excel`/`Sheet` の共通規約）。
        # report_master.py の雛形生成と同じく、保存済みファイルを開き直して
        # スタイルだけを当てる（Sheet の private 属性には触れない）
        workbook = load_workbook(output_path)
        worksheet = workbook["PY_最新ステータス"]
        fill = PatternFill("solid", fgColor=Color.PINK)
        for row_number in failure_row_numbers:
            for column in range(1, len(_COLUMNS) + 1):
                worksheet.cell(row=row_number, column=column).fill = fill
        workbook.save(output_path)
        workbook.close()


def _latest_rows_by_key(history_path: Path) -> dict[str, dict[str, str]]:
    """履歴から、管理番号ごとに「実行日時が最大」の行を返す。

    実行日時の書式は ``%Y-%m-%d %H:%M:%S`` で固定（``history.record`` がそう
    書いている）なので、文字列のまま大小比較できる。空文字は無視する。
    """
    latest: dict[str, dict[str, str]] = {}
    for row in history.read_all(history_path):
        timestamp = row.get("実行日時", "")
        key = row.get("管理番号", "")
        if not timestamp or not key:
            continue
        existing = latest.get(key)
        if existing is None or timestamp > existing["実行日時"]:
            latest[key] = row
    return latest
