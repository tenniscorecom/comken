"""examples/ のオフラインサンプルが実際に動くことを確認するテスト。

外部システム（Excel COM・ブラウザ）を必要としないサンプルを、
出力先を tmp_path に差し替えて main() まで通し、成果物ができることを検証する。
README・ドキュメントが「そのまま動く」と案内している主張をテストで担保する。
"""

import pytest
from openpyxl import load_workbook


class TestBasicExamples:
    def test_csv_read(self, tmp_path, monkeypatch):
        from examples import csv_read

        csv_path = tmp_path / "受注明細.csv"
        monkeypatch.setattr(csv_read, "CSV_PATH", csv_path)
        csv_read.main()

        assert csv_path.exists()
        assert "株式会社アルファ" in csv_path.read_text(encoding="utf-8")

    def test_csv_write(self, tmp_path, monkeypatch):
        from examples import csv_write

        output_path = tmp_path / "作業記録.csv"
        monkeypatch.setattr(csv_write, "OUTPUT_PATH", output_path)
        csv_write.main()

        assert output_path.exists()
        assert len(output_path.read_text(encoding="utf-8-sig").splitlines()) == 5

    def test_excel_read(self, tmp_path, monkeypatch):
        from examples import excel_read

        excel_path = tmp_path / "在庫一覧.xlsx"
        monkeypatch.setattr(excel_read, "EXCEL_PATH", excel_path)
        excel_read.main()

        assert load_workbook(excel_path, read_only=True)["PY_在庫"].max_row == 3

    def test_excel_write(self, tmp_path, monkeypatch):
        from examples import excel_write

        output_path = tmp_path / "売上帳票.xlsx"
        monkeypatch.setattr(excel_write, "OUTPUT_PATH", output_path)
        excel_write.main()

        worksheet = load_workbook(output_path)[excel_write.SHEET_NAME]
        assert worksheet.freeze_panes == "A2"
        assert worksheet["D4"].value == 4300

    def test_column_mapping(self, tmp_path, monkeypatch):
        from examples import column_mapping

        monkeypatch.setattr(column_mapping, "OUTPUT_FOLDER", tmp_path)
        monkeypatch.setattr(column_mapping, "CONFIG_PATH", tmp_path / "config.ini")
        monkeypatch.setattr(column_mapping, "SOURCE_CSV", tmp_path / "受注.csv")
        monkeypatch.setattr(column_mapping, "OUTPUT_PATH", tmp_path / "請求一覧.xlsx")
        column_mapping.main()

        worksheet = load_workbook(column_mapping.OUTPUT_PATH)["PY_請求一覧"]
        assert worksheet["B2"].value == "株式会社アルファ"
        assert worksheet["C2"].value == 12000

    def test_state(self, tmp_path, monkeypatch):
        from examples import state

        state_path = tmp_path / "state.ini"
        monkeypatch.setattr(state, "STATE_PATH", state_path)
        state.main()
        state.main()

        assert "1002" in state_path.read_text(encoding="utf-8")

    def test_logger(self, tmp_path, monkeypatch):
        from examples import logger

        monkeypatch.setattr(logger, "LOG_FOLDER", tmp_path / "logs")
        logger.main()

    def test_runtime(self, tmp_path, monkeypatch):
        from examples import runtime

        monkeypatch.setattr(runtime, "OUTPUT_FOLDER", tmp_path)
        monkeypatch.setattr(runtime, "SOURCE_PATH", tmp_path / "source.txt")
        monkeypatch.setattr(runtime, "DRY_RUN_PATH", tmp_path / "dry-run.txt")
        monkeypatch.setattr(runtime, "ACTUAL_PATH", tmp_path / "actual.txt")
        runtime.main()

        assert not runtime.DRY_RUN_PATH.exists()
        assert runtime.ACTUAL_PATH.exists()

    def test_files(self, tmp_path, monkeypatch):
        from examples import files

        monkeypatch.setattr(files, "OUTPUT_FOLDER", tmp_path)
        monkeypatch.setattr(files, "ARCHIVE_PATH", tmp_path / "日次資料.zip")
        files.main()

        assert files.ARCHIVE_PATH.exists()
        assert len(list((tmp_path / "展開").glob("*.csv"))) == 2

    def test_utils(self):
        from examples import utils

        utils.main()

    def test_constants(self, tmp_path, monkeypatch):
        from examples import constants

        monkeypatch.setattr(constants, "OUTPUT_FOLDER", tmp_path)
        monkeypatch.setattr(constants, "CSV_PATH", tmp_path / "名簿.csv")
        constants.main()

        assert constants.CSV_PATH.exists()

    def test_exceptions(self, tmp_path, monkeypatch):
        from examples import exceptions

        monkeypatch.setattr(exceptions, "CSV_PATH", tmp_path / "例外確認.csv")
        exceptions.main()


class TestCsvToExcelReport:
    def test_creates_report(self, tmp_path, monkeypatch):
        """CSV を読んで Excel レポートを作る例が xlsx を出力する。"""
        from examples.advanced.csv_to_excel_report import run

        monkeypatch.setattr(run, "OUTPUT_FOLDER", tmp_path)
        run.main()

        outputs = list(tmp_path.glob("*.xlsx"))
        assert len(outputs) == 1
        # 合計行まで書けている（ヘッダー + データ3件 + 合計 = 5行以上）
        ws = load_workbook(outputs[0]).active
        assert ws.max_row >= 5
        assert ws.cell(row=ws.max_row, column=1).value == "合計"


class TestExcelKeyTransfer:
    @pytest.fixture
    def transferred(self, tmp_path, monkeypatch):
        """サンプルを実行し、転記後のシートを {注文番号: 行} で返す。"""
        from examples.advanced.excel_key_transfer import run

        monkeypatch.setattr(run, "OUTPUT_FOLDER", tmp_path)
        monkeypatch.setattr(run, "MASTER_CSV", tmp_path / "master.csv")
        monkeypatch.setattr(run, "DETAIL_CSV", tmp_path / "detail.csv")
        monkeypatch.setattr(run, "INVOICE_XLSX", tmp_path / "invoice.xlsx")
        run.main()

        ws = load_workbook(tmp_path / "invoice.xlsx").active
        return {r[0]: r for r in ws.iter_rows(min_row=2, values_only=True)}

    def test_transfers_matched_rows(self, transferred):
        """マスタにあるキーだけ転記される。"""
        assert transferred["A001"][1] == "株式会社アルファ"
        assert "Z999" not in transferred

    def test_sums_multiple_detail_rows(self, transferred):
        """1対多の明細は合計して転記される（後の行で上書きしない）。"""
        # A001 は 48000 + 12000 + 1500。index() で引くと最後の 1500 になってしまう
        assert transferred["A001"][2] == 61500
        assert transferred["A002"][2] == 18000

    def test_amount_is_number_not_text(self, transferred):
        """金額が文字列でなく数値で入る（Excel 側で集計できる）。"""
        assert isinstance(transferred["A001"][2], int)


class TestCsvDiffReport:
    def test_detects_added_removed_changed(self, tmp_path, monkeypatch):
        """差分レポートの例が追加・削除・変更を検出して xlsx を出す。"""
        from examples.advanced.csv_diff_report import run

        monkeypatch.setattr(run, "OUTPUT_FOLDER", tmp_path)
        monkeypatch.setattr(run, "YESTERDAY_CSV", tmp_path / "yesterday.csv")
        monkeypatch.setattr(run, "TODAY_CSV", tmp_path / "today.csv")
        run.main()

        outputs = list(tmp_path.glob("*.xlsx"))
        assert len(outputs) == 1
        statuses = {
            row[0]
            for row in load_workbook(outputs[0]).active.iter_rows(min_row=2, values_only=True)
        }
        # 追加(004)・削除(003)・変更(002) がそれぞれ検出されている
        assert {"追加", "削除", "変更"} <= statuses


class TestCsvDateMove:
    def test_moves_only_file_with_matching_date(self, tmp_path):
        """指定列とファイル名の日付が一致する CSV だけを移動する。"""
        from examples.advanced.csv_date_move.run import move_matching_files

        input_folder = tmp_path / "input"
        output_folder = tmp_path / "output"
        input_folder.mkdir()
        matching = input_folder / "売上_20260729.csv"
        mismatching = input_folder / "売上_20260730.csv"
        matching.write_text("日付\n2026/07/29\n", encoding="utf-8")
        mismatching.write_text("日付\n2026/07/29\n", encoding="utf-8")

        result = move_matching_files(input_folder, output_folder, "日付", "%Y/%m/%d", "*.csv")

        assert result == (1, 1)
        assert (output_folder / matching.name).exists()
        assert not matching.exists()
        assert mismatching.exists()
        assert not (output_folder / mismatching.name).exists()


class TestTableTransferDesign:
    def test_writes_invoice_with_skip_and_unmatched(self, tmp_path, monkeypatch):
        """Transfer.matched_rows / unmatched / apply_mapping を使ったサンプルが
        Excel 出力を作る。matched_rows で continue した行は除かれ、
        read にしか無い行は新規追加、write にしか無い行は「転記元に無し」と
        印が付けられる。
        """
        from examples.advanced.table_transfer_design import run

        monkeypatch.setattr(run, "OUTPUT_FOLDER", tmp_path)
        monkeypatch.setattr(run, "SOURCE_CSV", tmp_path / "受注.csv")
        monkeypatch.setattr(run, "OUTPUT_PATH", tmp_path / "請求一覧.xlsx")
        run.main()

        assert (tmp_path / "請求一覧.xlsx").exists()
        ws = load_workbook(tmp_path / "請求一覧.xlsx")["PY_請求一覧"]
        # A001: 通常転記 / A003: 新規追加 / A099: 転記元に無し
        # A002 は continue でスキップされ備考が空欄、filter で落ちる
        rows = {r[0]: r for r in ws.iter_rows(min_row=2, values_only=True)}
        assert set(rows) == {"A001", "A003", "A099"}
        assert rows["A001"][1] == "株式会社アルファ"
        assert rows["A001"][2] == 12000
        assert rows["A001"][3] == "消費税: 1200"
        assert rows["A003"][3] == "新規追加"
        assert rows["A099"][3] == "転記元に無し"
