"""
excel/sheet.py — ワークシートの高レベルラッパー

ExcelWriter.sheet() から取得し、セル書き込み・行書き込み・列幅調整・
ヘッダー固定などをシート単位で行う（sheet_name を毎回渡さなくてよい）。

使い方:
    from comken.csv import CsvReader
    from comken.excel import ExcelWriter

    # 既存ファイルを開いて書き込む
    with ExcelWriter("report.xlsx") as f:
        s = f.sheet("Sheet1")
        s["A1"] = "売上レポート"          # セル参照で書き込み
        title = s["A1"]                   # セル参照で読み取り
        s.write_row(3, ["日付", "金額"])  # 3行目に横並びで書く
        s.append_row(["2026-07-12", 1000])  # 最終行の下に追記
        s.auto_width()                    # 列幅を内容に合わせる（日本語対応）
        s.freeze_header()                 # 1行目を固定
        f.save()

    # 新規ブックを作ってレポートを出力する
    rows = CsvReader("data.csv").rows()
    with ExcelWriter.create(r"C:\\作業\\report.xlsx") as f:
        s = f.sheet("Sheet1")
        s.write_table(rows)               # ヘッダー行 + データ行をまとめて書く
        s.auto_width()
        s.freeze_header()
        f.save()
"""

import logging
import re

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from ..exceptions import (
    InvalidTableNameError,
    TableAlreadyExistsError,
    TableNotFoundError,
    _warn_coerce,
)
from ..utils.data import col_to_num, column_number
from ..utils.timer import measure

logger = logging.getLogger(__name__)

DEFAULT_TABLE_STYLE = "TableStyleMedium2"
_CELL_REFERENCE = re.compile(r"^[A-Z]{1,3}[1-9]\d*$", re.IGNORECASE)
_R1C1_REFERENCE = re.compile(r"^R[1-9]\d*C[1-9]\d*$", re.IGNORECASE)
_TABLE_NAME = re.compile(r"^(?:[^\W\d]|[\\_])[\w.]*$", re.UNICODE)


class Sheet:
    """1枚のワークシートのラッパー。ExcelWriter.sheet() から取得する。

    ここにないシート操作は .ws から生の openpyxl Worksheet を使える。
    """

    def __init__(self, ws: Worksheet) -> None:
        self.ws = ws  # 生の openpyxl Worksheet（高度な操作用に公開）
        if not hasattr(ws, "_comken_written_bounds"):
            ws._comken_written_bounds = None

    # ------------------------------------------------------------ セル参照
    def __getitem__(self, ref: str):
        """セル参照で値を読む（例: s["A1"]）。"""
        return self.ws[ref].value

    def __setitem__(self, ref: str, value) -> None:
        """セル参照で値を書く（例: s["A1"] = "タイトル"）。"""
        self.ws[ref] = value
        row, col = coordinate_to_tuple(ref)
        self._record_written_range(row, col, row, col)

    def write_cell(self, row: int, col: int | str, value) -> None:
        """行番号と列番号・列記号を指定してセルに値を書き込む。"""
        row_num = int(row)
        col_num = column_number(col)
        self.ws.cell(row=row_num, column=col_num).value = value
        self._record_written_range(row_num, col_num, row_num, col_num)

    # ------------------------------------------------------------ 行の書き込み
    def write_row(self, row: int, values: list, start_col: int = 1) -> None:
        """1行に値を横並びで書き込む。

        Args:
            row: 行番号（1始まり）。
            values: 書き込む値のリスト（左から順に並ぶ）。
            start_col: 開始列番号（1始まり。デフォルト: A列から）。
        """
        for i, value in enumerate(values):
            self.ws.cell(row=int(row), column=start_col + i).value = value
        if values:
            self._record_written_range(int(row), start_col, int(row), start_col + len(values) - 1)

    def write_rows(self, start_row: int, rows: list[list], start_col: int = 1) -> None:
        """複数行をまとめて書き込む。

        Args:
            start_row: 開始行番号（1始まり）。
            rows: 行のリスト（値のリストのリスト）。
            start_col: 開始列番号（1始まり）。
        """
        for i, values in enumerate(rows):
            self.write_row(int(start_row) + i, values, start_col)

    def append_row(self, values: list) -> None:
        """最終行の下に1行追記する（空シートなら1行目に書く）。"""
        self.write_row(self.last_row + 1 if not self.is_empty else 1, values)

    def write_table(
        self, rows: list[dict], start_row: int = 1, headers: list[str] | None = None
    ) -> None:
        """ヘッダー行 + データ行の値を書き込む（構造化テーブルにはしない）。

        CsvReader.rows() や read_rows_as_dicts() の結果をそのまま渡せる。
        Excel の構造化テーブルにする場合は、書き込み後に add_table() を呼ぶ。

        Args:
            rows: 辞書のリスト（キーが列名になる）。
            start_row: ヘッダー行の行番号（1始まり）。
            headers: 列の並び順。省略すると最初の行のキー順。
        """
        if not rows:
            return
        headers = headers or list(rows[0].keys())
        self.write_row(int(start_row), headers)
        for i, row in enumerate(rows, start=int(start_row) + 1):
            self.write_row(i, [row.get(h, "") for h in headers])

    @measure
    def transfer_by_key(
        self,
        key_col: int | str,
        lookup: dict[str, dict],
        column_mapping: dict[str, str],
        start_row: int = 2,
    ) -> int:
        """キー列の値で lookup を引き、一致した行へ値を転記する。"""
        key_col_num = column_number(key_col)
        mapping = {col_to_num(letter): name for letter, name in column_mapping.items()}
        logger.info("シート「%s」: 最終行 %d行", self.ws.title, self.ws.max_row)
        matched = 0
        for row in range(int(start_row), self.ws.max_row + 1):
            key_value = self.ws.cell(row=row, column=key_col_num).value
            if key_value is None or str(key_value).strip() == "":
                continue
            if isinstance(key_value, float) and key_value.is_integer():
                key_value = int(key_value)
            lookup_row = lookup.get(str(key_value).strip())
            if lookup_row is None:
                logger.debug("%d行目: キー「%s」が lookup に存在しません", row, key_value)
                continue
            for col_num, name in mapping.items():
                self.ws.cell(row=row, column=col_num).value = lookup_row.get(name, "")
            if mapping:
                self._record_written_range(row, min(mapping), row, max(mapping))
            matched += 1
        logger.info("転記完了: %d件一致（シート: %s）", matched, self.ws.title)
        return matched

    # ------------------------------------------------------------ 構造化テーブル
    def add_table(self, name: str, ref: str) -> None:
        """指定範囲を Excel の構造化テーブルにする。

        write_table() は値だけを書き、このメソッドは既存の値をテーブルにする。
        スタイルを変えたい場合は .ws から openpyxl を直接使用する。
        """
        _validate_table_name(name)
        workbook_table_names = [
            table_name for worksheet in self.ws.parent.worksheets for table_name in worksheet.tables
        ]
        if name.casefold() in {table_name.casefold() for table_name in workbook_table_names}:
            raise TableAlreadyExistsError(name)
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name=DEFAULT_TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        self.ws.add_table(table)

    def append_to_table(self, name: str, rows: list[dict]) -> None:
        """構造化テーブルの末尾にデータ行を追記する。

        openpyxl は計算列を自動入力しない。数式の列がある場合は、
        ``{"税込": "=[@金額]*1.1"}`` のように行データへ数式文字列を含める。
        ``[@列名]`` の構造化参照はテーブル内のセルでのみ有効。
        """
        if not rows:
            return
        table = self._table(name)
        min_col, header_row, max_col, last_row = range_boundaries(table.ref)
        headers = [
            self.ws.cell(row=header_row, column=col).value for col in range(min_col, max_col + 1)
        ]
        _validate_table_rows(rows, headers)

        for row_number, row in enumerate(rows, start=last_row + 1):
            self.write_row(row_number, [row.get(header, "") for header in headers], min_col)

        # NOTE: 範囲を広げないと追記行はただのセルになり、構造化参照や集計に入らない。
        new_last_row = last_row + len(rows)
        table.ref = (
            f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{new_last_row}"
        )

    def clear_table(self, name: str) -> None:
        """構造化テーブルのデータ行だけを消す（見出し行は残す）。"""
        table = self._table(name)
        min_col, header_row, max_col, last_row = range_boundaries(table.ref)
        if last_row > header_row:
            for row in self.ws.iter_rows(
                min_row=header_row + 1,
                max_row=last_row,
                min_col=min_col,
                max_col=max_col,
            ):
                for cell in row:
                    cell.value = None
            self._record_written_range(header_row + 1, min_col, last_row, max_col)

        # NOTE: 0件でも見出し行を範囲として残さないと、テーブル定義が壊れる。
        table.ref = (
            f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{header_row}"
        )

    def replace_table(self, name: str, rows: list[dict]) -> None:
        """構造化テーブルのデータ行をすべて入れ替える。

        openpyxl は計算列を自動入力しない。数式の列がある場合は、
        ``{"税込": "=[@金額]*1.1"}`` のように行データへ数式文字列を含める。
        ``[@列名]`` の構造化参照はテーブル内のセルでのみ有効。
        """
        table = self._table(name)
        min_col, header_row, max_col, _ = range_boundaries(table.ref)
        headers = [
            self.ws.cell(row=header_row, column=col).value for col in range(min_col, max_col + 1)
        ]
        _validate_table_rows(rows, headers)
        self.clear_table(name)
        self.append_to_table(name, rows)

    def _table(self, name: str) -> Table:
        if name not in self.ws.tables:
            raise TableNotFoundError(name, _table_names(self.ws))
        return self.ws.tables[name]

    # ------------------------------------------------------------ 見た目の調整
    def set_fill(self, row: int, col: int | str, color: str) -> None:
        """セルの背景色を16進 RGB（# なし）で設定する。"""
        color = _warn_coerce(color, str, "color", stacklevel=2)
        self.ws.cell(row=int(row), column=column_number(col)).fill = PatternFill(
            fill_type="solid", fgColor=color
        )

    def set_column_width(self, col: int | str, width: float) -> None:
        """列番号または列記号を指定して列幅を設定する。"""
        letter = get_column_letter(column_number(col))
        self.ws.column_dimensions[letter].width = float(width)

    def set_number_format(self, row: int, col: int | str, fmt: str) -> None:
        """セルの数値フォーマットを設定する。"""
        fmt = _warn_coerce(fmt, str, "fmt", stacklevel=2)
        self.ws.cell(row=int(row), column=column_number(col)).number_format = fmt

    def set_bold(self, row: int, col: int | str, bold: bool = True) -> None:
        """セルの太字を設定または解除する。"""
        self.ws.cell(row=int(row), column=column_number(col)).font = Font(bold=bool(bold))

    def auto_width(self, min_width: float = 8, max_width: float = 60) -> None:
        """全列の幅を内容に合わせて調整する（全角文字は2文字ぶんで計算）。

        Args:
            min_width: 最小の列幅（内容が短くても これより狭くしない）。
            max_width: 最大の列幅（長文があっても これより広げない）。
        """
        for col_num, column in enumerate(self.ws.iter_cols(), start=1):
            width = max(
                (_display_width(cell.value) for cell in column if cell.value is not None),
                default=0,
            )
            letter = get_column_letter(col_num)
            self.ws.column_dimensions[letter].width = min(max(width + 2, min_width), max_width)

    def freeze_header(self, rows: int = 1) -> None:
        """ヘッダー行を固定する（スクロールしても見出しが消えない）。

        Args:
            rows: 固定する行数（デフォルト: 1行目のみ）。
        """
        self.ws.freeze_panes = f"A{int(rows) + 1}"

    # ------------------------------------------------------------ 状態の取得
    @property
    def last_row(self) -> int:
        """データがある最終行の番号（1始まり）。空シートでも 1 が返る点に注意。"""
        return self.ws.max_row

    @property
    def is_empty(self) -> bool:
        """シートに値が1つもないか返す。"""
        if self.ws.max_row > 1 or self.ws.max_column > 1:
            return False
        return self.ws.cell(row=1, column=1).value is None

    def _written_range(self) -> str | None:
        bounds = self.ws._comken_written_bounds
        if bounds is None:
            return None
        min_row, min_col, max_row, max_col = bounds
        return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    def _record_written_range(self, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
        # NOTE: 飛び地も包む矩形にすることで、数万セルでもセル単位の記録を避ける。
        bounds = self.ws._comken_written_bounds
        if bounds is None:
            self.ws._comken_written_bounds = (min_row, min_col, max_row, max_col)
            return
        old_min_row, old_min_col, old_max_row, old_max_col = bounds
        self.ws._comken_written_bounds = (
            min(old_min_row, min_row),
            min(old_min_col, min_col),
            max(old_max_row, max_row),
            max(old_max_col, max_col),
        )


def _display_width(value) -> int:
    """列幅計算用の表示幅を返す（全角文字は2文字ぶんとして数える）。"""
    return sum(2 if ord(ch) > 0xFF else 1 for ch in str(value))


def _validate_table_name(name: str) -> None:
    if (
        len(name) > 255
        or not _TABLE_NAME.fullmatch(name)
        or _CELL_REFERENCE.fullmatch(name)
        or _R1C1_REFERENCE.fullmatch(name)
    ):
        raise InvalidTableNameError(name)


def _table_names(ws: Worksheet) -> list[str]:
    return list(ws.tables)


def _validate_table_rows(rows: list[dict], headers: list) -> None:
    header_set = set(headers)
    for row_number, row in enumerate(rows, start=1):
        unknown_keys = [key for key in row if key not in header_set]
        if unknown_keys:
            raise ValueError(
                f"{row_number}件目のデータにテーブルの見出しにないキーがあります: "
                f"{unknown_keys}。使用できる見出し: {headers}。"
                "キー名を見出しに合わせてください。"
            )
