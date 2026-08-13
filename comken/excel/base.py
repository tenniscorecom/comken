"""comken/excel/base.py — ExcelReader と ExcelWriter に共通するワークブック操作。

ブックを開く・閉じる、シートの存在確認、行を読み取るといった共通処理を持つ。
読み取り専用の ExcelReader と書き込み用の ExcelWriter がこの基底クラスを継承する。
数式の計算結果が必要な場合は、内部で win32com（pywin32）にフォールバックする。
"""

import logging
import shutil
import tempfile
from collections.abc import Generator
from contextlib import closing
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..exceptions import (
    EmptyHeaderCellError,
    ExcelFileNotFoundError,
    ExcelHeadersTooFewError,
    SheetNotFoundError,
    _warn_coerce,
)
from ..utils.files.base import FileBase
from ..utils.timer import measure

logger = logging.getLogger(__name__)


class ExcelBase(FileBase):
    """ExcelReader と ExcelWriter の共通基盤。

    with 文によるブックのクローズ、シートの存在確認、行の読み取りを受け持つ。
    書き込み・書式設定は Sheet、保存は ExcelWriter が提供する。
    数式の計算結果を読む read_computed_rows() は、必要な場合に openpyxl から
    win32com（pywin32）へ自動的にフォールバックする。
    利用例は用途に応じて ExcelReader または ExcelWriter の docstring を参照する。
    """

    SUFFIXES = (".xlsx", ".xlsm", ".xltx", ".xltm")

    def __init__(
        self,
        path: str | Path,
        data_only: bool = False,
        read_only: bool = False,
        local_copy_threshold_mb: float = 10,
        headers: list[str] | None = None,
    ) -> None:
        """
        Args:
            path: Excel ファイルのパス。
            data_only: True にすると数式セルのキャッシュ値を読む（read_computed_rows 推奨）。
            read_only: True にすると読み取り専用で開く（大きなファイルで高速化）。
            local_copy_threshold_mb: この MB 以上のファイルはローカルにコピーしてから開く。
                NAS・ネットワークドライブのファイルが遅い・不安定な場合に有効。
                0 を指定するとローカルコピーを無効化できる。
            headers: ヘッダー行がない Excel の場合に、列名のリストをここで付ける。
                     指定すると read_rows_as_dicts() は全行をデータとして読む。
                     例: ExcelReader("data.xlsx", headers=["注文番号", "金額", "担当者"])
        """
        # save() の保存先は常に元のファイル（ローカルコピーに保存すると close で消えてしまう）
        super().__init__(path)
        self._original_path = self._path
        src = self._original_path

        # 素の FileNotFoundError ではなく、対処法つきの ExcelError にする
        if not src.exists():
            raise ExcelFileNotFoundError(src)

        # ── NAS・ネットワークファイルのローカルコピー ──────────────────
        # 社内ルールでローカルへのコピーが不可の場合は、
        # このブロックを丸ごと削除し「self._tmp = None」だけ残す。
        # close() 内の対応する削除ブロックも併せて削除すること。
        self._tmp = None
        if local_copy_threshold_mb and src.stat().st_size > local_copy_threshold_mb * 1024 * 1024:
            # NOTE: openpyxl がパスから開けるよう、名前を確保して即座に閉じる。
            tmp = tempfile.NamedTemporaryFile(suffix=src.suffix, delete=False)  # noqa: SIM115
            self._tmp = Path(tmp.name)
            tmp.close()
            shutil.copy2(src, self._tmp)
            src = self._tmp
        # ────────────────────────────────────────────────────────────────

        self._working_path = src
        self._headers = headers
        # マクロ入りブック（.xlsm/.xlsb）は keep_vba=True で開かないと save() で VBA が消える
        keep_vba = self._original_path.suffix.lower() in (".xlsm", ".xlsb", ".xltm")
        self._wb: Workbook = load_workbook(
            self._working_path, data_only=data_only, read_only=read_only, keep_vba=keep_vba
        )

    def __enter__(self) -> "ExcelBase":
        return self

    def __exit__(self, *args) -> None:
        self.close()

    def _sheet(self, name: str) -> Worksheet:
        """シートオブジェクトを返す。

        Args:
            name: シート名。

        Raises:
            SheetNotFoundError: 指定したシートが存在しない場合。
        """
        name = _warn_coerce(name, str, "sheet_name", stacklevel=3)
        if name not in self._wb.sheetnames:
            raise SheetNotFoundError(name, self._wb.sheetnames)
        return self._wb[name]

    @measure
    def read_rows(self, sheet_name: str, min_row: int = 2) -> list[tuple]:
        """指定シートの行データをタプルのリストで返す。

        Args:
            sheet_name: シート名。
            min_row: 読み始める行番号（デフォルト: 2 でヘッダーをスキップ）。

        Returns:
            各行を値のタプルにしたリスト。
        """
        return list(self._sheet(sheet_name).iter_rows(min_row=int(min_row), values_only=True))

    @measure
    def read_rows_as_dicts(self, sheet_name: str, header_row: int = 1) -> list[dict]:
        """ヘッダー行をキーとした辞書のリストで返す。

        ヘッダー行がないファイルは ExcelReader(path, headers=[...]) で列名を指定すること。

        Args:
            sheet_name: シート名。
            header_row: ヘッダーが存在する行番号（デフォルト: 1）。
                        __init__ で headers を指定した場合は無視される。

        Returns:
            [{"列名": 値, ...}, ...] の形式のリスト。全セルが空の行は除外される。

        Raises:
            SheetNotFoundError: 指定したシートが存在しない場合。
            ExcelError: ヘッダー行に空のセルがある場合（headers 未指定時のみ）、
                        または headers の列数がシートの列数より少ない場合。
        """
        ws = self._sheet(sheet_name)
        if self._headers is not None:
            all_rows = list(ws.iter_rows(min_row=1, values_only=True))
            if all_rows and len(all_rows[0]) > len(self._headers):
                raise ExcelHeadersTooFewError(len(self._headers), len(all_rows[0]))
            return [
                # headers が実データ列より多い場合は、従来どおり余った見出しを含めない。
                dict(zip(self._headers, row, strict=False))
                for row in all_rows
                if not all(c is None for c in row)
            ]
        all_rows = list(ws.iter_rows(min_row=int(header_row), values_only=True))
        if not all_rows:
            return []
        file_headers = all_rows[0]
        if all(h is None for h in file_headers):
            return []
        none_cols = [i + 1 for i, h in enumerate(file_headers) if h is None]
        if none_cols:
            raise EmptyHeaderCellError(none_cols)
        # openpyxl の行幅差は末尾空セルによるため、見出しとの対応範囲だけ辞書化する。
        return [dict(zip(file_headers, row, strict=False)) for row in all_rows[1:]]

    def iter_rows(
        self, sheet_name: str, min_row: int = 2
    ) -> Generator[tuple[Any, ...], None, None]:
        """大量データ向け。行をジェネレータで1行ずつ返す（メモリ効率優先）。

        read_rows はファイル全体をメモリに乗せるため、数万行以上のファイルでは
        この メソッドを使って1行ずつ処理する。

        複数ファイルを同時に処理する場合（目安: 10ファイル以上）は
        concurrent.futures.ThreadPoolExecutor を使うとさらに高速化できる。

        Args:
            sheet_name: シート名。
            min_row: 読み始める行番号（デフォルト: 2 でヘッダーをスキップ）。

        Yields:
            各行の値のタプル。
        """
        self._sheet(sheet_name)  # シート名の存在チェック（間違いを分かりやすいエラーにする）
        # openpyxl の Workbook は with に対応していないため closing で包む。
        # read_only=True はファイルを開いたままにするので、閉じ忘れると次の処理が失敗する。
        with closing(load_workbook(self._working_path, data_only=True, read_only=True)) as wb:
            yield from wb[str(sheet_name)].iter_rows(min_row=int(min_row), values_only=True)

    @measure
    def read_computed_rows(self, sheet_name: str, min_row: int = 2) -> list[tuple]:
        """数式の計算結果を含む行を読む。

        openpyxl で数式の位置とキャッシュ値を別々に読み、数式セルの
        キャッシュ値がない場合だけ win32com（pywin32）にフォールバックする。
        数式がない場合や、すべての数式にキャッシュ値がある場合は Excel を起動しない。

        Args:
            sheet_name: シート名。
            min_row: 読み始める行番号（デフォルト: 2 でヘッダーをスキップ）。

        Returns:
            各行を値のタプルにしたリスト。数式は計算後の値になっている。

        Raises:
            SheetNotFoundError: 指定したシートが存在しない場合。
        """
        self._sheet(sheet_name)  # シート名の存在チェック（間違いを分かりやすいエラーにする）
        try:
            with closing(
                load_workbook(self._working_path, data_only=False, read_only=True)
            ) as formula_wb:
                formula_rows = list(
                    formula_wb[str(sheet_name)].iter_rows(min_row=int(min_row), values_only=True)
                )

            with closing(
                load_workbook(self._working_path, data_only=True, read_only=True)
            ) as value_wb:
                rows = list(
                    value_wb[str(sheet_name)].iter_rows(min_row=int(min_row), values_only=True)
                )

            needs_calculation = any(
                isinstance(formula_cell, str)
                and formula_cell.startswith("=")
                and value_cell is None
                for formula_row, value_row in zip(formula_rows, rows, strict=True)
                for formula_cell, value_cell in zip(formula_row, value_row, strict=True)
            )
            if not needs_calculation:
                return rows
        except Exception as e:
            logger.debug("openpyxl での読み込みに失敗（%s）。win32com にフォールバックします", e)

        from ..windows.handler import ExcelComHandler

        with ExcelComHandler(self._working_path) as com:
            return com.read_rows(sheet_name, min_row)

    def close(self) -> None:
        """ワークブックを閉じる。with 文を使う場合は自動で呼ばれる。"""
        self._wb.close()

        # ── ローカルコピーの後処理（__init__ の対応ブロックと一緒に削除）──
        if self._tmp:
            self._tmp.unlink(missing_ok=True)
        # ──────────────────────────────────────────────────────────────────
