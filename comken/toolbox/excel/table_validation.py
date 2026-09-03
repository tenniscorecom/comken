"""comken/toolbox/excel/table_validation.py — Excel 範囲の安全性判定ヘルパー。

``workbook.py`` の ``Excel.convert_range_to_table`` は、
既存セル範囲を Excel テーブルに変換する前に「範囲の妥当性」を複数の観点で
判定する責務を持っていた（寸法内・見出し行の決定・見出し空欄・結合セル・
空データ行・重複見出し）。判定ロジック自体は ``Excel`` インスタンス状態に
依存しないため、本モジュールに**モジュールレベルの通常関数**として切り出した。

切り出しの経緯:

- インスタンスメソッド (``self._xxx`` / ``Excel._yyy``) だった頃は、``Excel``
  クラスが「ブックを開く／読む／書く／閉じる」に加えて「テーブル化の安全
  判定」も抱えていた。 ``Excel.convert_range_to_table`` 周辺だけが肥大化し、
  テスト時に ``Excel`` を ``with`` で開かずに ``Worksheet`` だけで呼び出す
  経路を試しにくかった。
- ``static`` 化されていた補助関数も含め、判定ロジックを**副作用なしの
  トップレベル関数**に揃えることで、 ``convert_range_to_table`` は判定を
  ``validate_range_for_table(worksheet, range_spec, header_row)`` 一本に
  集約できる形になった。テストは ``Worksheet`` を直接用意して判定だけを
  試せる。

公開 API は ``validate_range_for_table`` のみ。あとは内部用として
``_`` プレフィックスで隠す。``workbook.py`` からは ``from
comken.toolbox.excel.table_validation import validate_range_for_table``
だけで参照する。
"""

import builtins
from typing import Any

from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from comken.exceptions import (
    DuplicateHeaderCellError,
    EmptyHeaderCellError,
    InvalidTableInputError,
)


def validate_range_for_table(
    worksheet: Worksheet,
    range_spec: str,
    explicit_header_row: int | None,
) -> tuple[int, int, int, int, int, list[Any]]:
    """``convert_range_to_table`` の安全性判定を集約する。

    戻り値は ``(min_col, min_row, max_col, max_row, header_row, header_cells)``。
    各判定に違反したら対応する既存例外で止める。
    """
    boundaries = range_boundaries(range_spec)
    # ``range_boundaries`` は型ヒント上 ``int | None`` を返すが、利用者が ``range_spec``
    # を ``range_boundaries`` を通せる形で渡しているため通常 ``None`` にはならない。
    # ``None`` が現れた場合のフォールバックは 0 とする（``cell(...)`` 呼び出しが
    # 「先頭セル」として動くため、見える不整合は出ない）。
    range_min_col = 0 if boundaries[0] is None else boundaries[0]
    range_min_row = 0 if boundaries[1] is None else boundaries[1]
    range_max_col = 0 if boundaries[2] is None else boundaries[2]
    range_max_row = 0 if boundaries[3] is None else boundaries[3]
    _check_range_inside_dimensions(
        worksheet, range_spec, range_min_col, range_min_row, range_max_col, range_max_row
    )
    header_row = _resolve_header_row(
        worksheet, explicit_header_row, range_min_row, range_min_col, range_max_col
    )
    # 引数 ``range`` が組み込み ``range`` を覆い隠すため ``builtins.range`` を直接呼ぶ
    header_cells = [
        worksheet.cell(row=header_row, column=column).value
        for column in builtins.range(range_min_col, range_max_col + 1)
    ]
    _check_header_not_empty(header_cells)
    _check_no_merged_cells_in_range(
        worksheet,
        range_spec,
        range_min_row,
        range_max_row,
        range_min_col,
        range_max_col,
        header_row,
    )
    _check_no_blank_data_rows(worksheet, header_row, range_max_row, range_min_col, range_max_col)
    _check_no_duplicate_headers(header_cells)
    return range_min_col, range_min_row, range_max_col, range_max_row, header_row, header_cells


def _check_range_inside_dimensions(
    worksheet: Worksheet,
    range_spec: str,
    range_min_col: int,
    range_min_row: int,
    range_max_col: int,
    range_max_row: int,
) -> None:
    """``range`` がシートの使用範囲 (``dimensions``) 内に収まっているか検証する。"""
    dim = worksheet.dimensions
    # ``dimensions`` が空（新規シート等）のときは ``range`` をそのまま許容する。
    if not dim or dim == "A1:A1":
        return
    # ``range_boundaries`` は型ヒント上 ``int | None`` を返すが、既存の
    # ``_table_boundaries`` と同じく「dimensions から取った値は通常 None にならない」
    # 前提で 0 フォールバックを ``int`` に揃える。
    boundaries = range_boundaries(dim)
    dim_min_col = 0 if boundaries[0] is None else boundaries[0]
    dim_min_row = 0 if boundaries[1] is None else boundaries[1]
    dim_max_col = 0 if boundaries[2] is None else boundaries[2]
    dim_max_row = 0 if boundaries[3] is None else boundaries[3]
    if (
        range_min_col < dim_min_col
        or range_min_row < dim_min_row
        or range_max_col > dim_max_col
        or range_max_row > dim_max_row
    ):
        raise InvalidTableInputError(
            f"指定した範囲 {range_spec} がシート「{worksheet.title}」の使用範囲 "
            f"({dim}) を超えています。"
        )


def _resolve_header_row(
    worksheet: Worksheet,
    explicit_header_row: int | None,
    range_min_row: int,
    range_min_col: int,
    range_max_col: int,
) -> int:
    """見出し行の番号を決定する（明示 > A2 ルール > 先頭行）。"""
    if explicit_header_row is not None:
        return explicit_header_row
    # A2 ルール: 先頭行に結合セルがあるなら次行を見出し行とする（タイトル行の許容）。
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row == range_min_row
            and merged_range.min_col >= range_min_col
            and merged_range.max_col <= range_max_col
        ):
            return range_min_row + 1
    return range_min_row


def _check_header_not_empty(header_cells: list[Any]) -> None:
    """見出し行のセルが全て埋まっているか検証する。"""
    empty_columns = [
        column for column, value in enumerate(header_cells, start=1) if value is None or value == ""
    ]
    if empty_columns:
        raise EmptyHeaderCellError(empty_columns)


def _check_no_merged_cells_in_range(
    worksheet: Worksheet,
    range_spec: str,
    range_min_row: int,
    range_max_row: int,
    range_min_col: int,
    range_max_col: int,
    header_row: int,
) -> None:
    """``range`` 内に結合セルが無いことを検証する。

    A2 が発火した「見出し行より前の行」にある結合はタイトル行として許容する。
    """
    merged_in_range: list[str] = []
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row < range_min_row
            or merged_range.max_row > range_max_row
            or merged_range.min_col < range_min_col
            or merged_range.max_col > range_max_col
        ):
            # 完全に範囲外
            continue
        # ``header_row`` 以下の行に完全に収まる結合は「タイトル領域」として除外する。
        if merged_range.max_row < header_row:
            continue
        merged_in_range.append(str(merged_range))
    if merged_in_range:
        raise InvalidTableInputError(
            f"指定した範囲 {range_spec} に結合セルがあります: "
            f"{', '.join(merged_in_range)}。"
            "convert_range_to_table は結合セルを含まない範囲をテーブル化します。"
            "タイトル行として許容されるのは見出し行より上にある結合だけです。"
        )


def _check_no_blank_data_rows(
    worksheet: Worksheet,
    header_row: int,
    range_max_row: int,
    range_min_col: int,
    range_max_col: int,
) -> None:
    """データ行（見出し行より下、range 内）に「全セル空」の行が無いことを検証する。"""
    blank_data_rows: list[int] = []
    for row_index in builtins.range(header_row + 1, range_max_row + 1):
        row_values = [
            worksheet.cell(row=row_index, column=column).value
            for column in builtins.range(range_min_col, range_max_col + 1)
        ]
        if all(value is None or value == "" for value in row_values):
            blank_data_rows.append(row_index)
    if blank_data_rows:
        raise InvalidTableInputError(
            f"指定した範囲のデータ行に空行があります: 行 {blank_data_rows}。"
            "データは連続している必要があります。"
        )


def _check_no_duplicate_headers(header_cells: list[Any]) -> None:
    """見出し行の値に重複が無いことを検証する。"""
    header_strings = ["" if value is None else str(value) for value in header_cells]
    duplicates: list[str] = []
    seen: set[str] = set()
    for header in header_strings:
        if header in seen:
            duplicates.append(header)
        else:
            seen.add(header)
    if duplicates:
        raise DuplicateHeaderCellError(duplicates)
