"""利用者向けExcel入口のエンジン切替と保存契約を確認する。"""

from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook, load_workbook

from comken import dry_run
from comken.exceptions import ExcelMacroPreservationError, ExcelSaveValidationError
from comken.toolbox.excel import Excel


def _book(path, value: str = "変更前") -> None:
    workbook = Workbook()
    workbook.active["A1"] = value
    workbook.save(path)


def _value(path) -> str:
    workbook = load_workbook(path)
    try:
        return workbook.active["A1"].value
    finally:
        workbook.close()


class TestExcelAutomaticSave:
    def test_normal_exit_saves_dirty_book(self, tmp_path) -> None:
        path = tmp_path / "normal.xlsx"
        _book(path)

        with Excel(path) as excel:
            excel.sheet("Sheet").write_value("A1", "変更後")

        assert _value(path) == "変更後"

    def test_exception_exit_does_not_save(self, tmp_path) -> None:
        path = tmp_path / "exception.xlsx"
        _book(path)

        with pytest.raises(RuntimeError), Excel(path) as excel:
            excel.sheet("Sheet").write_value("A1", "保存しない")
            raise RuntimeError("処理失敗")

        assert _value(path) == "変更前"

    def test_read_only_and_dry_run_do_not_save(self, tmp_path) -> None:
        read_only_path = tmp_path / "read_only.xlsx"
        dry_run_path = tmp_path / "dry_run.xlsx"
        _book(read_only_path)
        _book(dry_run_path)

        with Excel(read_only_path, read_only=True):
            pass
        with dry_run(), Excel(dry_run_path) as excel:
            excel.sheet("Sheet").write_value("A1", "保存しない")

        assert _value(read_only_path) == "変更前"
        assert _value(dry_run_path) == "変更前"

    def test_local_copy_saves_to_original_and_removes_copy(self, tmp_path) -> None:
        path = tmp_path / "local.xlsx"
        _book(path)
        with Excel(path, local_copy=True) as excel:
            copy_path = excel._local_copy_path
            excel.sheet("Sheet").write_value("A1", "変更後")

        assert _value(path) == "変更後"
        assert copy_path is not None and not copy_path.exists()

    def test_explicit_save_resyncs_local_copy_before_later_com_use(self, tmp_path) -> None:
        path = tmp_path / "local-then-com.xlsx"
        _book(path)
        with Excel(path, local_copy=True) as excel:
            copy_path = excel._local_copy_path
            excel.sheet("Sheet").write_value("A1", "明示保存後")
            excel.save()
            assert excel._working_copy_is_stale

            excel._sync_working_file()
            assert not excel._working_copy_is_stale
            assert copy_path is not None and _value(copy_path) == "明示保存後"

    def test_invalid_temporary_book_keeps_original(self, tmp_path) -> None:
        path = tmp_path / "invalid.xlsx"
        _book(path)
        original = path.read_bytes()
        with Excel(path) as excel:
            excel.sheet("Sheet").write_value("A1", "変更後")

            with (
                patch.object(
                    excel._workbook, "save", side_effect=lambda target: target.write_bytes(b"bad")
                ),
                pytest.raises(ExcelSaveValidationError),
            ):
                excel.save()
            excel.close(save=False)

        assert path.read_bytes() == original

    def test_unc_detection_can_be_explicitly_overridden(self, tmp_path) -> None:
        assert Excel._is_unc_path(r"\\server\share\book.xlsx")
        assert not Excel._is_unc_path(r"C:\work\book.xlsx")

        path = tmp_path / "unc-like.xlsx"
        _book(path)

        # UNCパス（モック）と判定されたときは、作業コピーが作られる。
        with (
            patch.object(Excel, "_is_unc_path", return_value=True),
            Excel(path) as excel,
        ):
            assert excel._local_copy_path is not None
            assert excel._local_copy_path.exists()

        # local_copy=False を明示すれば、作業コピーを使わない。
        with (
            patch.object(Excel, "_is_unc_path", return_value=True),
            Excel(path, local_copy=False) as excel,
        ):
            assert excel._local_copy_path is None

        # 通常のパスで UNC判定モックが無効なら、コピーは走らない。
        with Excel(path) as excel:
            assert excel._local_copy_path is None

    def test_vba_change_in_temporary_book_keeps_original(self, tmp_path) -> None:
        path = tmp_path / "macro-check.xlsx"
        _book(path)
        original = path.read_bytes()
        with Excel(path) as excel:
            excel.sheet("Sheet").write_value("A1", "変更後")

            with (
                patch.object(excel, "_vba_digest", side_effect=[b"before", b"after"]),
                pytest.raises(ExcelMacroPreservationError),
            ):
                excel.save()
            excel.close(save=False)

        assert path.read_bytes() == original


class TestExcelComPromotion:
    def test_uncomputed_formula_uses_bulk_com_read(self, tmp_path) -> None:
        path = tmp_path / "formula.xlsx"
        workbook = Workbook()
        workbook.active.append(["値", "結果"])
        workbook.active.append([10, "=A2*2"])
        workbook.save(path)
        com = MagicMock()
        com.__enter__.return_value.read_rows.return_value = [(10, 20)]

        with (
            patch("comken.toolbox.windows.handler.ExcelCOMHandler", return_value=com) as handler,
            Excel(path) as excel,
        ):
            rows = excel.read_computed_rows("Sheet")

        assert rows == [(10, 20)]
        handler.assert_called_once_with(path, local_copy_threshold_mb=0)
        com.__enter__.return_value.read_rows.assert_called_once_with("Sheet", 2)

    def test_formula_free_book_stays_on_openpyxl(self, tmp_path) -> None:
        path = tmp_path / "values.xlsx"
        workbook = Workbook()
        workbook.active.append(["値"])
        workbook.active.append([10])
        workbook.save(path)

        with (
            patch("comken.toolbox.windows.handler.ExcelCOMHandler") as handler,
            Excel(path) as excel,
        ):
            assert excel.read_computed_rows("Sheet") == [(10,)]

        handler.assert_not_called()

    def test_cached_formula_does_not_start_com(self, tmp_path) -> None:
        path = tmp_path / "cached_formula.xlsx"
        _book(path)

        with (
            patch("comken.toolbox.windows.handler.ExcelCOMHandler") as handler,
            Excel(path) as excel,
            patch.object(excel, "_cached_rows", return_value=([(20,)], False)),
        ):
            assert excel.read_computed_rows("Sheet") == [(20,)]

        handler.assert_not_called()

    def test_run_macro_promotes_and_marks_book_for_save(self, tmp_path) -> None:
        path = tmp_path / "macro.xlsm"
        _book(path)
        com = MagicMock()

        with (
            patch("comken.toolbox.windows.handler.ExcelCOMHandler", return_value=com) as handler,
            patch.object(Excel, "_reload_workbook"),
            Excel(path) as excel,
        ):
            excel.run_macro("Module1.Update")
            working_path = excel._working_path
            excel._is_dirty = False  # モックCOMが実ファイルを作らないため後段の保存だけ抑止

        assert working_path != path
        handler.assert_called_once_with(working_path, local_copy_threshold_mb=0)
        com.__enter__.return_value.run_macro.assert_called_once_with("Module1.Update")
        com.__enter__.return_value.save.assert_called_once_with()

    def test_successful_macro_is_saved_to_original_on_normal_exit(self, tmp_path) -> None:
        path = tmp_path / "macro_success.xlsm"
        _book(path)

        class FakeExcelComHandler:
            """作業コピーだけを書き換える、保存契約確認用の偽COM。"""

            def __init__(self, working_path, **_kwargs) -> None:
                self.working_path = working_path

            def __enter__(self):
                return self

            def __exit__(self, *_args) -> None:
                return None

            def run_macro(self, _macro_name: str) -> None:
                workbook = load_workbook(self.working_path, keep_vba=True)
                try:
                    workbook.active["A1"] = "マクロ実行後"
                    workbook.save(self.working_path)
                finally:
                    workbook.close()

            def save(self) -> None:
                return None

        with (
            patch("comken.toolbox.windows.handler.ExcelCOMHandler", FakeExcelComHandler),
            Excel(path) as excel,
        ):
            excel.run_macro("Module1.Update")

        assert _value(path) == "マクロ実行後"

    def test_macro_failure_does_not_change_original(self, tmp_path) -> None:
        path = tmp_path / "macro_failure.xlsm"
        _book(path)
        original = path.read_bytes()
        com = MagicMock()
        com.__enter__.return_value.run_macro.side_effect = RuntimeError("macro failed")

        with (
            patch("comken.toolbox.windows.handler.ExcelCOMHandler", return_value=com),
            pytest.raises(RuntimeError, match="macro failed"),
            Excel(path) as excel,
        ):
            excel.run_macro("Module1.Update")

        assert path.read_bytes() == original

    def test_macro_then_outer_exception_does_not_change_original(self, tmp_path) -> None:
        path = tmp_path / "macro_outer_failure.xlsm"
        _book(path)
        original = path.read_bytes()
        com = MagicMock()

        with (
            patch("comken.toolbox.windows.handler.ExcelCOMHandler", return_value=com),
            patch.object(Excel, "_reload_workbook"),
            pytest.raises(RuntimeError, match="later failure"),
            Excel(path) as excel,
        ):
            excel.run_macro("Module1.Update")
            raise RuntimeError("later failure")

        assert path.read_bytes() == original

    def test_new_book_macro_uses_temporary_working_file(self, tmp_path) -> None:
        path = tmp_path / "new.xlsm"
        com = MagicMock()

        with (
            patch("comken.toolbox.windows.handler.ExcelCOMHandler", return_value=com) as handler,
            patch.object(Excel, "_reload_workbook"),
            Excel(path) as excel,
        ):
            excel.run_macro("Module1.Update")
            working_path = excel._working_path
            excel._is_dirty = False

        assert working_path != path
        assert working_path.suffix == ".xlsm"
        handler.assert_called_once_with(working_path, local_copy_threshold_mb=0)
        assert not working_path.exists()


class TestSheetFormat:
    def test_format_sets_only_specified_attributes_and_keeps_others(self, tmp_path) -> None:
        """``format("A1", bold=True)`` だけでも、他のフォント属性は保持される。"""
        path = tmp_path / "format-preserve.xlsx"
        with Excel(path) as excel:
            sheet = excel.create_sheet("集計")
            sheet.write_value("A1", "x")
            # 先に italic とサイズを設定しておき、bold だけの format() で
            # それらが消えないことを確認する
            sheet.format("A1", italic=True, size=20, name="Arial")
            sheet.format("A1", bold=True)
        with Excel(path, read_only=True) as excel:
            font = excel.sheet("集計")._worksheet["A1"].font
            assert font.bold is True
            assert font.italic is True
            assert font.size == 20
            assert font.name == "Arial"

    def test_format_accepts_size_and_number_format_together(self, tmp_path) -> None:
        """``bold / size / number_format`` を同時に渡せる。"""
        path = tmp_path / "format-multi.xlsx"
        with Excel(path) as excel:
            sheet = excel.create_sheet("集計")
            sheet.write_value("B1", 12.5)
            sheet.format("B1", bold=True, size=14, number_format="0.00")
        with Excel(path, read_only=True) as excel:
            cell = excel.sheet("集計")._worksheet["B1"]
            assert cell.font.bold is True
            assert cell.font.size == 14
            assert cell.number_format == "0.00"

    def test_format_rejects_unknown_keyword_at_call_time(self) -> None:
        """未知のキーワードは Python の呼び出し時点で TypeError。"""
        import inspect

        from comken.toolbox.excel.sheet import Sheet

        sig = inspect.signature(Sheet.format)
        # format() は名前付きキーワードしか受け付けない（**kwargs は持たない）。
        # weight を bind しようとすると TypeError になる
        with pytest.raises(TypeError):
            sig.bind("A1", weight=2)
