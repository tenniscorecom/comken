"""comken/toolbox/excel/sheet.py — Excel シートを操作する。"""

import re
from copy import copy
from typing import TYPE_CHECKING, Any, Literal

from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.table import Table as OpenPyXLTable
from openpyxl.worksheet.table import TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from comken.core.table.model import Table
from comken.exceptions import (
    DataSheetAccessError,
    InvalidTableInputError,
    InvalidTableNameError,
    InvalidTableOperationError,
    TableAlreadyExistsError,
    TableNotFoundError,
)
from comken.toolbox.excel.table import ExcelTable

if TYPE_CHECKING:
    from comken.core.table.model import Table
    from comken.toolbox.excel.workbook import Excel

# openpyxl の Side が受け付ける境界線のスタイル。
# 値の一覧は openpyxl.styles.Side.style の NoneSet に従う。
BorderStyle = Literal[
    "dashDot",
    "dashDotDot",
    "dashed",
    "dotted",
    "double",
    "hair",
    "medium",
    "mediumDashDot",
    "mediumDashDotDot",
    "mediumDashed",
    "slantDashDot",
    "thick",
    "thin",
]

# Excel テーブル名でセル参照と紛らわしい形（A1, R1C1 など）に該当するパターン。
# 数字始まりは別条件ではじくので、ここでは数字を含まない / 含み得る両方を許容する。
_CELL_REFERENCE_PATTERN = re.compile(r"^[A-Z]+[0-9]+$|^R[0-9]+C[0-9]+$", re.IGNORECASE)
# Excel がテーブル名に許さない特殊文字。バッククォート・鉤括弧・パス区切りなどを含む。
_FORBIDDEN_TABLE_CHARACTERS = frozenset("[]/\\:*?\"<>|'`#%@$&+={}~")


class Sheet:
    """Excel シートのデータ領域または表示領域を操作する。"""

    PY_TABLE_PREFIX = "PY_T_"

    def __init__(self, excel: "Excel", worksheet: Worksheet) -> None:
        self._excel = excel
        self._worksheet = worksheet

    @property
    def is_data_sheet(self) -> bool:
        """プレフィックス付きのデータシートか返す。"""
        return self._excel._is_data_sheet_name(self._worksheet.title)

    def table(self, name: str | None = None) -> ExcelTable:
        """データシート全体を扱うテーブルを返す。"""
        if not self.is_data_sheet:
            raise DataSheetAccessError(self._worksheet.title, "table")
        table_names = list(self._worksheet.tables)
        if name is not None:
            name = self._with_table_prefix(name)
            if name not in table_names:
                raise TableNotFoundError(name, table_names)
        if name is None and len(table_names) > 1:
            raise InvalidTableOperationError(
                "1シートに複数テーブルがあります。table(name)で指定してください。"
            )
        return ExcelTable(self._excel, self._worksheet, name)

    def create_table(self, name: str, table: "Table", start_cell: str = "A1") -> ExcelTable:
        """Python管理用の実テーブルを新規作成する。

        ``start_cell`` は見出しの左上セルです。作成直後の Table はメモリ上の
        現在値で、Excel ファイルへの保存は Excel の save/with 契約で後から行います。
        """
        if not self.is_data_sheet:
            raise DataSheetAccessError(self._worksheet.title, "create_table")
        full_name = self._with_table_prefix(name)
        self._validate_table_name(name)
        if any(full_name in worksheet.tables for worksheet in self._excel._workbook.worksheets):
            raise TableAlreadyExistsError(full_name)
        if not isinstance(table, Table):
            raise InvalidTableInputError("create_table には Table を指定してください。")
        if not table.columns:
            raise InvalidTableInputError("列がないTableはExcelテーブルにできません。")
        try:
            start_column, start_row = coordinate_from_string(start_cell)
            start_column_number = column_index_from_string(start_column)
        except (TypeError, ValueError):
            raise InvalidTableInputError(f"start_cell が不正です: {start_cell!r}") from None
        for column, header in enumerate(table.columns, start_column_number):
            self._worksheet.cell(start_row, column, header)
        for row_number, row in enumerate(table.read(), start_row + 1):
            for column, header in enumerate(table.columns, 1):
                self._worksheet.cell(
                    row_number,
                    start_column_number + column - 1,
                    row.get(header, ""),
                )
        end_row = max(start_row + len(table.read()), start_row + 1)
        end_column = start_column_number + len(table.columns) - 1
        # Excel の実テーブル範囲は、無関係なセルではなく見出し・データから計算します。
        ref = f"{start_cell.upper()}:{get_column_letter(end_column)}{end_row}"
        excel_table = OpenPyXLTable(displayName=full_name, ref=ref)
        excel_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        self._worksheet.add_table(excel_table)
        self._excel._mark_dirty()
        return ExcelTable(self._excel, self._worksheet, full_name)

    @staticmethod
    def _with_table_prefix(name: str) -> str:
        """短いテーブル名に、Python管理用の ``PY_T_`` を補う。"""
        return name if name.startswith(Sheet.PY_TABLE_PREFIX) else f"{Sheet.PY_TABLE_PREFIX}{name}"

    @staticmethod
    def _validate_table_name(name: str) -> None:
        """Excel のテーブル名の制約に合っているか検証する。

        利用者が ``create_table`` に渡した名前を直接判定する。``PY_T_`` プレフィックスは
        ``_with_table_prefix`` が補うため、利用者が ``A1`` と書いた時点でセル参照と
        紛らわしい。Excel が拒否する名前は openpyxl の例外からでは原因が分かりにくいため、
        ここで先に弾く。
        """
        if not name:
            raise InvalidTableNameError(name)
        # Excel はテーブル名に空白・制御文字を許さない
        if any(ch.isspace() for ch in name):
            raise InvalidTableNameError(name)
        # 先頭が数字だとセル参照と紛らわしい
        if name[0].isdigit():
            raise InvalidTableNameError(name)
        # Excel がセル参照と解釈し得る形（A1, R1C1 など）を拒否する
        if _CELL_REFERENCE_PATTERN.fullmatch(name):
            raise InvalidTableNameError(name)
        # Excel が許さない特殊文字を拒否する
        for character in name:
            if character in _FORBIDDEN_TABLE_CHARACTERS:
                raise InvalidTableNameError(name)

    def write_value(self, cell: str, value: Any) -> None:
        """セルへ値を書き込む。"""
        self._ensure_display_sheet("write_value")
        self._worksheet[cell] = value
        self._excel._mark_dirty()

    def read_value(self, cell: str, *, force_com: bool = False) -> Any:
        """セルの値を読む。数式は計算結果を返す。

        ブックは ``data_only`` 以外の状態で開くため、メモリ上の ``cell.value`` は
        数式セルでは ``"=SUM(A1:A3)"`` という文字列になる。``read_value`` は
        数式セルでは保存済み計算値（無ければ COM で再計算）を返す。
        ``force_com=True`` でキャッシュを無視して Excel 実機で強制再計算させる。
        """
        self._ensure_display_sheet("read_value")
        raw = self._worksheet[cell].value
        is_formula = isinstance(raw, str) and raw.startswith("=")
        if not is_formula and not force_com:
            return "" if raw is None else raw
        column, row = coordinate_from_string(cell)
        column_index = column_index_from_string(column)
        self._excel._ensure_open()
        if force_com:
            rows = self._excel._read_range_with_com(
                self._worksheet.title, column_index, row, column_index, row
            )
            return rows[0][0] if rows else ("" if raw is None else raw)
        cached_rows, needs_com = self._excel._cached_range(
            self._worksheet.title, column_index, row, column_index, row
        )
        if not needs_com and cached_rows:
            return cached_rows[0][0]
        rows = self._excel._read_range_with_com(
            self._worksheet.title, column_index, row, column_index, row
        )
        return rows[0][0] if rows else ("" if raw is None else raw)

    def read_formula(self, cell: str) -> str:
        """セルの数式を読む。数式でなければ空文字を返す。

        ``read_value`` は計算結果を返すため、もう数式の判定には使えない。
        ワークシートの生の値を直接見る。
        """
        self._ensure_display_sheet("read_formula")
        value = self._worksheet[cell].value
        return value if isinstance(value, str) and value.startswith("=") else ""

    def write_range(self, cell_range: str, values: list[list[Any]]) -> None:
        """指定範囲へ二次元の値を書き込む。"""
        self._ensure_display_sheet("write_range")
        cells = self._worksheet[cell_range]
        if len(cells) != len(values) or any(
            len(row) != len(value_row) for row, value_row in zip(cells, values, strict=False)
        ):
            raise ValueError("指定範囲と values の行数・列数が一致しません。")
        for cell_row, value_row in zip(cells, values, strict=True):
            for cell, value in zip(cell_row, value_row, strict=True):
                cell.value = value
        self._excel._mark_dirty()

    def read_range(self, cell_range: str, *, force_com: bool = False) -> list[dict[str, Any]]:
        """指定範囲の先頭行を見出しとして辞書のリストで読む。

        数式セルがある範囲では保存済み計算値、無ければ COM で再計算した値を返す。
        ``force_com=True`` でキャッシュを無視して Excel 実機で強制再計算させる。
        """
        self._ensure_display_sheet("read_range")
        cells = self._worksheet[cell_range]
        # MultiCellRange は iter_rows 形式で展開して扱う。
        if hasattr(cells, "min_row"):
            min_row, min_col = cells.min_row, cells.min_col
            max_row, max_col = cells.max_row, cells.max_col
            row_iter = self._worksheet.iter_rows(
                min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
            )
        else:
            row_iter = iter(cells)
            min_row = min_col = max_row = max_col = 0
        has_formula = any(
            isinstance(cell.value, str) and cell.value.startswith("=")
            for row in row_iter
            for cell in row
        )
        if has_formula or force_com:
            if force_com:
                rows = self._excel._read_range_with_com(
                    self._worksheet.title, min_col, min_row, max_col, max_row
                )
            else:
                cached_rows, needs_com = self._excel._cached_range(
                    self._worksheet.title, min_col, min_row, max_col, max_row
                )
                if needs_com:
                    rows = self._excel._read_range_with_com(
                        self._worksheet.title, min_col, min_row, max_col, max_row
                    )
                else:
                    rows = cached_rows
        else:
            rows = [
                tuple("" if cell.value is None else cell.value for cell in row) for row in cells
            ]
        if not rows:
            return []
        headers = [str(value) for value in rows[0]]
        return [dict(zip(headers, row, strict=True)) for row in rows[1:]]

    def get_used_range(self) -> tuple[str, str]:
        """使用範囲の左上と右下のセル参照を返す。"""
        self._ensure_display_sheet("get_used_range")
        return "A1", f"{get_column_letter(self._worksheet.max_column)}{self._worksheet.max_row}"

    def set_row_height(self, row: int, height: float) -> None:
        """行の高さを設定する。"""
        self._ensure_display_sheet("set_row_height")
        self._worksheet.row_dimensions[row].height = height
        self._excel._mark_dirty()

    def set_column_width(self, col: str, width: float) -> None:
        """列の幅を設定する。"""
        self._ensure_display_sheet("set_column_width")
        self._worksheet.column_dimensions[col].width = width
        self._excel._mark_dirty()

    def hide_row(self, row: int) -> None:
        """指定した行を非表示にする。

        行の表示設定はデータ表の内容ではなく画面レイアウトなので、データシート
        ではなく表示シートに限定している。
        """
        self._set_row_hidden(row, True)

    def show_row(self, row: int) -> None:
        """指定した行の非表示を解除する。"""
        self._set_row_hidden(row, False)

    def hide_column(self, col: str) -> None:
        """指定した列を非表示にする。"""
        self._set_column_hidden(col, True)

    def show_column(self, col: str) -> None:
        """指定した列の非表示を解除する。"""
        self._set_column_hidden(col, False)

    def insert_row(self, row: int) -> None:
        """指定位置に表示用の行を挿入する。"""
        self._ensure_display_sheet("insert_row")
        self._worksheet.insert_rows(row)
        self._excel._mark_dirty()

    def delete_row(self, row: int) -> None:
        """指定位置の表示用の行を削除する。"""
        self._ensure_display_sheet("delete_row")
        self._worksheet.delete_rows(row)
        self._excel._mark_dirty()

    def insert_column(self, col: str) -> None:
        """指定位置に表示用の列を挿入する。"""
        self._ensure_display_sheet("insert_column")
        self._worksheet.insert_cols(column_index_from_string(col))
        self._excel._mark_dirty()

    def delete_column(self, col: str) -> None:
        """指定位置の表示用の列を削除する。"""
        self._ensure_display_sheet("delete_column")
        self._worksheet.delete_cols(column_index_from_string(col))
        self._excel._mark_dirty()

    def format(
        self,
        cell: str,
        *,
        bold: bool | None = None,
        italic: bool | None = None,
        size: int | None = None,
        name: str | None = None,
        color: str | None = None,
        number_format: str | None = None,
    ) -> None:
        """セルのフォントと表示形式を設定する。

        渡した引数だけ反映し、``None`` の項目は既存の値を変えない。**指定しない
        項目がリセットされることはない**ので、``bold`` だけ書き換えるつもりで
        ``size`` が初期値に戻る、といった事故が起きない。

        Args:
            cell: 対象のセル参照 (例: ``"A1"``)。
            bold: ``True`` で太字、``False`` で解除、``None`` で変更しない。
            italic: イタリック。``True`` / ``False`` / ``None``。
            size: フォントサイズ。``None`` のとき変更しない。
            name: フォント名。``None`` のとき変更しない。
            color: 16進数 6 桁の色 (``#`` 付きでも可)。``None`` のとき変更しない。
            number_format: セルの表示形式 (例: ``"0.00"``)。``None`` のとき変更しない。

        Raises:
            TypeError: セル参照が不正な場合。
        """
        self._ensure_display_sheet("format")
        target = self._worksheet[cell]
        font = copy(target.font)
        if bold is not None:
            font.bold = bold
        if italic is not None:
            font.italic = italic
        if size is not None:
            font.size = size
        if name is not None:
            font.name = name
        if color is not None:
            # ``set_border()`` と同じく ``#`` 付きを許容する
            font.color = color.removeprefix("#")
        if any(value is not None for value in (bold, italic, size, name, color)):
            # 既存フォントの属性（太字・色など）を保ったまま反映するため、
            # 1つでもフォント指定があれば必ず書き戻す
            target.font = font
        if number_format is not None:
            target.number_format = str(number_format)
        self._excel._mark_dirty()

    def set_background(self, cell: str, color: str) -> None:
        """セルの背景色を設定する。"""
        self._ensure_display_sheet("set_background")
        self._worksheet[cell].fill = PatternFill("solid", fgColor=color.removeprefix("#"))
        self._excel._mark_dirty()

    def set_border(
        self,
        cell: str,
        *,
        style: BorderStyle = "thin",
        color: str = "000000",
    ) -> None:
        """セルの四辺に同じ境界線を設定する。

        よく使う ``style``: ``"thin"`` / ``"medium"`` / ``"thick"`` /
        ``"dashed"`` / ``"double"``。全種類は ``BorderStyle`` 型を参照。

        Args:
            cell: 対象のセル参照 (例: ``"A1"``)。
            style: 線の種類。 ``BorderStyle`` で定義したいずれかの値。
            color: 16進数 6 桁の色 (``#`` 付きでも可)。既定は ``"000000"``。

        Raises:
            ValueError: ``style`` が ``BorderStyle`` のいずれにも該当しない
                (openpyxl の検証による)。
        """
        self._ensure_display_sheet("set_border")
        side = Side(style=style, color=color.removeprefix("#"))
        self._worksheet[cell].border = Border(left=side, right=side, top=side, bottom=side)
        self._excel._mark_dirty()

    def merge_cells(self, cell_range: str) -> None:
        """指定範囲のセルを結合する。"""
        self._ensure_display_sheet("merge_cells")
        self._worksheet.merge_cells(cell_range)
        self._excel._mark_dirty()

    def unmerge_cells(self, cell_range: str) -> None:
        """指定範囲のセル結合を解除する。"""
        self._ensure_display_sheet("unmerge_cells")
        self._worksheet.unmerge_cells(cell_range)
        self._excel._mark_dirty()

    def freeze_panes(self, cell: str) -> None:
        """指定セルより上・左の領域を固定表示する。"""
        self._ensure_display_sheet("freeze_panes")
        self._worksheet.freeze_panes = cell
        self._excel._mark_dirty()

    def _ensure_display_sheet(self, operation: str) -> None:
        self._excel._ensure_open()
        if self.is_data_sheet:
            raise DataSheetAccessError(self._worksheet.title, operation)

    def _set_row_hidden(self, row: int, is_hidden: bool) -> None:
        self._ensure_display_sheet("hide/show_row")
        self._worksheet.row_dimensions[row].hidden = is_hidden
        self._excel._mark_dirty()

    def _set_column_hidden(self, col: str, is_hidden: bool) -> None:
        self._ensure_display_sheet("hide/show_column")
        self._worksheet.column_dimensions[col].hidden = is_hidden
        self._excel._mark_dirty()
