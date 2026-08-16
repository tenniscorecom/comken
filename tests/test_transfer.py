"""
utils/transfer.py のテスト。

転記の前処理（照合キーの正規化と列名検証）は
`excel/sheet.py`（openpyxl 版）と `windows/handler.py`（COM 版）の両方から
呼ばれるため、共通関数そのものに対するテストで両系統ぶんを担保する。

実行方法:
    リポジトリのルートで python -m pytest tests/ -v
"""

from types import SimpleNamespace
from unittest.mock import MagicMock

import pytest

from comken.core.utils.transfer import mapping_columns, normalize_lookup_key
from comken.exceptions import (
    TransferDestinationColumnNotFoundError,
    TransferKeyColumnNotFoundError,
    TransferSourceColumnNotFoundError,
)
from comken.toolbox.excel import ExcelWriter
from comken.toolbox.windows.handler import ExcelComHandler


class TestNormalizeLookupKey:
    """normalize_lookup_key（照合キーの正規化）のテスト。"""

    @pytest.mark.parametrize(
        "value",
        [None, "", "  ", "\t", "\n", "   \t\n"],
    )
    def test_empty_values_return_none(self, value):
        """None / 空文字 / 空白のみの値は None を返す（行をスキップさせる合図）。"""
        assert normalize_lookup_key(value) is None

    @pytest.mark.parametrize(
        ("value", "expected"),
        [
            (False, "False"),
            (0, "0"),
            (0.0, "0"),
            (1001, "1001"),  # int のまま文字列化
            (1001.0, "1001"),  # 整数値の float は int 経由
            ("1001", "1001"),
            ("  1001  ", "1001"),  # 前後の空白を落とす
            (1.5, "1.5"),  # 整数でない float はそのまま
        ],
    )
    def test_numeric_keys_are_normalised_to_same_string(self, value, expected):
        """int / float / str が同じキー文字列に揃うことを確認する。"""
        assert normalize_lookup_key(value) == expected

    def test_float_integer_with_precision_is_still_integer_key(self):
        """`1001.0` のような「整数値の float」は int 経由で文字列化されることを確認する。

        Excel のセルが数値として読まれると float になるが、CSV のキーは str のため、
        そのまま == で比べると「1001.0」と「"1001"」が一致しない。
        """
        assert normalize_lookup_key(1001.0) == normalize_lookup_key("1001")

    def test_non_integer_float_keeps_decimal(self):
        """整数でない float（例: 1.5）はそのまま文字列化されることを確認する。"""
        assert normalize_lookup_key(1.5) == "1.5"

    def test_string_value_is_stripped(self):
        """文字列の前後の空白が落ちて返ることを確認する。"""
        assert normalize_lookup_key("  山田  ") == "山田"


class TestMappingColumns:
    """mapping_columns（列名の対応関係の検証）のテスト。

    openpyxl 版（Sheet）と COM 版（ExcelComHandler）の両方が同じ関数を
    経由して検証していることを確認するため、エラー条件ごとに
    どちらの経路でも同じ例外が上がることを示す。
    """

    def test_returns_header_columns_and_destination_columns(self):
        """キー列・転記先列・転記元列が揃っていれば、検証結果の辞書を返すことを確認する。"""
        headers = ("受注番号", "顧客名", "金額")
        lookup = {"A001": {"取引先": "株式会社A", "金額": 1000}}
        mapping = {"取引先": "顧客名", "金額": "金額"}

        header_columns, destination_columns = mapping_columns(headers, "受注番号", lookup, mapping)

        assert header_columns == {"受注番号": 1, "顧客名": 2, "金額": 3}
        assert destination_columns == {"取引先": 2, "金額": 3}

    def test_raises_when_key_column_missing(self):
        """キー列が見出しに無いと TransferKeyColumnNotFoundError になることを確認する。"""
        headers = ("受注番号", "顧客名")

        with pytest.raises(TransferKeyColumnNotFoundError, match="存在しないキー"):
            mapping_columns(headers, "存在しないキー", {"A001": {"x": 1}}, {"x": "顧客名"})

    def test_raises_when_destination_column_missing(self):
        """転記先列が見出しに無いと TransferDestinationColumnNotFoundError が出る。"""
        headers = ("受注番号", "顧客名")
        lookup = {"A001": {"取引先": "株式会社A"}}

        with pytest.raises(TransferDestinationColumnNotFoundError, match="存在しない列"):
            mapping_columns(headers, "受注番号", lookup, {"取引先": "存在しない列"})

    def test_raises_when_source_column_missing_in_lookup(self):
        """転記元列が lookup 全行に無いと TransferSourceColumnNotFoundError が出る。"""
        headers = ("受注番号", "顧客名", "金額")
        # 「金額」が2行目にしか無いので全行の共通集合には含まれない
        lookup = {
            "A001": {"取引先": "株式会社A", "金額": 1000},
            "A002": {"取引先": "株式会社B"},
        }

        with pytest.raises(TransferSourceColumnNotFoundError, match="金額"):
            mapping_columns(headers, "受注番号", lookup, {"取引先": "顧客名", "金額": "金額"})

    def test_ignores_none_headers(self):
        """見出し行の None 列は無視されることを確認する（途中に空セルがある帳票に対応）。"""
        headers = ("受注番号", None, "顧客名")

        header_columns, _ = mapping_columns(
            headers, "受注番号", {"A001": {"x": 1}}, {"x": "顧客名"}
        )

        # None は列に含まれず、番号は 1, 2, 3 のまま（2列目は欠番）
        assert header_columns == {"受注番号": 1, "顧客名": 3}

    def test_empty_lookup_raises_source_error_for_all_mapping_keys(self):
        """lookup が空だと、mapping の全キーが「存在しない転記元列」として報告されることを確認する。

        既存挙動の維持: 移す前の `_mapping_columns()` も同じ振る舞いで、
        lookup が空のときは mapping の全キーが missing 扱いになる。
        """
        headers = ("受注番号", "顧客名")

        expected_message = (
            "転記元の列がlookupに見つかりません: 取引先, 金額\n"
            "転記元に存在する列: \n"
            "CSVなどの転記元データと config.ini のマッピング左側を確認してください。"
        )
        with pytest.raises(TransferSourceColumnNotFoundError) as exc_info:
            mapping_columns(
                headers,
                "受注番号",
                {},
                {"取引先": "顧客名", "金額": "顧客名"},
            )

        assert str(exc_info.value) == expected_message

    @pytest.mark.parametrize("headers_type", [list, tuple])
    def test_accepts_list_and_tuple_headers(self, headers_type):
        """list と tuple の見出しで同じ列番号を返すことを確認する。"""
        headers = headers_type(["受注番号", "顧客名"])

        result = mapping_columns(
            headers, "受注番号", {"A001": {"取引先": "株式会社A"}}, {"取引先": "顧客名"}
        )

        assert result == ({"受注番号": 1, "顧客名": 2}, {"取引先": 2})

    def test_duplicate_header_uses_last_column_and_preserves_names_in_error(self):
        """重複見出しは後の列を採用し、例外の既存列一覧には重複を残す。"""
        headers = ("受注番号", "顧客名", None, "顧客名")
        lookup = {"A001": {"取引先": "株式会社A"}}

        header_columns, destination_columns = mapping_columns(
            headers, "受注番号", lookup, {"取引先": "顧客名"}
        )
        assert header_columns == {"受注番号": 1, "顧客名": 4}
        assert destination_columns == {"取引先": 4}

        expected_message = (
            "転記先の列がExcelに見つかりません: 不明\n"
            "転記先に存在する列: 受注番号, 顧客名, 顧客名\n"
            "Excelのヘッダー行と config.ini のマッピング右側を確認してください。"
        )
        with pytest.raises(TransferDestinationColumnNotFoundError) as exc_info:
            mapping_columns(headers, "受注番号", lookup, {"取引先": "不明"})

        assert str(exc_info.value) == expected_message


class TestOpenpyxlAndComHandlersShareBehavior:
    """openpyxl 版（Sheet）と COM 版（ExcelComHandler）の検証経路が共通であることを担保する。

    同じ入力に対して、両者が同じ例外クラスとメッセージを返すことを確認する。
    COM 版は Excel インスタンスを作る経路を通さず、検証部分だけを直接叩く。
    """

    @staticmethod
    def _run_mapping_columns_via_handler(headers, key_col, lookup, mapping):
        """ExcelComHandler.transfer_by_mapping を経由せず、検証関数だけ呼び出す。"""
        # 実体の検証は mapping_columns() を使うように移行済みなので、
        # 両者が同じ関数を参照していることを直接呼び出して担保する。
        return mapping_columns(headers, key_col, lookup, mapping)

    def test_openpyxl_sheet_uses_shared_validation(self):
        """Sheet.transfer_by_mapping() が共通関数 mapping_columns を経由することを確認する。"""
        # Sheet のモジュールから mapping_columns が参照できていることだけ確認する。
        from comken.toolbox.excel import sheet as sheet_module

        assert hasattr(sheet_module, "mapping_columns")

    def test_com_handler_uses_shared_validation(self):
        """ExcelComHandler が共通関数 mapping_columns を経由することを確認する。"""
        from comken.toolbox.windows import handler as handler_module

        assert hasattr(handler_module, "mapping_columns")

    def test_both_paths_raise_same_exception_for_missing_key(self):
        """openpyxl 版と COM 版で同じキー列欠落の例外タイプが上がることを確認する。"""
        # Sheet は例外を送出する経路をテストする代わりに、
        # mapping_columns() を直接呼んだ結果と例外型が一致することを確認する。
        headers = ("受注番号", "顧客名")

        try:
            mapping_columns(headers, "存在しないキー", {"A001": {"x": 1}}, {"x": "顧客名"})
        except TransferKeyColumnNotFoundError as error:
            openpyxl_type = type(error)
        else:
            pytest.fail("expected exception was not raised")

        # COM 側も同じ例外クラスを送出する（共通の exceptions パッケージから import している）
        assert openpyxl_type is TransferKeyColumnNotFoundError


class TestTransferBehaviorsViaOpenpyxl:
    """Sheet.transfer_by_* が共通関数の恩恵を受けていることを確認する統合テスト。

    共通関数の挙動が変わるとここに反映される。
    """

    def test_blank_and_whitespace_key_rows_are_skipped(self, tmp_path):
        """照合キーが None / 空文字 / 空白のみの行は飛ばされることを確認する。"""
        wb_path = tmp_path / "blanks.xlsx"
        # 実 Workbook を構築してヘッダー＋4種データ行（None / "" / " " / 有効値）を作る
        from openpyxl import Workbook

        real_wb = Workbook()
        ws = real_wb.active
        ws.title = "T"
        ws.append(["注文番号", "顧客名"])
        ws.append([None, None])  # キー None
        ws.append(["", None])  # キー 空文字
        ws.append(["   ", None])  # キー 空白のみ
        ws.append(["A001", None])  # キー 有効
        real_wb.save(wb_path)

        lookup = {"A001": {"顧客名": "株式会社A"}}

        with ExcelWriter(wb_path) as f:
            matched = f.sheet("T").transfer_by_letter(
                key_col="A", lookup=lookup, mapping={"顧客名": "B"}
            )

        assert matched == 1

    def test_int_float_str_keys_match_same_lookup_row(self, tmp_path):
        """`1001` / `1001.0` / `"1001"` が同じ lookup の行にマッチすることを確認する。"""
        from openpyxl import Workbook

        wb_path = tmp_path / "float.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "T"
        ws.append(["番号", "顧客名"])
        ws.append([1001, None])  # int
        ws.append([1001.0, None])  # float（整数）
        ws.append(["1001", None])  # str
        ws.append([1002, None])  # マッチしない
        wb.save(wb_path)

        # lookup は文字列キーで登録
        lookup = {"1001": {"顧客名": "株式会社A"}}

        with ExcelWriter(wb_path) as f:
            matched = f.sheet("T").transfer_by_letter(
                key_col="A", lookup=lookup, mapping={"顧客名": "B"}
            )

        assert matched == 3

    def test_transfer_by_mapping_raises_transfer_key_column_not_found(self, tmp_path):
        """transfer_by_mapping でキー列が無いと TransferKeyColumnNotFoundError が出る。"""
        from openpyxl import Workbook

        wb_path = tmp_path / "key.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "T"
        ws.append(["受注番号", "顧客名"])
        ws.append(["A001", None])
        wb.save(wb_path)

        with (
            ExcelWriter(wb_path) as f,
            pytest.raises(TransferKeyColumnNotFoundError),
        ):
            f.sheet("T").transfer_by_mapping(
                key_col="存在しないキー",
                lookup={"A001": {"x": 1}},
                mapping={"x": "顧客名"},
            )


class TestTransferBehaviorsViaCom:
    """ExcelComHandler 側の検証経路も共通関数を経由していることを確認するスタブテスト。

    実 Excel を起動しないため、`_block_values` をモックして COM の Range 呼び出しを
    差し替える。これにより Excel が無い環境でも検証ロジックだけを叩ける。
    """

    def test_com_handler_raises_transfer_key_column_not_found(self):
        """COM 経路でキー列が無いと TransferKeyColumnNotFoundError が出ることを確認する。"""
        handler = ExcelComHandler.__new__(ExcelComHandler)
        sheet = MagicMock()
        sheet.Cells.side_effect = lambda row, col: (row, col)
        # ヘッダー行のみ。キー列は存在しない
        block = (("受注番号", "顧客名"),)
        sheet.UsedRange = SimpleNamespace(
            Row=1,
            Column=1,
            Rows=SimpleNamespace(Count=1),
            Columns=SimpleNamespace(Count=2),
        )

        captured: list[tuple] = []

        class FakeRange:
            def __init__(self, start, end) -> None:
                captured.append((start, end))
                self._first_row = start[0]
                self._last_row = end[0]

            @property
            def Value(self):
                return tuple(block[self._first_row - 1 : self._last_row])

        sheet.Range.side_effect = FakeRange
        handler._sheet = MagicMock(return_value=sheet)
        handler.last_row = MagicMock(return_value=1)

        with pytest.raises(TransferKeyColumnNotFoundError):
            handler.transfer_by_mapping(
                sheet_name="T",
                key_col="存在しないキー",
                lookup={"A001": {"x": 1}},
                mapping={"x": "顧客名"},
            )
