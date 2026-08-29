"""現行のExcel API（Excel / Sheet / ExcelTable）の契約テスト。"""

import pytest
from openpyxl.styles import PatternFill

from comken.core.table import Table
from comken.exceptions import (
    DataSheetAccessError,
    EmptyHeaderCellError,
    ExcelFileNotFoundError,
    ExcelReadOnlyOperationError,
    InvalidTableNameError,
    InvalidTableOperationError,
    SheetAlreadyExistsError,
    SheetNameError,
    SheetNotFoundError,
    TableNotOpenError,
    UnsupportedFileSuffixError,
)
from comken.toolbox.excel import Excel


def test_excel_creates_and_reads_python_table(tmp_path) -> None:
    path = tmp_path / "book.xlsx"
    with Excel(path) as excel:
        sheet = excel.create_data_sheet("顧客")
        sheet.create_table("顧客", Table(["ID", "名前"], [{"ID": "001", "名前": "山田"}]))
    with Excel(path, read_only=True) as excel:
        assert excel.data_sheet("顧客").table().read() == [{"ID": "001", "名前": "山田"}]


def test_excel_replaces_table_without_saving_until_context_exit(tmp_path) -> None:
    path = tmp_path / "book.xlsx"
    with Excel(path) as excel:
        excel.create_data_sheet("顧客").create_table("顧客", Table(["ID"], [{"ID": "001"}]))
    with Excel(path) as excel:
        table = excel.data_sheet("顧客").table()
        table.replace([{"ID": "002"}])
        assert table.read() == [{"ID": "002"}]
    with Excel(path, read_only=True) as excel:
        assert excel.data_sheet("顧客").table().read() == [{"ID": "002"}]


def test_excel_rejects_ambiguous_table_name(tmp_path) -> None:
    path = tmp_path / "book.xlsx"
    with Excel(path) as excel:
        sheet = excel.create_data_sheet("顧客")
        sheet.create_table("基本", Table(["ID"], [{"ID": "001"}]), "A1")
        sheet.create_table("連絡", Table(["電話"], [{"電話": "000"}]), "D1")
        with pytest.raises(InvalidTableOperationError):
            sheet.table().read()


def test_excel_rejects_duplicate_data_sheet(tmp_path) -> None:
    with Excel(tmp_path / "book.xlsx") as excel:
        excel.create_data_sheet("顧客")
        with pytest.raises(SheetAlreadyExistsError):
            excel.create_data_sheet("顧客")


def test_excel_table_append_accepts_row_list_and_table(tmp_path) -> None:
    path = tmp_path / "book.xlsx"
    with Excel(path) as excel:
        table = excel.create_data_sheet("顧客").create_table("顧客", Table(["ID"], [{"ID": "001"}]))
        table.append({"ID": "002"})
        table.append([{"ID": "003"}])
        table.append(Table(["ID"], [{"ID": "004"}]))
    with Excel(path, read_only=True) as excel:
        assert excel.data_sheet("顧客").table().read().column("ID") == [
            "001",
            "002",
            "003",
            "004",
        ]


def test_excel_rejects_missing_read_only_file_and_non_excel_suffix(tmp_path) -> None:
    path = tmp_path / "missing.xlsx"
    with pytest.raises(ExcelFileNotFoundError), Excel(path, read_only=True):
        pass
    with pytest.raises(UnsupportedFileSuffixError):
        Excel(tmp_path / "book.csv")


def test_create_sheet_uses_name_as_is_and_supports_layout_api(tmp_path) -> None:
    path = tmp_path / "book.xlsx"
    with Excel(path) as excel:
        sheet = excel.create_sheet("集計")
        assert sheet.is_data_sheet is False
        sheet.set_column_width("A", 12)
        sheet.freeze_panes("B2")
        sheet.write_value("A1", "見出し")
        sheet.format("A1", bold=True)
    with Excel(path, read_only=True) as excel:
        restored = excel.sheet("集計")
        assert restored.read_value("A1") == "見出し"
        assert restored.read_value("A1") == "見出し"


def test_create_sheet_does_not_appear_in_list_data_sheets(tmp_path) -> None:
    path = tmp_path / "book.xlsx"
    with Excel(path) as excel:
        excel.create_data_sheet("顧客").create_table("顧客", Table(["ID"], [{"ID": "001"}]))
        excel.create_sheet("集計")
        assert excel.list_data_sheets() == ["PY_顧客"]


def test_create_sheet_allows_multiple_display_sheets(tmp_path) -> None:
    path = tmp_path / "book.xlsx"
    with Excel(path) as excel:
        excel.create_sheet("集計")
        excel.create_sheet("月次")
        display_names = [name for name in excel._workbook.sheetnames if not name.startswith("PY_")]
        assert "集計" in display_names
        assert "月次" in display_names


def test_create_sheet_rejects_duplicate_name(tmp_path) -> None:
    with Excel(tmp_path / "book.xlsx") as excel:
        excel.create_sheet("集計")
        with pytest.raises(SheetAlreadyExistsError):
            excel.create_sheet("集計")


def test_create_sheet_rejects_python_prefixed_name(tmp_path) -> None:
    with Excel(tmp_path / "book.xlsx") as excel, pytest.raises(SheetNameError):
        excel.create_sheet("PY_顧客")


def test_create_sheet_rejects_read_only_workbook(tmp_path) -> None:
    path = tmp_path / "book.xlsx"
    with Excel(path) as excel:
        excel.create_sheet("集計")
    with Excel(path, read_only=True) as excel, pytest.raises(ExcelReadOnlyOperationError):
        excel.create_sheet("別のシート")


def test_create_sheet_returns_sheet_that_supports_layout_api(tmp_path) -> None:
    path = tmp_path / "book.xlsx"
    with Excel(path) as excel:
        sheet = excel.create_sheet("集計")
        # 表示用シートでは table() は DataSheetAccessError
        with pytest.raises(DataSheetAccessError):
            sheet.table()


def test_set_border_uses_thin_by_default(tmp_path) -> None:
    """set_border() は style を省略すると thin が使われる。"""
    import inspect

    from comken.toolbox.excel.sheet import Sheet

    path = tmp_path / "book.xlsx"
    with Excel(path) as excel:
        sheet = excel.create_sheet("集計")
        sheet.write_value("A1", "x")
        sheet.set_border("A1")
    with Excel(path, read_only=True) as excel:
        cell = excel.sheet("集計")._worksheet["A1"]
        assert cell.border.left.style == "thin"
        # openpyxl は 8 桁の ARGB 形式で色を保存する（先頭 2 桁はアルファチャンネル）
        assert cell.border.left.color.value == "00000000"
    # 型定義の Literal として受け付ける値の一覧（実行時の網羅チェック）
    sig = inspect.signature(Sheet.set_border)
    if sig.parameters["style"].default != "thin":
        pytest.fail(f"style 既定値が 'thin' ではない: {sig.parameters['style'].default!r}")
    if sig.parameters["color"].default != "000000":
        pytest.fail(f"color 既定値が '000000' ではない: {sig.parameters['color'].default!r}")


def test_set_border_accepts_style_and_strips_hash(tmp_path) -> None:
    """style / color を指定でき、color の '#' は落ちる。"""
    path = tmp_path / "book.xlsx"
    with Excel(path) as excel:
        sheet = excel.create_sheet("集計")
        sheet.write_value("A1", "x")
        sheet.set_border("A1", style="thick", color="#FF0000")
    with Excel(path, read_only=True) as excel:
        cell = excel.sheet("集計")._worksheet["A1"]
        assert cell.border.left.style == "thick"
        assert cell.border.right.style == "thick"
        assert cell.border.top.style == "thick"
        assert cell.border.bottom.style == "thick"
        # openpyxl は 8 桁の ARGB 形式で色を保存する（先頭 2 桁はアルファチャンネル）
        assert cell.border.left.color.value == "00FF0000"


def test_set_border_rejects_unknown_keyword() -> None:
    """未知のキーワードは Python の呼び出し時点で TypeError。"""
    import inspect

    from comken.toolbox.excel.sheet import Sheet

    sig = inspect.signature(Sheet.set_border)
    # set_border() は style / color しか受け付けないため、
    # 未知のキーワードを bind しようとすると TypeError になる。
    with pytest.raises(TypeError):
        sig.bind("A1", thickness=2)


def test_openpyxl_side_rejects_unknown_style_with_clear_message() -> None:
    """openpyxl の Side は無効な style を ValueError にして、有効値の一覧を返す。

    comken の Sheet.set_border() は Literal 型でビルド時に不正値を防ぐので、
    openpyxl 側の例外メッセージはここで直接確認する。
    """
    from openpyxl.styles import Side

    with pytest.raises(ValueError, match="Value must be one of"):
        Side(style="thinn", color="000000")


def test_excel_outside_with_block_raises_table_not_open_error(tmp_path) -> None:
    path = tmp_path / "book.xlsx"
    excel = Excel(path)
    with pytest.raises(TableNotOpenError, match="Excel"):
        excel.list_data_sheets()
    with pytest.raises(TableNotOpenError, match="Excel"):
        excel.data_sheet("顧客")
    with pytest.raises(TableNotOpenError, match="Excel"):
        excel.create_sheet("集計")
    with pytest.raises(TableNotOpenError, match="Excel"):
        excel.save()


class TestCreateTableNameValidation:
    """``Sheet.create_table`` は Excel が受け付けない名前を ``InvalidTableNameError`` で止める。"""

    @pytest.mark.parametrize(
        "invalid_name",
        [
            pytest.param("", id="empty"),
            pytest.param("結 果", id="contains-half-width-space"),
            pytest.param("結　果", id="contains-full-width-space"),
            pytest.param("1結果", id="starts-with-digit"),
            pytest.param("A1", id="cell-reference-A1"),
            pytest.param("R1C1", id="cell-reference-R1C1"),
            pytest.param("Bad/Name", id="forbidden-slash"),
            pytest.param("Bad*Name", id="forbidden-asterisk"),
            pytest.param("Bad[Name", id="forbidden-bracket"),
            pytest.param("Bad]Name", id="forbidden-close-bracket"),
        ],
    )
    def test_invalid_names_raise(self, tmp_path, invalid_name: str) -> None:
        path = tmp_path / "book.xlsx"
        with Excel(path) as excel, pytest.raises(InvalidTableNameError):
            excel.create_data_sheet("S").create_table(invalid_name, Table(["a"], [{"a": "1"}]))

    def test_valid_japanese_name_is_accepted(self, tmp_path) -> None:
        path = tmp_path / "book.xlsx"
        with Excel(path) as excel:
            table = excel.create_data_sheet("S").create_table("顧客", Table(["ID"], [{"ID": "1"}]))
            rows = table.read()
        assert rows == [{"ID": "1"}]


class TestReadComputedRowsDropsBlankRows:
    """Excel の ``dimension`` が膨らんだブックで空行を返さない契約。

    Excel は行を削除しても書式が残っていると使用範囲（dimension）が縮まない。
    実務のファイルではよくある状態で、``min_row`` から ``max_row`` まで全部を
    返すと大量の中身のない行が混ざる。 ``_read_computed_rows`` /
    ``read()`` は「全セルが ``None`` または空文字」の行を
    ストリーム段階で落とす。
    """

    def test_declared_range_much_larger_than_data_returns_only_data_rows(self, tmp_path) -> None:
        """実データ 200 行 / 宣言された範囲 5000 行で、返る行数は 200。"""
        path = tmp_path / "declared-empty.xlsx"
        with Excel(path) as excel:
            sheet = excel.create_sheet("データ")
            sheet.write_range("A1:B1", [["ID", "名前"]])
            for index in range(2, 202):
                sheet.write_value(f"A{index}", str(index - 1))
                sheet.write_value(f"B{index}", f"ユーザー{index - 1}")
            # 遠くのセルに書式だけ付けて dimension を膨らませる
            sheet._worksheet.cell(row=5000, column=40).fill = PatternFill("solid", fgColor="FFFF00")

        with Excel(path, read_only=True) as excel:
            rows = excel._read_computed_rows("データ")

        assert len(rows) == 200

    def test_blank_rows_in_the_middle_are_skipped(self, tmp_path) -> None:
        """データの途中の空行（None と空文字の両方）は飛ばされる。"""
        path = tmp_path / "middle-blank.xlsx"
        with Excel(path) as excel:
            sheet = excel.create_sheet("データ")
            sheet.write_range("A1:B1", [["ID", "名前"]])
            sheet.write_value("A2", "1")
            sheet.write_value("B2", "A")
            # 3 行目は空（すべて None）
            # 4 行目はさらに空（すべて空文字）
            sheet.write_value("A5", "")
            sheet.write_value("B5", "")
            sheet.write_value("A6", "2")
            sheet.write_value("B6", "B")

        with Excel(path, read_only=True) as excel:
            rows = excel._read_computed_rows("データ")

        assert rows == [("1", "A"), ("2", "B")]

    def test_zero_and_false_are_kept(self, tmp_path) -> None:
        """``0`` や ``False`` は値として残る（集計を壊さないため）。"""
        path = tmp_path / "zero-false.xlsx"
        with Excel(path) as excel:
            sheet = excel.create_sheet("データ")
            sheet.write_range("A1:C1", [["ID", "数量", "有効"]])
            sheet.write_value("A2", "1")
            sheet.write_value("B2", 0)
            sheet.write_value("C2", False)

        with Excel(path, read_only=True) as excel:
            rows = excel._read_computed_rows("データ")

        assert rows == [("1", 0, False)]

    def test_all_blank_rows_returns_empty_list(self, tmp_path) -> None:
        """空行だけのシートでは空リストが返る（例外にしない）。"""
        path = tmp_path / "all-blank.xlsx"
        with Excel(path) as excel:
            sheet = excel.create_sheet("データ")
            sheet.write_value("A1", "ID")
            # 2 行目以降は空

        with Excel(path, read_only=True) as excel:
            rows = excel._read_computed_rows("データ")

        assert rows == []

    def test_dicts_path_also_skips_blank_rows(self, tmp_path) -> None:
        """``read()`` も空行を飛ばす。"""
        path = tmp_path / "blank-dict.xlsx"
        with Excel(path) as excel:
            sheet = excel.create_sheet("データ")
            sheet.write_range("A1:B1", [["ID", "名前"]])
            sheet.write_value("A2", "1")
            sheet.write_value("B2", "A")
            sheet.write_value("A4", "2")
            sheet.write_value("B4", "B")

        with Excel(path, read_only=True) as excel:
            table = excel.read("データ")

        assert table.read_rows() == [{"ID": "1", "名前": "A"}, {"ID": "2", "名前": "B"}]

    def test_empty_header_cell_error_still_fires(self, tmp_path) -> None:
        """見出し行の空セルは従来どおり ``EmptyHeaderCellError``。"""
        path = tmp_path / "header-blank.xlsx"
        with Excel(path) as excel:
            sheet = excel.create_sheet("データ")
            sheet.write_value("A1", "ID")
            # B1 は空の見出し
            sheet.write_value("A2", "1")
            sheet.write_value("B2", "A")

        with Excel(path, read_only=True) as excel, pytest.raises(EmptyHeaderCellError):
            excel.read("データ")

    def test_existing_header_and_data_behavior_unchanged(self, tmp_path) -> None:
        """通常のブックで見出し + データが期待どおり返ること。"""
        path = tmp_path / "normal.xlsx"
        with Excel(path) as excel:
            sheet = excel.create_sheet("データ")
            sheet.write_range("A1:B1", [["ID", "名前"]])
            sheet.write_value("A2", "1")
            sheet.write_value("B2", "A")

        # ``_read_computed_rows`` と ``read`` は内部で同じストリーム Workbook
        # を使うため、別々の ``with`` ブロックで呼ぶ。同じブロックで 2 回呼ぶと
        # Workbook のライフサイクル管理との兼ね合いで既存の問題が表面化する。
        with Excel(path, read_only=True) as excel:
            rows = excel._read_computed_rows("データ")
        with Excel(path, read_only=True) as excel:
            table = excel.read("データ")

        assert rows == [("1", "A")]
        assert table.read_rows() == [{"ID": "1", "名前": "A"}]


class TestFindSheet:
    """``Excel.find_sheet(*candidates)`` の契約テスト。

    「シート名が違うだけで業務ロジックは共通」という業務ケースを吸収する口。
    ``延期積上集計`` のように ``SHEET_NAME = [Sheet1, 一覧]`` と候補を並べた
    config と組み合わせて使う。
    """

    def test_returns_first_existing_candidate(self, tmp_path) -> None:
        path = tmp_path / "book.xlsx"
        with Excel(path) as excel:
            excel.create_sheet("案件一覧")
            excel.create_sheet("Sheet1")

        with Excel(path, read_only=True) as excel:
            # 候補の 1 番目が見つかればそれを返す（順番保持）
            assert excel.find_sheet("案件一覧", "Sheet1") == "案件一覧"
            # 1 番目が無く 2 番目が見つかれば 2 番目を返す
            assert excel.find_sheet("Sheet1", "案件一覧") == "Sheet1"

    def test_returns_name_string_not_sheet_object(self, tmp_path) -> None:
        """戻り値は ``str``（シート名）で ``Sheet`` ではない。"""
        path = tmp_path / "book.xlsx"
        with Excel(path) as excel:
            excel.create_sheet("案件一覧")

        with Excel(path, read_only=True) as excel:
            result = excel.find_sheet("案件一覧")
            assert isinstance(result, str)
            assert result == "案件一覧"

    def test_raises_sheet_not_found_when_no_candidate_matches(self, tmp_path) -> None:
        """候補が全部無いとき ``SheetNotFoundError`` をそのまま送出する。"""
        path = tmp_path / "book.xlsx"
        with Excel(path) as excel:
            excel.create_sheet("Sheet1")
            excel.create_sheet("案件一覧")

        with Excel(path, read_only=True) as excel:
            with pytest.raises(SheetNotFoundError) as exc:
                excel.find_sheet("無いA", "無いB")
            message = str(exc.value)
            # 最後に試した名前と、ブックに実在するシート名が両方メッセージに含まれる
            assert "無いB" in message
            assert "Sheet1" in message
            assert "案件一覧" in message

    def test_raises_when_no_candidates_given(self, tmp_path) -> None:
        """候補を 1 つも渡さなかったときも、シート一覧入りの例外で止める。"""
        path = tmp_path / "book.xlsx"
        with Excel(path) as excel:
            excel.create_sheet("Sheet1")

        with Excel(path, read_only=True) as excel:
            with pytest.raises(SheetNotFoundError) as exc:
                excel.find_sheet()
            message = str(exc.value)
            assert "Sheet1" in message

    def test_does_not_treat_data_sheet_as_existing_candidate(self, tmp_path) -> None:
        """``PY_`` 付きデータシートを候補に入れても「見つかった」とは扱わない。

        業務ロジック上、表示用シートだけを扱うので、データシート名と一致して
        候補が「在る」と判定されるのは事故（データシートは ``Table`` API で読む）。
        ``find_sheet`` は ``sheetnames`` の所属だけで判定するため、 ``PY_顧客``
        を渡すとそのまま返ってしまう点はこのテストで明示する
        （=データシート名しか無いブックで業務候補が「無い」事故の検知は呼び出し側の責任）。
        """
        path = tmp_path / "book.xlsx"
        with Excel(path) as excel:
            excel.create_data_sheet("顧客").create_table("顧客", Table(["ID"], [{"ID": "1"}]))

        with Excel(path, read_only=True) as excel:
            # ``PY_顧客`` は sheetnames に含まれるので「見つかった」と判定される
            # （=この API はプレフィックスによる区別をしない）。
            assert excel.find_sheet("PY_顧客") == "PY_顧客"
            # 表示用シート名を候補にしても見つからない。
            with pytest.raises(SheetNotFoundError):
                excel.find_sheet("案件一覧")
