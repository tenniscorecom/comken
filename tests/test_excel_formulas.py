"""Excel の数式セル読み書きまわりの契約テスト。"""

import re
import tempfile
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from comken.core.table import Table
from comken.exceptions import TableColumnMismatchError, TableFormulaOverwriteError
from comken.toolbox.excel import Excel


def _book_with_cached_formulas(
    path: Path,
    *,
    sheet_name: str,
    formulas_with_values: dict[str, tuple[str, str]],
) -> None:
    """数式セルにキャッシュ値（計算結果）を埋めた xlsx を作る。

    openpyxl は保存時に再計算しないので、保存後の zip 内
    ``xl/worksheets/sheet1.xml`` を直接編集して ``<c><f>...</f><v>...</v></c>``
    の ``<v>`` に計算結果を入れる。Excel が開いたときは「キャッシュ済み」と
    みなされ、``Excel._cached_range`` は再計算しないで値を返す。
    """
    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:
        raise AssertionError("Workbook に active sheet がない")
    sheet.title = sheet_name
    for coord, formula_value in formulas_with_values.items():
        sheet[coord] = formula_value[0]
    workbook.save(path)
    # zip 内のシート XML に <v> を埋め込んでから再圧縮する。
    with tempfile.TemporaryDirectory() as tmp_str:
        tmp_path = Path(tmp_str)
        with zipfile.ZipFile(path) as archive:
            archive.extractall(tmp_path)
        sheet_xml_path = tmp_path / "xl" / "worksheets" / "sheet1.xml"
        text = sheet_xml_path.read_text(encoding="utf-8")
        for coord, formula_value in formulas_with_values.items():
            cached = formula_value[1]
            # openpyxl は lxml が入っていると ``<v></v>``、入っていないと
            # ``<v />`` を書く。どちらでも拾えるようにしておかないと、
            # 手元では通って CI（lxml なし）だけ落ちる
            pattern = re.compile(
                r'(<c r="' + re.escape(coord) + r'"[^/>]*>)'
                r"(<f[^<]*</f>)"
                r"(<v[^<]*</v>|<v\s*/>)"
                r"(</c>)"
            )
            replacement = r"\1\2<v>" + cached + r"</v>\4"
            text, count = pattern.subn(replacement, text)
            if count == 0:
                raise AssertionError(f"数式セルが見つかりません: {coord}")
        sheet_xml_path.write_text(text, encoding="utf-8")
        path.unlink()
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
            for file in sorted(tmp_path.rglob("*")):
                if file.is_file():
                    archive.write(file, file.relative_to(tmp_path).as_posix())


def test_read_value_returns_cached_formula_result(tmp_path: Path) -> None:
    """read_value() は数式セルのとき ``=...`` ではなく計算結果を返す。"""
    path = tmp_path / "cached.xlsx"
    _book_with_cached_formulas(
        path,
        sheet_name="集計",
        formulas_with_values={"A2": ("=A1*2", "4")},
    )
    with Excel(path) as excel:
        sheet = excel.sheet("集計")
        assert sheet.read_value("A2") == 4


def test_read_formula_returns_formula_string(tmp_path: Path) -> None:
    """read_formula() はセルの数式本体をそのまま返す。"""
    path = tmp_path / "formula.xlsx"
    _book_with_cached_formulas(
        path,
        sheet_name="集計",
        formulas_with_values={"A2": ("=A1*2", "4")},
    )
    with Excel(path) as excel:
        sheet = excel.sheet("集計")
        assert sheet.read_formula("A2") == "=A1*2"


def test_read_value_returns_raw_value_when_no_formula(tmp_path: Path) -> None:
    """数式が無いセルの read_value() は従来どおりの値を返す。"""
    path = tmp_path / "plain.xlsx"
    with Excel(path) as excel:
        sheet = excel.create_sheet("集計")
        sheet.write_value("A1", "hello")
        sheet.write_value("B1", 42)
        assert sheet.read_value("A1") == "hello"
        assert sheet.read_value("B1") == 42


def test_read_formula_returns_empty_when_not_a_formula(tmp_path: Path) -> None:
    """数式でないセルの read_formula() は空文字を返す。"""
    path = tmp_path / "plain.xlsx"
    with Excel(path) as excel:
        sheet = excel.create_sheet("集計")
        sheet.write_value("A1", "hello")
        assert sheet.read_formula("A1") == ""


def test_replace_with_formula_raises_default(tmp_path: Path) -> None:
    """人が入れた数式を含むテーブルへの replace() は既定で例外を出す。"""
    path = tmp_path / "formula-table.xlsx"
    with Excel(path) as excel:
        table = excel.create_data_sheet("顧客").create_table(
            "顧客", Table(["ID", "合計"], [{"ID": "001", "合計": 100}])
        )
        # 人が「合計」列へ入れた数式
        table._worksheet["B2"] = "=SUM(B2:B3)"

        with pytest.raises(TableFormulaOverwriteError) as exc_info:
            table.replace([{"ID": "002", "合計": 200}])

        assert "B2" in str(exc_info.value)
        # 数式セルはそのまま残っている
        assert table._worksheet["B2"].value == "=SUM(B2:B3)"


def test_replace_with_formula_overwrites_when_allowed(tmp_path: Path) -> None:
    """allow_formula_overwrite=True のときは従来どおり値で上書きする。"""
    path = tmp_path / "formula-overwrite.xlsx"
    with Excel(path) as excel:
        table = excel.create_data_sheet("顧客").create_table(
            "顧客", Table(["ID", "合計"], [{"ID": "001", "合計": 100}])
        )
        table._worksheet["B2"] = "=SUM(B2:B3)"
        table.replace(
            [{"ID": "002", "合計": 200}],
            allow_formula_overwrite=True,
        )
        assert table._worksheet["B2"].value == 200


def test_replace_without_formula_succeeds(tmp_path: Path) -> None:
    """数式が無いテーブルへの replace() は従来どおり通る。"""
    path = tmp_path / "no-formula.xlsx"
    with Excel(path) as excel:
        table = excel.create_data_sheet("顧客").create_table("顧客", Table(["ID"], [{"ID": "001"}]))
        table.replace([{"ID": "002"}])
        assert table.read() == [{"ID": "002"}]


def test_replace_only_checks_data_rows_not_header(tmp_path: Path) -> None:
    """見出し行の先頭が ``=`` だと、その文字列が見出し名として扱われる。

    見出しが ``=壊れ見出し`` に変わっているため、渡された Table の列 ``ID`` が
    既存の見出しと一致せず ``TableColumnMismatchError`` になる。
    （数式検出はデータ部のみを見ることが、新しい動作でも保たれている。）
    """
    path = tmp_path / "header-formula.xlsx"
    with Excel(path) as excel:
        # 見出し行に ``=`` 始まりの文字列を入れる異常系。通常起こらないが、
        # replace() の検査範囲がデータ部に限定されていることを確認する。
        excel.create_data_sheet("顧客").create_table("顧客", Table(["ID"], [{"ID": "001"}]))
        excel._workbook["PY_顧客"]["A1"] = "=壊れ見出し"
        # 既存の見出しと渡された Table の列名が一致しないため例外。
        with pytest.raises(TableColumnMismatchError) as exc_info:
            excel.data_sheet("顧客").table().replace([{"ID": "002"}])
        assert "ID" in str(exc_info.value)


class TestFormulaColumnPreservation:
    """数式列を含むテーブルへの ``replace()`` / ``append()`` の振る舞い。"""

    def _make_order_table(self, path: Path) -> None:
        """``| 顧客ID | 数量 | 単価 | 金額 |`` のテーブルを作る（D 列が数式）。"""
        with Excel(path) as excel:
            sheet = excel.create_data_sheet("注文")
            table = sheet.create_table(
                "注文",
                Table(
                    ["顧客ID", "数量", "単価", "金額"],
                    [
                        {"顧客ID": "A", "数量": 2, "単価": 100, "金額": 200},
                        {"顧客ID": "B", "数量": 3, "単価": 150, "金額": 450},
                    ],
                ),
            )
            # 人が「金額」列へ入れた相対参照の数式（B2*C2 の行方向シフト）
            table._worksheet["D2"] = "=B2*C2"
            table._worksheet["D3"] = "=B3*C3"

    def test_replace_preserves_formula_column(self, tmp_path: Path) -> None:
        """数式列を含まない Table で ``replace()`` すると、数式列が数式のまま残る。"""
        path = tmp_path / "formula-keep.xlsx"
        self._make_order_table(path)
        with Excel(path) as excel:
            table = excel.data_sheet("注文").table()
            table.replace(
                Table(
                    ["顧客ID", "数量", "単価"],
                    [
                        {"顧客ID": "C", "数量": 5, "単価": 80},
                        {"顧客ID": "D", "数量": 1, "単価": 1000},
                    ],
                )
            )
            # データ部が新しい値で置き換わっている
            assert table._worksheet["B2"].value == 5
            assert table._worksheet["C2"].value == 80
            # 数式列（D）は数式のまま残る
            assert table._worksheet["D2"].value == "=B2*C2"
            assert table._worksheet["D3"].value == "=B3*C3"

    def test_replace_extends_formula_with_translated_references(self, tmp_path: Path) -> None:
        """行が増えたとき、増えた行の数式列に相対参照が正しくずれた数式が入る。"""
        path = tmp_path / "formula-extend.xlsx"
        self._make_order_table(path)
        with Excel(path) as excel:
            table = excel.data_sheet("注文").table()
            # 既存 2 行 → 5 行へ拡張（D 列は渡さない → 保持される）
            table.replace(
                Table(
                    ["顧客ID", "数量", "単価"],
                    [
                        {"顧客ID": "A", "数量": 2, "単価": 100},
                        {"顧客ID": "B", "数量": 3, "単価": 150},
                        {"顧客ID": "C", "数量": 4, "単価": 200},
                        {"顧客ID": "D", "数量": 5, "単価": 250},
                        {"顧客ID": "E", "数量": 6, "単価": 300},
                    ],
                )
            )
            # 既存行の数式はそのまま
            assert table._worksheet["D2"].value == "=B2*C2"
            assert table._worksheet["D3"].value == "=B3*C3"
            # 増えた行の数式は Translator で下方向へコピーされている
            assert table._worksheet["D4"].value == "=B4*C4"
            assert table._worksheet["D5"].value == "=B5*C5"
            assert table._worksheet["D6"].value == "=B6*C6"

    def test_replace_shrinks_clears_removed_formula_cells(self, tmp_path: Path) -> None:
        """行が減ったとき、余った行の数式セルが消える。"""
        path = tmp_path / "formula-shrink.xlsx"
        self._make_order_table(path)
        with Excel(path) as excel:
            table = excel.data_sheet("注文").table()
            # 既存 2 行 → 1 行に減らす
            table.replace(
                Table(
                    ["顧客ID", "数量", "単価"],
                    [{"顧客ID": "A", "数量": 2, "単価": 100}],
                )
            )
            assert table._worksheet["B2"].value == 2
            assert table._worksheet["C2"].value == 100
            assert table._worksheet["D2"].value == "=B2*C2"
            # 余った行の数式セルは消える
            assert table._worksheet["B3"].value is None
            assert table._worksheet["C3"].value is None
            assert table._worksheet["D3"].value is None

    def test_replace_with_formula_column_in_passed_raises(self, tmp_path: Path) -> None:
        """数式列を含む Table で ``replace()`` すると ``TableFormulaOverwriteError``。"""
        path = tmp_path / "formula-included.xlsx"
        self._make_order_table(path)
        with Excel(path) as excel:
            table = excel.data_sheet("注文").table()
            with pytest.raises(TableFormulaOverwriteError) as exc_info:
                table.replace(
                    Table(
                        ["顧客ID", "数量", "単価", "金額"],
                        [{"顧客ID": "C", "数量": 5, "単価": 80, "金額": 400}],
                    )
                )
            # 数式セルの座標が示される
            assert "D2" in str(exc_info.value)
            # 数式はそのまま残っている（上書きされていない）
            assert table._worksheet["D2"].value == "=B2*C2"

    def test_replace_with_allow_overwrite_replaces_formula_cells(self, tmp_path: Path) -> None:
        """``allow_formula_overwrite=True`` なら値で上書きされる。"""
        path = tmp_path / "formula-overwrite-allowed.xlsx"
        self._make_order_table(path)
        with Excel(path) as excel:
            table = excel.data_sheet("注文").table()
            table.replace(
                Table(
                    ["顧客ID", "数量", "単価", "金額"],
                    [{"顧客ID": "C", "数量": 5, "単価": 80, "金額": 9999}],
                ),
                allow_formula_overwrite=True,
            )
            assert table._worksheet["D2"].value == 9999

    def test_replace_with_unknown_column_raises_mismatch(self, tmp_path: Path) -> None:
        """既存の見出しに無い列名を渡すと例外。"""
        path = tmp_path / "unknown-column.xlsx"
        self._make_order_table(path)
        with Excel(path) as excel:
            table = excel.data_sheet("注文").table()
            with pytest.raises(TableColumnMismatchError) as exc_info:
                table.replace(
                    Table(
                        ["顧客ID", "数量", "単価", "謎の列"],
                        [{"顧客ID": "A", "数量": 1, "単価": 10, "謎の列": "x"}],
                    )
                )
            assert "謎の列" in str(exc_info.value)
            # 数式列はそのまま残っている
            assert table._worksheet["D2"].value == "=B2*C2"

    def test_replace_with_omitted_non_formula_column_raises_mismatch(self, tmp_path: Path) -> None:
        """既存にある非数式列を Table から省くと、データ欠落を防ぐため例外。"""
        path = tmp_path / "omit-non-formula.xlsx"
        self._make_order_table(path)
        with Excel(path) as excel:
            table = excel.data_sheet("注文").table()
            with pytest.raises(TableColumnMismatchError) as exc_info:
                # 「数量」を省いて渡すと、非数式列を落とす形になる
                table.replace(
                    Table(
                        ["顧客ID", "単価", "金額"],
                        [{"顧客ID": "A", "単価": 100, "金額": 200}],
                    )
                )
            assert "数量" in str(exc_info.value)

    def test_append_preserves_formula_column(self, tmp_path: Path) -> None:
        """``append()`` で行を足しても数式列が保持される。"""
        path = tmp_path / "formula-append.xlsx"
        self._make_order_table(path)
        with Excel(path) as excel:
            table = excel.data_sheet("注文").table()
            table.append(
                Table(
                    ["顧客ID", "数量", "単価"],
                    [{"顧客ID": "C", "数量": 4, "単価": 200}],
                )
            )
            # 既存行の数式はそのまま
            assert table._worksheet["D2"].value == "=B2*C2"
            assert table._worksheet["D3"].value == "=B3*C3"
            # 追加行にも数式が入る（相対参照がずれる）
            assert table._worksheet["D4"].value == "=B4*C4"

    def test_append_with_formula_in_addition_raises(self, tmp_path: Path) -> None:
        """``append()`` で数式列の値を渡すと ``TableFormulaOverwriteError``。"""
        path = tmp_path / "formula-append-error.xlsx"
        self._make_order_table(path)
        with Excel(path) as excel:
            table = excel.data_sheet("注文").table()
            with pytest.raises(TableFormulaOverwriteError):
                table.append(
                    Table(
                        ["顧客ID", "数量", "単価", "金額"],
                        [{"顧客ID": "C", "数量": 5, "単価": 80, "金額": 400}],
                    )
                )

    def test_replace_table_without_formulas_unchanged(self, tmp_path: Path) -> None:
        """数式が無いテーブルは従来どおり動く。"""
        path = tmp_path / "no-formula-preserve.xlsx"
        with Excel(path) as excel:
            table = excel.create_data_sheet("顧客").create_table(
                "顧客", Table(["ID", "名前"], [{"ID": "001", "名前": "旧"}])
            )
            table.replace(Table(["ID", "名前"], [{"ID": "002", "名前": "新"}]))
            assert table._worksheet["A2"].value == "002"
            assert table._worksheet["B2"].value == "新"
